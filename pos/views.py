
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count, F, Q, ExpressionWrapper, DecimalField
from django.db import transaction
from django.core.paginator import Paginator
from django.views.decorators.http import require_POST, require_GET
from django.core.serializers.json import DjangoJSONEncoder
from django.utils import timezone
from datetime import datetime, timedelta
from decimal import Decimal
import json
import csv
from django.utils import timezone
from django.urls import reverse
from django.urls.base import reverse
# Add these imports at the top
from django.db.models.functions import ExtractHour, TruncHour, TruncMonth
from django.core.exceptions import ObjectDoesNotExist
from django.shortcuts import reverse
from io import StringIO
from django.db.models.functions import TruncDate, TruncMonth
from django.contrib import messages
from django.template.loader import render_to_string
from django.http import HttpResponseBadRequest
import pandas as pd
# Add these imports at the top of views.py
from django.db.models import Case, When, Value, Sum, Count, Max, Min, Avg, Q, F
from django.db.models.functions import ExtractHour, TruncHour, TruncMonth, TruncDate
from django.db.models import DecimalField
from django.contrib.auth.models import User  # Add this import
from django.db import OperationalError  # Add this import
import time 


# pos/views.py - Add these imports at the top with other imports
from django.contrib.auth.forms import UserCreationForm
from django import forms
from .models import (
    Role, UserProfile, UserActivityLog, PermissionGroup,
    Company, Supplier, Category, Product, Batch, Customer,
    Sale, SaleItem, Purchase, PurchaseItem, PendingPurchase,
    StockJournal, StockJournalItem, Expense, SupplierPayment,
    Discount, PendingSale, CompanyPrice, Employee, SupplierReturn,
    Receipt, CustomerPayment, CreditSaleItem, CreditPayment
)
from .forms import (
    CustomUserCreationForm, UserProfileForm, CustomUserEditForm,
    PasswordResetAdminForm, UserFilterForm, RoleForm,
    RoleFilterForm, BulkUserImportForm
)

from .models import (
    Product, Category, Customer, Supplier,
    Sale, SaleItem, Purchase, PurchaseItem,
    Expense, Batch, Discount, Company, CompanyPrice,
    SupplierPayment, PendingSale, PendingPurchase, CustomerPayment
)
from .forms import (
    ProductForm, CustomerForm, SupplierForm, PurchaseForm, 
    ExpenseForm, BatchForm, DiscountForm, CompanyPriceForm,
    BulkUploadForm, PurchaseReturnForm,CompanyForm, SupplierPaymentForm, ExpenseFormSet, BulkExpenseForm, StockJournalItemForm, 
    StockJournalItemFormSet, StockJournalForm, StockJournalItem
)


from decimal import Decimal, ROUND_HALF_UP

def round_to_nearest(value):
    """Round decimal to nearest whole number"""
    if isinstance(value, Decimal):
        return value.quantize(Decimal('1'), rounding=ROUND_HALF_UP)
    try:
        decimal_value = Decimal(str(value))
        return decimal_value.quantize(Decimal('1'), rounding=ROUND_HALF_UP)
    except:
        return Decimal('0')

def round_subtotal(subtotal):
    """Round subtotal to nearest whole number"""
    return round_to_nearest(subtotal)

def calculate_rounded_total(items):
    """Calculate total by rounding each item's total then summing"""
    rounded_total = Decimal('0')
    for item in items:
        # Calculate item total (price × quantity)
        item_total = Decimal(str(item.get('price', 0))) * Decimal(str(item.get('quantity', 0)))
        # Round each item total
        rounded_item_total = round_to_nearest(item_total)
        rounded_total += rounded_item_total
    return rounded_total
# ============== Decimal Formatting Utilities ==============

def format_decimal_4(value):
    """Format decimal to 4 decimal places consistently"""
    if value is None:
        return Decimal('0.0000')
    
    try:
        if isinstance(value, (Decimal, float, int)):
            decimal_value = Decimal(str(value))
            return decimal_value.quantize(Decimal('0.0001'))
        return Decimal(str(value)).quantize(Decimal('0.0001'))
    except:
        return Decimal('0.0000')

def format_currency(value):
    """Format as currency with 4 decimal places"""
    formatted = format_decimal_4(value)
    return f"KES {formatted}"

def get_formatted_sale_data(sale):
    """Helper to format all sale decimal values"""
    return {
        'id': sale.id,
        'sale_number': sale.sale_number,
        'date': sale.date,
        'customer': sale.customer.name if sale.customer else 'Walk-in Customer',
        'subtotal': format_currency(sale.subtotal),
        'tax': format_currency(sale.tax),
        'discount_amount': format_currency(sale.discount_amount),
        'total': format_currency(sale.total),
        'amount_paid': format_currency(sale.amount_paid),
        'balance': format_currency(sale.balance),
        'payment_method': sale.get_payment_method_display(),
        'is_credit': sale.is_credit,
        'items_count': sale.items.count(),
        'raw_total': sale.total,  # Keep raw for calculations if needed
    }

def get_formatted_sale_items(sale):
    """Helper to format all sale item decimal values"""
    formatted_items = []
    for item in sale.items.all():
        formatted_items.append({
            'product_name': item.product.name if item.product else 'Unknown',
            'quantity': item.quantity,
            'price': format_currency(item.price),
            'total': format_currency(item.total),
            'discount_amount': format_currency(item.discount_amount),
            'raw_price': item.price,  # Keep raw for calculations
            'raw_total': item.total,
        })
    return formatted_items

def get_formatted_product_data(product):
    """Helper to format product decimal values"""
    return {
        'id': product.id,
        'name': product.name,
        'barcode': product.barcode,
        'category': product.category.name if product.category else '',
        'purchase_price': format_currency(product.purchase_price),
        'selling_price': format_currency(product.selling_price),
        'least_selling_price': format_currency(product.least_selling_price),
        'wholesale_price': format_currency(product.wholesale_price),
        'quantity': product.quantity,
        'stock_value': format_currency(product.quantity * product.purchase_price),
        'profit_margin': f"{product.profit_margin:.2f}%" if hasattr(product, 'profit_margin') else '0.00%',
        'is_active': product.is_active,
        'raw_purchase_price': product.purchase_price,  # Keep raw values
        'raw_selling_price': product.selling_price,
    }
# ============== Dashboard Views ==============
@login_required
def dashboard(request):
    # Today's sales data
    today = timezone.now().date()
    today_sales = Sale.objects.filter(
        date__date=today,
        is_completed=True
    )
    
    # Monthly sales data
    monthly_sales = Sale.objects.filter(
        date__month=timezone.now().month,
        date__year=timezone.now().year,
        is_completed=True
    )
    
    # Calculate totals WITH formatting
    today_total = today_sales.aggregate(total=Sum('total'))['total'] or Decimal('0.0000')
    monthly_total = monthly_sales.aggregate(total=Sum('total'))['total'] or Decimal('0.0000')
    
    # Outstanding credit
    outstanding_credit = Sale.objects.filter(
        is_credit=True,
        balance__gt=0
    ).aggregate(total=Sum('balance'))['total'] or Decimal('0.0000')
    
    # Low stock items
    low_stock_count = Product.objects.filter(
        quantity__lte=F('reorder_level'),
        quantity__gt=0,
        is_active=True
    ).count()
    
    # Recent sales - use formatting helper
    recent_sales = Sale.objects.filter(
        is_completed=True
    ).order_by('-date')[:10]
    
    # Format recent sales
    formatted_recent_sales = []
    for sale in recent_sales:
        formatted_recent_sales.append({
            'id': sale.id,
            'sale_number': sale.sale_number,
            'date': sale.date,
            'customer_name': sale.customer.name if sale.customer else 'Walk-in',
            'total': format_currency(sale.total),
            'payment_method': sale.get_payment_method_display(),
            'is_completed': sale.is_completed,
            'raw_total': sale.total,  # Keep original for calculations if needed
        })
    
    # Fast moving products (last 7 days)
    seven_days_ago = timezone.now() - timedelta(days=7)
    fast_moving = SaleItem.objects.filter(
        sale__date__gte=seven_days_ago,
        sale__is_completed=True
    ).values('product__name').annotate(
        total_sold=Sum('quantity'),
        total_revenue=Sum('total')
    ).order_by('-total_sold')[:5]
    
    formatted_fast_moving = []
    for item in fast_moving:
        formatted_fast_moving.append({
            'name': item['product__name'],
            'total_sold': item['total_sold'],
            'total_revenue': format_currency(item['total_revenue'] or Decimal('0.0000')),
        })
    
    # Slow moving products (no sales in 30 days)
    thirty_days_ago = timezone.now() - timedelta(days=30)
    slow_moving = Product.objects.filter(
        is_active=True,
        quantity__gt=0
    ).exclude(
        saleitem__sale__date__gte=thirty_days_ago
    )[:5]
    
    formatted_slow_moving = []
    for product in slow_moving:
        last_sale = SaleItem.objects.filter(
            product=product
        ).order_by('-sale__date').first()
        
        days_since = None
        if last_sale:
            days_since = timezone.now().date() - last_sale.sale.date.date()
        
        formatted_slow_moving.append({
            'name': product.name,
            'quantity': product.quantity,
            'last_sale_date': last_sale.sale.date if last_sale else None,
            'days_since_sale': days_since,
        })
    
    # Sales trend data for chart (last 7 days)
    trend_dates = []
    trend_data = []
    for i in range(6, -1, -1):
        date = today - timedelta(days=i)
        daily_sales = Sale.objects.filter(
            date__date=date,
            is_completed=True
        ).aggregate(total=Sum('total'))['total'] or Decimal('0.0000')
        
        trend_dates.append(date.strftime('%a'))
        trend_data.append(float(daily_sales))
    
    # Payment methods data
    payment_methods_data = []
    for method_code, method_name in Sale.PAYMENT_METHODS:
        method_total = Sale.objects.filter(
            date__date=today,
            is_completed=True,
            payment_method=method_code
        ).aggregate(total=Sum('total'))['total'] or Decimal('0.0000')
        
        payment_methods_data.append({
            'name': method_name,
            'total': float(method_total),
        })
    
    # Today vs yesterday comparison
    yesterday = today - timedelta(days=1)
    yesterday_sales = Sale.objects.filter(
        date__date=yesterday,
        is_completed=True
    ).aggregate(total=Sum('total'))['total'] or Decimal('0.0000')
    
    if yesterday_sales > 0:
        today_vs_yesterday = ((today_total - yesterday_sales) / yesterday_sales * 100)
    else:
        today_vs_yesterday = 100 if today_total > 0 else 0
    
    # Month vs last month comparison
    last_month = timezone.now().replace(day=1) - timedelta(days=1)
    last_month_sales = Sale.objects.filter(
        date__year=last_month.year,
        date__month=last_month.month,
        is_completed=True
    ).aggregate(total=Sum('total'))['total'] or Decimal('0.0000')
    
    if last_month_sales > 0:
        monthly_vs_last = ((monthly_total - last_month_sales) / last_month_sales * 100)
    else:
        monthly_vs_last = 100 if monthly_total > 0 else 0
    
    context = {
        # Card data - formatted
        'today_sales': format_currency(today_total),
        'monthly_sales': format_currency(monthly_total),
        'outstanding_credit': format_currency(outstanding_credit),
        'low_stock_count': low_stock_count,
        
        # Comparison data
        'today_vs_yesterday': today_vs_yesterday,
        'today_vs_yesterday_abs': abs(today_vs_yesterday),
        'monthly_vs_last': monthly_vs_last,
        'monthly_vs_last_abs': abs(monthly_vs_last),
        
        # Tables - formatted
        'recent_sales': formatted_recent_sales,
        'fast_moving_products': formatted_fast_moving,
        'slow_moving_products': formatted_slow_moving,
        
        # Chart data
        'sales_trend_labels': json.dumps(trend_dates),
        'sales_trend_data': json.dumps(trend_data),
        'payment_method_labels': json.dumps([m['name'] for m in payment_methods_data]),
        'payment_method_data': json.dumps([m['total'] for m in payment_methods_data]),
        'payment_method_colors': json.dumps(['#3B82F6', '#10B981', '#F59E0B', '#EF4444', '#8B5CF6']),
        'payment_method_colors_hover': json.dumps(['#2563EB', '#059669', '#D97706', '#DC2626', '#7C3AED']),
        
        # For reference
        'raw_today_sales': today_total,
        'raw_monthly_sales': monthly_total,
    }
    return render(request, 'pos/dashboard.html', context)

# ============== POS & Sales Views ==============
@login_required
def pos(request):
    company = Company.objects.first()
    products = Product.objects.filter(is_active=True).order_by('name')
    categories = Category.objects.all()
    customers = Customer.objects.all().order_by('name')
    
    # Get currently valid discounts
    today = timezone.now().date()
    discounts = Discount.objects.filter(
        is_active=True,
        start_date__lte=today,
        end_date__gte=today
    )
    
    # Generate temporary display number
    today_sales_count = Sale.objects.filter(
        date__date=today
    ).count()
    display_number = f"SALE-{timezone.now().strftime('%Y%m%d')}-{today_sales_count + 1:04d}"
    
    # Check if we're editing an existing sale
    edit_sale_data = None
    edit_sale_id = None
    is_editing = False
    
    if 'edit_sale_data' in request.session:
        edit_sale_data = request.session.pop('edit_sale_data')
        edit_sale_id = request.session.pop('edit_sale_id', None)
        is_editing = True
        
        # Update display number to show we're editing
        if edit_sale_data and 'sale_number' in edit_sale_data:
            display_number = f"EDIT-{edit_sale_data['sale_number']}"
    
    context = {
        'products': products,
        'categories': categories,
        'customers': customers,
        'discounts': discounts,
        'display_number': display_number,
        'company': company,
        'edit_sale_data': json.dumps(edit_sale_data) if edit_sale_data else 'null',
        'edit_sale_id': edit_sale_id,
        'is_editing': is_editing
    }
    return render(request, 'pos/pos.html', context)

# views.py - Update the process_sale view with automatic batch splitting

# views.py - Update the process_sale view with purchase item tracking

@require_POST
def process_sale(request):
    max_retries = 3
    retry_delay = 0.1
    
    for attempt in range(max_retries):
        try:
            with transaction.atomic():
                data = json.loads(request.POST.get('sale_data'))
                pending_sale_id = request.POST.get('pending_sale_id')
                edit_sale_id = request.POST.get('edit_sale_id')
                
                company = Company.objects.first() or Company.objects.create(name="Default Company")
                
                # Round quantity to 5 decimal places
                def round_quantity(qty):
                    return Decimal(str(qty)).quantize(Decimal('0.00001'))
                
                # CHECK IF THIS IS EDITING AN EXISTING SALE
                if edit_sale_id:
                    original_sale = get_object_or_404(Sale, pk=edit_sale_id, is_completed=True)
                    
                    # First, restore original stock for all items in the original sale
                    for item in original_sale.items.all():
                        if item.batch:
                            batch = Batch.objects.select_for_update().get(id=item.batch.id)
                            batch.quantity += item.quantity
                            batch.save()
                            
                            # Restore purchase item remaining quantity
                            if item.purchase_item:
                                purchase_item = item.purchase_item
                                purchase_item.remaining_quantity += item.quantity
                                purchase_item.is_fully_sold = False
                                purchase_item.save()
                            
                            # Update product total quantity
                            product = batch.product
                            product.quantity += item.quantity
                            product.save()
                    
                    # Delete the original sale items
                    original_sale.items.all().delete()
                    
                    # Update the original sale with new data
                    original_sale.subtotal = Decimal(str(data['subtotal']))
                    original_sale.discount_amount = Decimal(str(data.get('discount_amount', 0)))
                    original_sale.discount_percent = Decimal(str(data.get('discount_percent', 0)))
                    original_sale.total = Decimal(str(data['total']))
                    original_sale.payment_method = data['payment_method']
                    original_sale.amount_paid = Decimal(str(data['amount_paid']))
                    original_sale.balance = max(Decimal(str(data['total'])) - Decimal(str(data['amount_paid'])), Decimal(0))
                    original_sale.is_credit = data.get('is_credit', False)
                    original_sale.is_completed = True
                    original_sale.updated_at = timezone.now()
                    
                    # Handle payment details
                    payment_details = data.get('payment_details', {})
                    original_sale.mpesa_amount = Decimal(str(payment_details.get('mpesa', 0)))
                    original_sale.cash_amount = Decimal(str(payment_details.get('cash', 0)))
                    original_sale.card_amount = Decimal(str(payment_details.get('card', 0)))
                    original_sale.cheque_amount = Decimal(str(payment_details.get('cheque', 0)))
                    original_sale.mpesa_code = payment_details.get('mpesa_code', '')
                    original_sale.cheque_number = payment_details.get('cheque_number', '')
                    
                    # Update customer if changed
                    if data.get('customer_id'):
                        original_sale.customer_id = data['customer_id']
                    else:
                        original_sale.customer = None
                    
                    original_sale.save()
                    sale = original_sale
                    
                else:
                    # CREATE NEW SALE
                    sale = Sale(
                        customer_id=data.get('customer_id'),
                        user=request.user,
                        sale_type=data.get('sale_type', 'retail'),
                        subtotal=Decimal(str(data['subtotal'])),
                        discount_amount=Decimal(str(data.get('discount_amount', 0))),
                        discount_percent=Decimal(str(data.get('discount_percent', 0))),
                        total=Decimal(str(data['total'])),
                        payment_method=data['payment_method'],
                        amount_paid=Decimal(str(data['amount_paid'])),
                        balance=max(Decimal(str(data['total'])) - Decimal(str(data['amount_paid'])), Decimal(0)),
                        is_credit=data.get('is_credit', False),
                        is_completed=True
                    )
                    
                    # Handle payment details
                    payment_details = data.get('payment_details', {})
                    sale.mpesa_amount = Decimal(str(payment_details.get('mpesa', 0)))
                    sale.cash_amount = Decimal(str(payment_details.get('cash', 0)))
                    sale.card_amount = Decimal(str(payment_details.get('card', 0)))
                    sale.cheque_amount = Decimal(str(payment_details.get('cheque', 0)))
                    sale.mpesa_code = payment_details.get('mpesa_code', '')
                    sale.cheque_number = payment_details.get('cheque_number', '')
                    
                    sale.save()
                
                # Create sale items and update inventory
                for item in data['items']:
                    product = Product.objects.select_for_update().get(id=item['id'])
                    batch_id = item.get('batch_id')
                    
                    # Round quantity to 5 decimal places
                    quantity = round_quantity(item['quantity'])
                    
                    # Validate quantity
                    if quantity <= 0:
                        raise ValueError(f"Invalid quantity for product {product.name}: {quantity}")
                    
                    # Handle batch selection
                    if batch_id:
                        # Use the specific batch selected
                        try:
                            batch = Batch.objects.select_for_update().get(id=batch_id, product=product)
                        except Batch.DoesNotExist:
                            raise ValueError(f"Selected batch not found for product {product.name}")
                        
                        # Get the purchase item linked to this batch
                        purchase_item = None
                        if hasattr(batch, 'purchase_item_source'):
                            purchase_item = batch.purchase_item_source
                        
                        # Check if enough stock in this batch
                        if batch.quantity < quantity:
                            # Try to find other batches with stock
                            other_batches = Batch.objects.filter(
                                product=product, 
                                quantity__gt=0,
                                is_active=True
                            ).exclude(id=batch_id).order_by('expiry_date', 'id')
                            
                            if other_batches.exists():
                                # We have other batches, but user selected specific batch
                                # We'll use the selected batch first, then others
                                batches_to_use = [batch] + list(other_batches)
                                remaining_qty = quantity
                                
                                for use_batch in batches_to_use:
                                    if remaining_qty <= 0:
                                        break
                                    
                                    take_from_batch = min(use_batch.quantity, remaining_qty)
                                    
                                    # Get purchase item for this batch
                                    use_purchase_item = None
                                    if hasattr(use_batch, 'purchase_item_source'):
                                        use_purchase_item = use_batch.purchase_item_source
                                    
                                    # Create sale item for this batch portion
                                    sale_item = SaleItem.objects.create(
                                        sale=sale,
                                        product=product,
                                        batch=use_batch,
                                        purchase_item=use_purchase_item,
                                        quantity=take_from_batch,
                                        price=Decimal(str(item['price'])),
                                        purchase_price=use_batch.purchase_price,
                                        discount_amount=Decimal(str(item.get('discount_amount', 0))),
                                        discount_percent=Decimal(str(item.get('discount_percent', 0))),
                                        total=round_to_nearest(take_from_batch * Decimal(str(item['price'])))
                                    )
                                    
                                    # Update batch stock
                                    use_batch.quantity -= take_from_batch
                                    if use_batch.quantity < 0:
                                        use_batch.quantity = Decimal('0')
                                    use_batch.save()
                                    
                                    # Update purchase item remaining quantity
                                    if use_purchase_item:
                                        use_purchase_item.update_sold_status(take_from_batch)
                                    
                                    remaining_qty -= take_from_batch
                                
                                # Update product total quantity
                                product.quantity -= quantity
                                if product.quantity < 0:
                                    product.quantity = Decimal('0')
                                product.save()
                                
                                # Continue to next item since we've handled this one
                                continue
                            else:
                                # No other batches available
                                total_available = batch.quantity
                                raise ValueError(
                                    f"Insufficient stock for {product.name}. "
                                    f"Available: {total_available}, Requested: {quantity}"
                                )
                    else:
                        # Auto-select batches using FIFO (oldest first)
                        available_batches = Batch.objects.filter(
                            product=product, 
                            quantity__gt=0, 
                            is_active=True
                        ).order_by('expiry_date', 'id')
                        
                        if not available_batches.exists():
                            raise ValueError(f"No available stock for product {product.name}")
                        
                        # Check total available quantity
                        total_available = available_batches.aggregate(total=Sum('quantity'))['total'] or 0
                        if total_available < quantity:
                            raise ValueError(
                                f"Insufficient stock for {product.name}. "
                                f"Available: {total_available}, Requested: {quantity}"
                            )
                        
                        # Split across multiple batches if needed
                        remaining_quantity = quantity
                        
                        for available_batch in available_batches:
                            if remaining_quantity <= 0:
                                break
                            
                            take_from_batch = min(available_batch.quantity, remaining_quantity)
                            
                            # Get purchase item for this batch
                            purchase_item = None
                            if hasattr(available_batch, 'purchase_item_source'):
                                purchase_item = available_batch.purchase_item_source
                            
                            # Create sale item for this batch portion
                            sale_item = SaleItem.objects.create(
                                sale=sale,
                                product=product,
                                batch=available_batch,
                                purchase_item=purchase_item,
                                quantity=take_from_batch,
                                price=Decimal(str(item['price'])),
                                purchase_price=available_batch.purchase_price,
                                discount_amount=Decimal(str(item.get('discount_amount', 0))),
                                discount_percent=Decimal(str(item.get('discount_percent', 0))),
                                total=round_to_nearest(take_from_batch * Decimal(str(item['price'])))
                            )
                            
                            # Update batch stock
                            available_batch.quantity -= take_from_batch
                            if available_batch.quantity < 0:
                                available_batch.quantity = Decimal('0')
                            available_batch.save()
                            
                            # Update purchase item remaining quantity
                            if purchase_item:
                                purchase_item.update_sold_status(take_from_batch)
                            
                            remaining_quantity -= take_from_batch
                        
                        # Update product total quantity
                        product.quantity -= quantity
                        if product.quantity < 0:
                            product.quantity = Decimal('0')
                        product.save()
                        
                        # Continue to next item since we've handled this one
                        continue
                    
                    # If we get here with batch_id and single batch, create single sale item
                    sale_item = SaleItem.objects.create(
                        sale=sale,
                        product=product,
                        batch=batch,
                        purchase_item=purchase_item,
                        quantity=quantity,
                        price=Decimal(str(item['price'])),
                        purchase_price=batch.purchase_price,
                        discount_amount=Decimal(str(item.get('discount_amount', 0))),
                        discount_percent=Decimal(str(item.get('discount_percent', 0))),
                        total=Decimal(str(item['total']))
                    )
                    
                    # Update batch stock
                    batch.quantity -= quantity
                    if batch.quantity < 0:
                        batch.quantity = Decimal('0')
                    batch.save()
                    
                    # Update purchase item remaining quantity
                    if purchase_item:
                        purchase_item.update_sold_status(quantity)
                    
                    # Update product total quantity
                    product.quantity -= quantity
                    if product.quantity < 0:
                        product.quantity = Decimal('0')
                    product.save()
                
                # Update customer balance if credit sale
                if sale.is_credit and sale.customer:
                    customer = Customer.objects.select_for_update().get(id=sale.customer.id)
                    customer.balance += sale.balance
                    customer.save()
                
                # Delete pending sale if it exists
                if pending_sale_id:
                    try:
                        pending_sale = PendingSale.objects.filter(id=pending_sale_id, user=request.user).first()
                        if pending_sale:
                            pending_sale.delete()
                    except PendingSale.DoesNotExist:
                        pass
                
                # Prepare receipt data - format quantity with 5 decimal places
                receipt = {
                    'company': {
                        'name': company.name,
                        'address': company.address or '',
                        'phone': company.phone or '',
                        'vat_number': company.vat_number or '',
                    },
                    'sale_id': sale.id,
                    'sale_number': sale.sale_number,
                    'date': timezone.localtime(sale.date).strftime('%d/%m/%Y %H:%M'),
                    'customer': sale.customer.name if sale.customer else 'Walk-in Customer',
                    'payment_method': sale.get_payment_method_display(),
                    'payment_details': {
                        'mpesa': str(sale.mpesa_amount),
                        'cash': str(sale.cash_amount),
                        'card': str(sale.card_amount),
                        'cheque': str(sale.cheque_amount)
                    },
                    'items': [{
                        'name': item.product.name,
                        'quantity': float(item.quantity),
                        'quantity_display': format(item.quantity, '.5f'),
                        'price': str(item.price),
                        'purchase_price': str(item.purchase_price),
                        'total': str(item.total),
                        'batch_number': item.batch.batch_number if item.batch and hasattr(item.batch, 'batch_number') else '',
                        'profit': float(item.profit),
                        'profit_margin': float(item.profit_margin)
                    } for item in sale.items.all()],
                    'subtotal': str(sale.subtotal),
                    'discount_percent': str(sale.discount_percent),
                    'discount_cash': str(sale.discount_amount),
                    'total': str(sale.total),
                    'amount_paid': str(sale.amount_paid),
                    'change': str(max(sale.amount_paid - sale.total, Decimal(0))),
                    'balance': str(sale.balance),
                    'sale_type': sale.sale_type
                }
                
                return JsonResponse({
                    'success': True,
                    'receipt': receipt,
                    'sale_id': sale.id,
                    'pending_sale_deleted': bool(pending_sale_id),
                    'is_edit': bool(edit_sale_id)
                })
        
        except OperationalError as e:
            if "database is locked" in str(e) and attempt < max_retries - 1:
                time.sleep(retry_delay)
                continue
            return JsonResponse({
                'success': False,
                'message': 'Database is locked. Please try again.'
            }, status=500)
        except ValueError as e:
            return JsonResponse({
                'success': False,
                'message': str(e)
            }, status=400)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error processing sale: {str(e)}'
            }, status=500)
    
    return JsonResponse({
        'success': False,
        'message': 'Database is locked. Please try again.'
    }, status=500)



# views.py - Add this view to get batches for a product

# views.py - Add or update this view

@login_required
@require_GET
def get_product_batches_with_prices(request, product_id):
    """Get all available batches for a product with their prices"""
    try:
        product = get_object_or_404(Product, pk=product_id, is_active=True)
        sale_type = request.GET.get('sale_type', 'retail')
        
        # Get all active batches with quantity > 0
        batches = product.batches.filter(quantity__gt=0, is_active=True).order_by('expiry_date', 'id')
        
        batches_data = []
        for batch in batches:
            # Determine price based on sale type and batch settings
            if sale_type == 'wholesale':
                price = batch.wholesale_price or product.wholesale_price
            else:
                price = batch.selling_price or product.selling_price
            
            # Calculate profit margin
            if batch.purchase_price > 0:
                profit_margin = ((price - batch.purchase_price) / batch.purchase_price) * 100
            else:
                profit_margin = 0
            
            batches_data.append({
                'id': batch.id,
                'batch_number': batch.batch_number,
                'quantity': float(batch.quantity),
                'purchase_price': float(batch.purchase_price),
                'selling_price': float(price),
                'expiry_date': batch.expiry_date.strftime('%Y-%m-%d') if batch.expiry_date else None,
                'profit_margin': float(profit_margin),
                'is_expired': batch.expiry_date and batch.expiry_date < timezone.now().date(),
            })
        
        # Also include default product price for reference
        default_price = product.wholesale_price if sale_type == 'wholesale' else product.selling_price
        
        return JsonResponse({
            'success': True,
            'product_id': product.id,
            'product_name': product.name,
            'default_price': float(default_price),
            'batches': batches_data,
            'total_available': float(product.quantity)
        })
        
    except Product.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Product not found'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)

# views.py - Add this view for stock management

@login_required
def stock_management(request):
    """View to manage stock and fix inconsistencies"""
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'create_batches':
            # Create batches for products without them
            products_without_batches = Product.objects.filter(
                quantity__gt=0,
                batches__isnull=True
            )
            
            created_count = 0
            for product in products_without_batches:
                Batch.objects.create(
                    product=product,
                    batch_number=f"AUTO-{product.id}-{timezone.now().strftime('%Y%m%d')}",
                    quantity=product.quantity,
                    purchase_price=product.purchase_price,
                    selling_price=product.selling_price,
                    wholesale_price=product.wholesale_price,
                    date_received=timezone.now().date(),
                    is_active=True
                )
                created_count += 1
            
            messages.success(request, f'Created {created_count} batches successfully!')
            
        elif action == 'fix_batch_quantities':
            # Fix batch quantities to match product totals
            products = Product.objects.filter(batches__isnull=False).distinct()
            fixed_count = 0
            
            for product in products:
                batch_total = product.batches.aggregate(total=models.Sum('quantity'))['total'] or Decimal('0')
                
                if batch_total != product.quantity:
                    difference = product.quantity - batch_total
                    latest_batch = product.batches.order_by('-date_received', '-id').first()
                    
                    if latest_batch:
                        latest_batch.quantity += difference
                        latest_batch.save()
                        fixed_count += 1
            
            messages.success(request, f'Fixed quantities for {fixed_count} products!')
            
        elif action == 'fix_sale_items':
            # Fix sale items without purchase_price
            sale_items = SaleItem.objects.filter(
                purchase_price=0,
                batch__isnull=False
            )
            
            fixed_count = 0
            for item in sale_items:
                if item.batch:
                    item.purchase_price = item.batch.purchase_price
                    item.save()
                    fixed_count += 1
            
            messages.success(request, f'Fixed {fixed_count} sale items!')
        
        return redirect('stock_management')
    
    # GET request - show stock status
    products = Product.objects.annotate(
        batch_count=models.Count('batches'),
        batch_total=models.Sum('batches__quantity')
    ).order_by('name')
    
    # Find issues
    issues = []
    
    for product in products:
        if product.quantity > 0 and product.batch_count == 0:
            issues.append({
                'product': product,
                'issue': 'Has stock but no batches',
                'severity': 'high'
            })
        elif product.batch_count > 0 and product.batch_total != product.quantity:
            issues.append({
                'product': product,
                'issue': f'Batch total ({product.batch_total}) != Product quantity ({product.quantity})',
                'severity': 'medium'
            })
    
    # Find batches with issues
    batches_with_issues = Batch.objects.filter(
        models.Q(quantity__lt=0) | 
        models.Q(is_active=False, quantity__gt=0)
    )
    
    context = {
        'products': products,
        'issues': issues,
        'batches_with_issues': batches_with_issues,
        'sale_items_without_purchase': SaleItem.objects.filter(purchase_price=0, batch__isnull=False).count(),
        'products_without_batches': Product.objects.filter(quantity__gt=0, batches__isnull=True).count(),
    }
    
    return render(request, 'pos/stock_management.html', context)

@login_required
@require_GET
def check_batch_availability(request, batch_id, quantity):
    """Check if a specific batch has enough quantity"""
    try:
        batch = get_object_or_404(Batch, pk=batch_id, is_active=True)
        quantity_needed = Decimal(str(quantity))
        
        if batch.quantity >= quantity_needed:
            return JsonResponse({
                'success': True,
                'available': True,
                'batch_quantity': float(batch.quantity),
                'requested_quantity': float(quantity_needed)
            })
        else:
            return JsonResponse({
                'success': True,
                'available': False,
                'batch_quantity': float(batch.quantity),
                'requested_quantity': float(quantity_needed),
                'message': f'Only {batch.quantity} available in this batch'
            })
            
    except Batch.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Batch not found'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


# pos/views.py (You need to update this view to properly check stock)
# Add/modify this view to get batches with available stock:

def get_product_batches(request, product_id):
    """Get available batches for a product with sufficient stock"""
    from inventory.models import Batch
    
    sale_type = request.GET.get('sale_type', 'retail')
    
    # Get batches with stock > 0
    batches = Batch.objects.filter(
        product_id=product_id,
        quantity__gt=0  # Only get batches with available stock
    ).order_by('expiry_date')  # Oldest expiry first
    
    batches_data = []
    for batch in batches:
        # Get the appropriate price based on sale type
        if sale_type == 'wholesale' and batch.wholesale_price:
            selling_price = batch.wholesale_price
        else:
            selling_price = batch.selling_price or batch.product.selling_price
            
        batches_data.append({
            'id': batch.id,
            'batch_number': batch.batch_number,
            'quantity': float(batch.quantity),
            'expiry_date': batch.expiry_date.strftime('%Y-%m-%d') if batch.expiry_date else None,
            'purchase_price': float(batch.purchase_price),
            'selling_price': float(selling_price),
        })
    
    # Also include product default for products without batches
    product = Batch.objects.filter(id=product_id).first()
    if not batches_data and product:
        batches_data.append({
            'id': None,
            'batch_number': 'Default',
            'quantity': product.quantity,
            'expiry_date': None,
            'purchase_price': float(product.purchase_price),
            'selling_price': float(product.selling_price) if sale_type == 'retail' else float(product.wholesale_price or product.selling_price),
        })
    
    return JsonResponse({'batches': batches_data, 'success': True})

@login_required
@require_GET
def get_sale_id(request, sale_number):
    """Get sale ID from sale number"""
    try:
        sale = Sale.objects.filter(sale_number=sale_number).first()
        if sale:
            return JsonResponse({'sale_id': sale.id})
        else:
            return JsonResponse({'error': 'Sale not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.db import transaction
from decimal import Decimal
import json
from .models import PendingSale, Sale, SaleItem, Product, Batch, Customer, Company


# pos/views.py
@login_required
@require_POST
@login_required
@login_required
@require_POST
def save_pending_sale(request):
    try:
        # Get raw POST data
        data = json.loads(request.body.decode('utf-8'))
        sale_data = data.get('sale_data')
        pending_sale_id = data.get('pending_sale_id')  # Get pending sale ID if provided
        
        if not sale_data:
            return JsonResponse({
                'success': False,
                'message': 'No sale data provided'
            }, status=400)
        
        # Validate required fields
        required_fields = ['items', 'total']
        for field in required_fields:
            if field not in sale_data:
                return JsonResponse({
                    'success': False,
                    'message': f'Missing required field: {field}'
                }, status=400)
        
        # Ensure items is a list
        if not isinstance(sale_data['items'], list):
            return JsonResponse({
                'success': False,
                'message': 'Items must be an array'
            }, status=400)
        
        # If we have a pending_sale_id, try to update the existing pending sale
        if pending_sale_id:
            try:
                pending_sale = PendingSale.objects.get(
                    id=pending_sale_id, 
                    user=request.user
                )
                # Update the existing pending sale
                pending_sale.customer_id = sale_data.get('customer_id')
                pending_sale.data = sale_data
                pending_sale.updated_at = timezone.now()  # Track when it was last updated
                pending_sale.save()
                
                message = 'Pending sale updated successfully'
                
            except PendingSale.DoesNotExist:
                # If pending sale doesn't exist, create a new one
                pending_sale = PendingSale(
                    user=request.user,
                    customer_id=sale_data.get('customer_id'),
                    data=sale_data
                )
                pending_sale.save()
                message = 'New pending sale created successfully'
        else:
            # Create a new pending sale
            pending_sale = PendingSale(
                user=request.user,
                customer_id=sale_data.get('customer_id'),
                data=sale_data
            )
            pending_sale.save()
            message = 'Sale saved as pending successfully'
        
        return JsonResponse({
            'success': True,
            'pending_sale_id': pending_sale.id,
            'message': message,
            'is_update': bool(pending_sale_id)  # Indicate if this was an update
        })
    
    except json.JSONDecodeError as e:
        return JsonResponse({
            'success': False,
            'message': f'Invalid JSON data: {str(e)}'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


@login_required
def pending_sales_list(request):
    pending_sales = PendingSale.objects.filter(user=request.user).order_by('-created_at')
    
    sales_data = []
    for sale in pending_sales:
        try:
            data = sale.data
            sales_data.append({
                'id': sale.id,
                'created_at': sale.created_at.strftime("%Y-%m-%d %H:%M:%S"),
                'customer': sale.customer.name if sale.customer else "Walk-in",
                'items_count': len(data.get('items', [])),
                'total': float(data.get('total', 0)),
                'sale_type': data.get('sale_type', 'retail'),
                'data': data  # Include full sale data for reference
            })
        except Exception as e:
            print(f"Error processing pending sale {sale.id}: {str(e)}")
            continue
    
    return JsonResponse({
        'success': True,
        'pending_sales': sales_data
    })

@login_required
@login_required
def load_pending_sale(request, pk):
    try:
        pending_sale = get_object_or_404(PendingSale, pk=pk, user=request.user)
        sale_data = pending_sale.data
        
        # Ensure items exists and is a list
        if 'items' not in sale_data:
            sale_data['items'] = []
        elif not isinstance(sale_data['items'], list):
            sale_data['items'] = []
        
        # Get product details for all items
        product_ids = [item.get('id') for item in sale_data.get('items', []) if item.get('id')]
        products = Product.objects.filter(id__in=product_ids).in_bulk()
        
        # Enhance items with product names
        for item in sale_data.get('items', []):
            product = products.get(int(item.get('id'))) if item.get('id') else None
            if product:
                item['name'] = product.name
                item['barcode'] = product.barcode
        
        return JsonResponse({
            'success': True,
            'sale_data': sale_data,
            'customer': {
                'id': pending_sale.customer.id if pending_sale.customer else None,
                'name': pending_sale.customer.name if pending_sale.customer else 'Walk-in Customer'
            }
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=400)

@login_required
@login_required
@require_POST
@login_required
@require_POST
def complete_pending_sale(request, pk):
    """Convert a pending purchase to a completed purchase"""
    pending_sale = get_object_or_404(PendingSale, pk=pk, user=request.user)
    
    try:
        with transaction.atomic():
            # Get data from pending sale
            pending_data = pending_sale.data
            
            # Create the sale
            sale = Sale.objects.create(
                customer_id=pending_data.get('customer_id'),
                user=request.user,
                sale_type=pending_data.get('sale_type', 'retail'),
                subtotal=Decimal(pending_data.get('subtotal', '0.00')),
                discount_amount=Decimal(pending_data.get('discount_amount', '0.00')),
                discount_percent=Decimal(pending_data.get('discount_percent', '0.00')),
                total=Decimal(pending_data.get('total', '0.00')),
                payment_method=pending_data.get('payment_method', 'cash'),
                amount_paid=Decimal(pending_data.get('amount_paid', '0.00')),
                balance=Decimal(pending_data.get('balance', '0.00')),
                is_credit=pending_data.get('is_credit', False),
                is_completed=True
            )
            
            # Add items to the sale and update inventory
            for item_data in pending_data.get('items', []):
                try:
                    product = Product.objects.get(id=item_data['id'])
                    quantity = Decimal(str(item_data['quantity']))
                    price = Decimal(str(item_data['price']))
                    
                    # Create sale item
                    SaleItem.objects.create(
                        sale=sale,
                        product=product,
                        quantity=quantity,
                        price=price,
                        discount_amount=Decimal(item_data.get('discount_amount', '0.00')),
                        discount_percent=Decimal(item_data.get('discount_percent', '0.00')),
                        total=Decimal(item_data['total'])
                    )
                    
                    # Update product stock
                    product.quantity -= quantity
                    product.save()
                    
                except Product.DoesNotExist:
                    # If product doesn't exist, skip it but continue with others
                    continue
            
            # Delete pending sale AFTER successful completion
            pending_sale.delete()
            
            return JsonResponse({
                'success': True,
                'message': 'Sale completed successfully!',
                'sale_id': sale.id,
                'pending_sale_deleted': True
            })
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error completing sale: {str(e)}',
            'pending_sale_deleted': False
        }, status=500)

# pos/views.py
from django.shortcuts import render, get_object_or_404
from .models import PendingSale, Customer, Product, Company

@login_required
def view_pending_sale(request, pk):
    pending_sale = get_object_or_404(PendingSale, pk=pk, user=request.user)
    data = pending_sale.data  # Access the JSON data
    
    # Get customer object if exists
    customer = None
    if data.get('customer_id'):
        try:
            customer = Customer.objects.get(id=data['customer_id'])
        except Customer.DoesNotExist:
            customer = None
    
    # Get product details for display
    product_ids = [item['id'] for item in data.get('items', [])]
    products = Product.objects.filter(id__in=product_ids).in_bulk()
    
    context = {
        'pending_sale': pending_sale,
        'sale_data': {
            'customer': customer.name if customer else "Walk-in",
            'customer_id': data.get('customer_id'),
            'sale_type': data.get('sale_type', 'retail'),
            'subtotal': data.get('subtotal', 0),
            'discount_amount': data.get('discount_amount', 0),
            'discount_percent': data.get('discount_percent', 0),
            'total': data.get('total', 0),
            'payment_method': data.get('payment_method'),
            'payment_details': data.get('payment_details', {}),
            'amount_paid': data.get('amount_paid', 0),
            'balance': data.get('balance', 0),
            'is_credit': data.get('is_credit', False),
            'items': [
                {
                    'id': item['id'],
                    'name': products.get(int(item['id'])).name if products.get(int(item['id'])) else "Unknown Product",
                    'price': item['price'],
                    'quantity': item['quantity'],
                    'total': item['total'],
                    'batch_id': item.get('batch_id'),
                    'discount_amount': item.get('discount_amount', 0),
                    'discount_percent': item.get('discount_percent', 0)
                }
                for item in data.get('items', [])
            ]
        },
        'company': Company.objects.first()
    }
    
    return render(request, 'pos/view_pending_sale.html', context)


from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404
from .models import PendingSale

@login_required
@require_POST
def delete_pending_sale(request, pk):
    try:
        pending_sale = get_object_or_404(PendingSale, pk=pk, user=request.user)
        pending_sale.delete()
        return JsonResponse({'success': True, 'message': 'Pending sale deleted successfully'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


# views.py - Update credit_payments view
@login_required
def credit_payments(request):
    credit_sales = Sale.objects.filter(
        is_credit=True, 
        balance__gt=0
    ).select_related('customer').prefetch_related('items').order_by('-date')
    
    # Filter by customer
    customer_id = request.GET.get('customer')
    if customer_id:
        credit_sales = credit_sales.filter(customer_id=customer_id)
    
    # Filter by date range
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if start_date:
        credit_sales = credit_sales.filter(date__date__gte=start_date)
    if end_date:
        credit_sales = credit_sales.filter(date__date__lte=end_date)
    
    # Filter by balance status
    balance_status = request.GET.get('balance_status')
    if balance_status:
        if balance_status == 'high':
            credit_sales = credit_sales.filter(balance__gt=10000)
        elif balance_status == 'medium':
            credit_sales = credit_sales.filter(balance__range=(1000, 10000))
        elif balance_status == 'low':
            credit_sales = credit_sales.filter(balance__lt=1000)
        elif balance_status == 'overdue':
            # Sales older than 30 days
            thirty_days_ago = timezone.now() - timedelta(days=30)
            credit_sales = credit_sales.filter(date__lt=thirty_days_ago)
    
    # Calculate totals
    credit_sales_total = credit_sales.aggregate(
        total=Sum('total'),
        amount_paid=Sum('amount_paid'),
        balance=Sum('balance')
    )
    
    # Calculate average days outstanding
    if credit_sales.exists():
        total_days = 0
        for sale in credit_sales:
            days = (timezone.now() - sale.date).days
            total_days += days
        average_days_outstanding = total_days // credit_sales.count()
    else:
        average_days_outstanding = 0
    
    customers = Customer.objects.filter(balance__gt=0).order_by('name')
    
    # Export functionality
    export_format = request.GET.get('export')
    if export_format:
        return generate_credit_payments_export(export_format, credit_sales)
    
    # Pagination
    paginator = Paginator(credit_sales, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Date calculations for quick filters
    today = timezone.now().date()
    last_7_days = today - timedelta(days=7)
    this_month_start = today.replace(day=1)
    last_30_days = today - timedelta(days=30)
    
    context = {
        'credit_sales': page_obj,
        'customers': customers,
        'credit_sales_total': credit_sales_total,
        'total_outstanding': credit_sales_total['balance'] or Decimal('0.00'),
        'selected_customer': customer_id,
        'start_date': start_date,
        'end_date': end_date,
        'balance_status': balance_status,
        'average_days_outstanding': average_days_outstanding,
        'today': today,
        'last_7_days': last_7_days,
        'this_month_start': this_month_start,
        'last_30_days': last_30_days,
    }
    return render(request, 'pos/credit_payments.html', context)
@login_required
@login_required
@login_required
def process_credit_payment(request, pk):
    sale = get_object_or_404(Sale, pk=pk, is_credit=True, balance__gt=0)
    
    if request.method == 'POST':
        try:
            amount = Decimal(request.POST.get('amount', 0))
            payment_method = request.POST.get('payment_method', 'cash')
            reference = request.POST.get('mpesa_code', '') or request.POST.get('cheque_number', '')
            
            if amount <= 0:
                messages.error(request, 'Amount must be greater than 0')
            elif amount > sale.balance:
                messages.error(request, f'Amount cannot be greater than the outstanding balance of {sale.balance}')
            else:
                with transaction.atomic():
                    # Create customer payment record
                    CustomerPayment.objects.create(
                        sale=sale,
                        customer=sale.customer,  # This can be None for walk-in customers
                        amount=amount,
                        date=timezone.now().date(),
                        payment_method=payment_method,
                        reference=reference,
                        user=request.user
                    )
                    
                    # Update sale payment details
                    sale.amount_paid += amount
                    sale.balance -= amount
                    
                    # Mark as paid if balance is cleared
                    if sale.balance <= 0:
                        sale.is_paid = True
                    
                    sale.save()
                    
                    # Update customer balance only if customer exists
                    if sale.customer:
                        customer = sale.customer
                        customer.balance -= amount
                        customer.save()
                    
                    messages.success(request, f'Payment of {amount} recorded successfully')
                    return redirect('credit_payments')
        
        except Exception as e:
            messages.error(request, f'Error processing payment: {str(e)}')
    
    context = {
        'sale': sale,
        'max_amount': sale.balance
    }
    return render(request, 'pos/process_credit_payment.html', context)


@login_required
def view_sale(request, pk):
    sale = get_object_or_404(Sale, pk=pk)
    context = {
        'sale': sale,
        'company': Company.objects.first()
    }
    return render(request, 'pos/view_sale.html', context)

@login_required
def edit_sale(request, pk):
    sale = get_object_or_404(Sale, pk=pk)
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Get posted data
                customer_id = request.POST.get('customer')
                payment_method = request.POST.get('payment_method')
                amount_paid = Decimal(request.POST.get('amount_paid', '0.00'))
                is_credit = request.POST.get('is_credit') == 'on'
                notes = request.POST.get('notes', '')
                
                # Handle payment details
                mpesa_code = request.POST.get('mpesa_code', '')
                cheque_number = request.POST.get('cheque_number', '')
                
                # First, restore original stock for all items
                for item in sale.items.all():
                    product = item.product
                    product.quantity += item.quantity
                    product.save()
                    
                    if item.batch:
                        batch = item.batch
                        batch.quantity += item.quantity
                        batch.save()
                
                # Update customer balance if credit sale
                if sale.is_credit and sale.customer:
                    customer = sale.customer
                    customer.balance -= sale.balance
                    customer.save()
                
                # Update sale details
                sale.customer_id = customer_id if customer_id else None
                sale.payment_method = payment_method
                sale.amount_paid = amount_paid
                sale.is_credit = is_credit
                sale.notes = notes
                sale.mpesa_code = mpesa_code
                sale.cheque_number = cheque_number
                
                # Initialize totals
                subtotal = Decimal('0.00')
                discount_amount = Decimal('0.00')
                
                # Update sale items
                for item in sale.items.all():
                    product_id = str(item.product.id)
                    
                    # Get new quantity and price from POST data
                    quantity_key = f'quantity_{product_id}'
                    price_key = f'price_{product_id}'
                    
                    new_quantity = Decimal(request.POST.get(quantity_key, str(item.quantity)))
                    new_price = Decimal(request.POST.get(price_key, str(item.price)))
                    
                    # Validate quantity doesn't exceed available stock
                    available_stock = item.product.quantity + item.quantity  # Original stock restored + current quantity
                    if new_quantity > available_stock:
                        messages.error(request, f"Quantity for {item.product.name} exceeds available stock ({available_stock})")
                        return redirect('edit_sale', pk=sale.id)
                    
                    # Update item
                    item.quantity = new_quantity
                    item.price = new_price
                    item.total = new_quantity * new_price
                    item.save()
                    
                    # Add to subtotal
                    subtotal += item.total
                    
                    # Update product stock
                    product = item.product
                    product.quantity -= new_quantity
                    if product.quantity < 0:
                        product.quantity = Decimal('0')
                    product.save()
                    
                    # Update batch stock if exists
                    if item.batch:
                        batch = item.batch
                        batch.quantity -= new_quantity
                        if batch.quantity < 0:
                            batch.quantity = Decimal('0')
                        batch.save()
                
                # Get discount values
                discount_amount = Decimal(request.POST.get('discount_amount', '0.00'))
                discount_percent = Decimal(request.POST.get('discount_percent', '0.00'))
                
                # Calculate discount if percentage is provided
                if discount_percent > 0:
                    discount_amount = subtotal * (discount_percent / 100)
                
                # Calculate totals
                sale.subtotal = subtotal
                sale.discount_amount = discount_amount
                sale.discount_percent = discount_percent
                sale.total = subtotal - discount_amount
                sale.balance = sale.total - amount_paid if not is_credit else sale.total
                
                # Handle payment details
                if sale.payment_method == 'cash':
                    sale.cash_amount = amount_paid
                elif sale.payment_method == 'mpesa':
                    sale.mpesa_amount = amount_paid
                elif sale.payment_method == 'card':
                    sale.card_amount = amount_paid
                elif sale.payment_method == 'cheque':
                    sale.cheque_amount = amount_paid
                
                sale.save()
                
                # Update customer balance if credit sale
                if sale.is_credit and sale.customer:
                    customer = sale.customer
                    customer.balance += sale.balance
                    customer.save()
                
                messages.success(request, f'Sale #{sale.sale_number} updated successfully!')
                return redirect('sale_detail', pk=sale.id)
                
        except Exception as e:
            messages.error(request, f'Error updating sale: {str(e)}')
    
    # GET request - prepare data for editing
    customers = Customer.objects.all().order_by('name')
    company = Company.objects.first()
    
    context = {
        'sale': sale,
        'customers': customers,
        'company': company
    }
    return render(request, 'pos/edit_sale.html', context)

# views.py
from django.shortcuts import render, get_object_or_404

@login_required
def sale_detail(request, pk):
    sale = get_object_or_404(Sale, pk=pk)
    context = {
        'sale': sale,
        'company': Company.objects.first()
    }
    return render(request, 'pos/sale_detail.html', context)

# ============== Product & Inventory Views ==============
@login_required
def product_list(request):
    products = Product.objects.all().order_by('name')
    
    # Search functionality - removed description from the filter
    search_query = request.GET.get('search')
    if search_query:
        products = products.filter(
            Q(name__icontains=search_query) |
            Q(barcode__icontains=search_query)
        )
    # Rest of your view remains the same...
    category_id = request.GET.get('category')
    if category_id:
        products = products.filter(category_id=category_id)
    # Stock filter
    stock_filter = request.GET.get('stock')
    if stock_filter == 'low':
        products = products.filter(quantity__lte=F('reorder_level'), quantity__gt=0)
    elif stock_filter == 'out':
        products = products.filter(quantity__lte=0)
    elif stock_filter == 'active':
        products = products.filter(is_active=True)
    elif stock_filter == 'inactive':
        products = products.filter(is_active=False)
    
    # Get all categories for filter dropdown
    categories = Category.objects.all()
    
    # Calculate total stock value
    total_value = products.aggregate(
        total=Sum(F('quantity') * F('selling_price'))
    )['total'] or 0
    
    # Pagination
    paginator = Paginator(products, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    export_format = request.GET.get('export')
    if export_format:
        return generate_product_list_export(export_format, products)
    
    context = {
        'products': page_obj,
        'categories': categories,
        'search_query': search_query,
        'selected_category': category_id,
        'stock_filter': stock_filter,
        'total_value': total_value
    }
    return render(request, 'pos/products.html', context)

def generate_product_list_export(format, queryset):
    if format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="product_list.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            'Name', 'Barcode', 'Category', 'Supplier', 
            'Purchase Price', 'Selling Price', 'Wholesale Price',
            'Current Stock', 'Reorder Level', 'Stock Value', 'Status'
        ])
        
        for product in queryset:
            stock_value = product.quantity * product.purchase_price
            writer.writerow([
                product.name,
                product.barcode or '',
                product.category.name if product.category else '',
                product.supplier.name if product.supplier else '',
                float(product.purchase_price),
                float(product.selling_price),
                float(product.wholesale_price),
                product.quantity,
                product.reorder_level,
                float(stock_value),
                'Active' if product.is_active else 'Inactive'
            ])
        
        return response

@login_required
def add_product(request):
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES)
        if form.is_valid():
            product = form.save(commit=False)
            
            # Generate barcode if not provided
            if not product.barcode:
                last_product = Product.objects.order_by('-id').first()
                next_id = last_product.id + 1 if last_product else 1
                product.barcode = f"PRD{next_id:08d}"
            
            # Ensure least_selling_price is valid
            if product.least_selling_price > product.selling_price:
                messages.error(request, "Least selling price cannot be higher than retail price.")
                suppliers = Supplier.objects.all()
                categories = Category.objects.all()
                context = {
                    'form': form,
                    'suppliers': suppliers,
                    'categories': categories
                }
                return render(request, 'pos/add_product.html', context)
            
            if product.least_selling_price < product.purchase_price:
                messages.error(request, "Least selling price cannot be less than purchase price.")
                suppliers = Supplier.objects.all()
                categories = Category.objects.all()
                context = {
                    'form': form,
                    'suppliers': suppliers,
                    'categories': categories
                }
                return render(request, 'pos/add_product.html', context)
            
            # Ensure wholesale price is valid
            if product.wholesale_price < product.purchase_price:
                messages.error(request, "Wholesale price cannot be less than purchase price.")
                suppliers = Supplier.objects.all()
                categories = Category.objects.all()
                context = {
                    'form': form,
                    'suppliers': suppliers,
                    'categories': categories
                }
                return render(request, 'pos/add_product.html', context)
            
            # Save the product with all decimal values
            try:
                product.save()
                messages.success(request, f"Product '{product.name}' added successfully!")
                return redirect('product_list')
            except Exception as e:
                messages.error(request, f"Error saving product: {str(e)}")
                suppliers = Supplier.objects.all()
                categories = Category.objects.all()
                context = {
                    'form': form,
                    'suppliers': suppliers,
                    'categories': categories
                }
                return render(request, 'pos/add_product.html', context)
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = ProductForm()
        # Set default values
        form.fields['least_selling_price'].initial = Decimal('0.0000')
        form.fields['quantity'].initial = Decimal('0.000000')
        form.fields['reorder_level'].initial = Decimal('5.000000')
        form.fields['wholesale_min_quantity'].initial = Decimal('0.000000')
    
    suppliers = Supplier.objects.all()
    categories = Category.objects.all()
    
    context = {
        'form': form,
        'suppliers': suppliers,
        'categories': categories
    }
    return render(request, 'pos/add_product.html', context)

@login_required
def edit_product(request, pk):
    product = get_object_or_404(Product, pk=pk)
    
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES, instance=product)
        if form.is_valid():
            updated_product = form.save(commit=False)
            
            # Validate least_selling_price
            if updated_product.least_selling_price > updated_product.selling_price:
                messages.error(request, "Least selling price cannot be higher than retail price.")
                suppliers = Supplier.objects.all()
                categories = Category.objects.all()
                context = {
                    'form': form,
                    'product': product,
                    'suppliers': suppliers,
                    'categories': categories
                }
                return render(request, 'pos/edit_product.html', context)
            
            if updated_product.least_selling_price < updated_product.purchase_price:
                messages.error(request, "Least selling price cannot be less than purchase price.")
                suppliers = Supplier.objects.all()
                categories = Category.objects.all()
                context = {
                    'form': form,
                    'product': product,
                    'suppliers': suppliers,
                    'categories': categories
                }
                return render(request, 'pos/edit_product.html', context)
            
            # Validate wholesale price
            if updated_product.wholesale_price < updated_product.purchase_price:
                messages.error(request, "Wholesale price cannot be less than purchase price.")
                suppliers = Supplier.objects.all()
                categories = Category.objects.all()
                context = {
                    'form': form,
                    'product': product,
                    'suppliers': suppliers,
                    'categories': categories
                }
                return render(request, 'pos/edit_product.html', context)
            
            # Save with proper decimal handling
            try:
                form.save()
                messages.success(request, f"Product '{updated_product.name}' updated successfully!")
                return redirect('product_list')
            except Exception as e:
                messages.error(request, f"Error updating product: {str(e)}")
                suppliers = Supplier.objects.all()
                categories = Category.objects.all()
                context = {
                    'form': form,
                    'product': product,
                    'suppliers': suppliers,
                    'categories': categories
                }
                return render(request, 'pos/edit_product.html', context)
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = ProductForm(instance=product)
        # Ensure decimal fields are properly displayed
        if product.quantity:
            form.initial['quantity'] = product.quantity.quantize(Decimal('0.000001'))
        if product.reorder_level:
            form.initial['reorder_level'] = product.reorder_level.quantize(Decimal('0.000001'))
        if product.wholesale_min_quantity:
            form.initial['wholesale_min_quantity'] = product.wholesale_min_quantity.quantize(Decimal('0.000001'))
    
    suppliers = Supplier.objects.all()
    categories = Category.objects.all()
    
    context = {
        'form': form,
        'product': product,
        'suppliers': suppliers,
        'categories': categories
    }
    return render(request, 'pos/edit_product.html', context)

@login_required
def toggle_product_status(request, pk):
    product = get_object_or_404(Product, pk=pk)
    product.is_active = not product.is_active
    product.save()
    return redirect('product_list')

@login_required
def product_detail(request, pk):
    product = get_object_or_404(Product, pk=pk)
    batches = product.batches.filter(quantity__gt=0).order_by('expiry_date')
    return render(request, 'pos/product_detail.html', {'product': product, 'batches': batches})

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import F, Sum, Q, DecimalField
from django.db.models.expressions import ExpressionWrapper
from django.http import HttpResponse
from django.core.paginator import Paginator
from decimal import Decimal
import csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font
from .models import Product, Category, Supplier

@login_required
def inventory_report(request):
    # Debug: print received parameters
    print(f"GET parameters: {dict(request.GET)}")
    
    # Start with base queryset - FIXED: Use the annotated queryset
    products = Product.objects.annotate(
        stock_value=ExpressionWrapper(
            F('quantity') * F('purchase_price'),
            output_field=DecimalField()
        )
    ).order_by('name')
    
    print(f"Initial products count: {products.count()}")
    
    # Apply filters with proper validation
    category_id = request.GET.get('category')
    if category_id and category_id.isdigit():
        products = products.filter(category_id=int(category_id))
        print(f"After category filter ({category_id}): {products.count()}")
    
    supplier_id = request.GET.get('supplier')
    if supplier_id and supplier_id.isdigit():
        products = products.filter(supplier_id=int(supplier_id))
        print(f"After supplier filter ({supplier_id}): {products.count()}")
    
    # Stock status filter - FIXED: Handle all cases properly
    stock_filter = request.GET.get('stock', '')
    if stock_filter:
        if stock_filter == 'low':
            products = products.filter(quantity__gt=0, quantity__lte=F('reorder_level'))
            print(f"After low stock filter: {products.count()}")
        elif stock_filter == 'out':
            products = products.filter(quantity__lte=0)
            print(f"After out of stock filter: {products.count()}")
        elif stock_filter == 'active':
            products = products.filter(is_active=True)
            print(f"After active filter: {products.count()}")
        elif stock_filter == 'inactive':
            products = products.filter(is_active=False)
            print(f"After inactive filter: {products.count()}")
    
    # Search functionality - FIXED: Handle search properly
    search_query = request.GET.get('search', '').strip()
    if search_query:
        products = products.filter(
            Q(name__icontains=search_query) |
            Q(barcode__icontains=search_query)
        )
        print(f"After search filter ({search_query}): {products.count()}")
    
    print(f"Final products count: {products.count()}")
    
    # Calculate summary statistics using the filtered queryset
    total_products = products.count()
    total_stock_value = products.aggregate(total=Sum('stock_value'))['total'] or Decimal('0.00')
    
    # Calculate counts for the filtered results
    in_stock_count = products.filter(quantity__gt=0).count()
    low_stock_count = products.filter(quantity__gt=0, quantity__lte=F('reorder_level')).count()
    out_of_stock_count = products.filter(quantity__lte=0).count()
    
    # Get filter options
    categories = Category.objects.all()
    suppliers = Supplier.objects.all()
    
    # Export functionality
    export_format = request.GET.get('export')
    if export_format:
        return generate_inventory_export(export_format, products)
    
    # Pagination - FIXED: Use the filtered queryset
    paginator = Paginator(products, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Pass filter values to template for proper display
    context = {
        'products': page_obj,
        'categories': categories,
        'suppliers': suppliers,
        'total_products': total_products,
        'total_stock_value': total_stock_value,
        'in_stock_count': in_stock_count,
        'low_stock_count': low_stock_count,
        'out_of_stock_count': out_of_stock_count,
        'selected_category': category_id if category_id else '',
        'selected_supplier': supplier_id if supplier_id else '',
        'stock_filter': stock_filter,
        'search_query': search_query,
    }
    return render(request, 'pos/inventory_report.html', context)



def generate_inventory_export(format, queryset):
    if format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="inventory_report_{}.csv"'.format(
            datetime.now().strftime('%Y%m%d_%H%M%S')
        )
        
        writer = csv.writer(response)
        writer.writerow([
            'Product Name', 'Barcode', 'Category', 'Supplier', 
            'Quantity', 'Reorder Level', 'Purchase Price (KES)', 
            'Selling Price (KES)', 'Stock Value (KES)', 'Status'
        ])
        
        for product in queryset:
            stock_value = product.quantity * product.purchase_price
            writer.writerow([
                product.name,
                product.barcode or '',
                product.category.name if product.category else '',
                product.supplier.name if product.supplier else '',
                product.quantity,
                product.reorder_level,
                float(product.purchase_price),
                float(product.selling_price),
                float(stock_value),
                'Active' if product.is_active else 'Inactive'
            ])
        
        return response
    
    elif format == 'excel':
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="inventory_report_{}.xlsx"'.format(
            datetime.now().strftime('%Y%m%d_%H%M%S')
        )
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory Report"
        
        # Add headers (REMOVED SKU)
        headers = [
            'Product Name', 'Barcode', 'Category', 'Supplier', 
            'Quantity', 'Reorder Level', 'Purchase Price (KES)', 
            'Selling Price (KES)', 'Stock Value (KES)', 'Status'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
        
        # Add data (FIXED column mapping)
        for row_num, product in enumerate(queryset, 2):
            stock_value = product.quantity * product.purchase_price
            ws.cell(row=row_num, column=1, value=product.name)
            ws.cell(row=row_num, column=2, value=product.barcode or '')
            ws.cell(row=row_num, column=3, value=product.category.name if product.category else '')
            ws.cell(row=row_num, column=4, value=product.supplier.name if product.supplier else '')
            ws.cell(row=row_num, column=5, value=product.quantity)
            ws.cell(row=row_num, column=6, value=product.reorder_level)
            ws.cell(row=row_num, column=7, value=float(product.purchase_price))
            ws.cell(row=row_num, column=8, value=float(product.selling_price))
            ws.cell(row=row_num, column=9, value=float(stock_value))
            ws.cell(row=row_num, column=10, value='Active' if product.is_active else 'Inactive')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(response)
        return response
    
    return HttpResponse('Invalid export format', status=400)

# views.py - Update profit_margin_report to use batch data

@login_required
def profit_margin_report(request):
    """Enhanced profit margin report using batch data"""
    # Date filtering
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if not start_date:
        # Default to beginning of current month
        start_date = timezone.now().date().replace(day=1)
        end_date = timezone.now().date()
    
    # Get sales items in date range with batch info
    sale_items = SaleItem.objects.filter(
        sale__date__date__range=[start_date, end_date],
        sale__is_completed=True,
        batch__isnull=False  # Only include items with batch info
    ).select_related('product', 'batch', 'product__category')
    
    # Calculate profit by product using batch purchase prices
    product_data = {}
    total_revenue = Decimal('0')
    total_cost = Decimal('0')
    total_profit = Decimal('0')
    
    for item in sale_items:
        if item.product:
            product_id = item.product.id
            if product_id not in product_data:
                product_data[product_id] = {
                    'product': item.product,
                    'category': item.product.category.name if item.product.category else 'Uncategorized',
                    'quantity_sold': Decimal('0'),
                    'revenue': Decimal('0'),
                    'cost': Decimal('0'),
                    'profit': Decimal('0'),
                    'batches': set(),
                }
            
            # Use the purchase price from the batch (not product default)
            cost = item.purchase_price * item.quantity
            revenue = item.total
            profit = revenue - cost
            
            product_data[product_id]['quantity_sold'] += item.quantity
            product_data[product_id]['revenue'] += revenue
            product_data[product_id]['cost'] += cost
            product_data[product_id]['profit'] += profit
            if item.batch:
                product_data[product_id]['batches'].add(item.batch.batch_number)
            
            total_revenue += revenue
            total_cost += cost
            total_profit += profit
    
    # Calculate margins and prepare final data
    final_data = []
    for data in product_data.values():
        profit_margin = (data['profit'] / data['revenue'] * 100) if data['revenue'] > 0 else 0
        
        final_data.append({
            'product': data['product'],
            'category': data['category'],
            'quantity_sold': data['quantity_sold'],
            'revenue': data['revenue'],
            'cost': data['cost'],
            'profit': data['profit'],
            'profit_margin': profit_margin,
            'batches_used': len(data['batches']),
        })
    
    # Sort by profit margin
    final_data.sort(key=lambda x: x['profit_margin'], reverse=True)
    
    # Calculate overall margin
    overall_margin = (total_profit / total_revenue * 100) if total_revenue > 0 else 0
    
    context = {
        'profit_data': final_data,
        'start_date': start_date,
        'end_date': end_date,
        'total_revenue': total_revenue,
        'total_cost': total_cost,
        'total_profit': total_profit,
        'overall_margin': overall_margin,
        'total_products': len(final_data),
    }
    
    return render(request, 'pos/profit_margin_report.html', context)



# ============== Batch Management Views ==============
# views.py - Enhanced batch_list view

@login_required
def batch_list(request):
    """View all batches with filtering and search"""
    # Base queryset - get all batches with related data
    batches = Batch.objects.select_related(
        'product', 
        'product__category', 
        'purchase', 
        'purchase__supplier'
    ).all().order_by('-created_at')
    
    # Apply filters
    product_id = request.GET.get('product')
    supplier_id = request.GET.get('supplier')
    status = request.GET.get('status')
    expiry_status = request.GET.get('expiry')
    
    if product_id:
        batches = batches.filter(product_id=product_id)
    
    if supplier_id:
        batches = batches.filter(purchase__supplier_id=supplier_id)
    
    if status == 'active':
        batches = batches.filter(is_active=True, quantity__gt=0)
    elif status == 'inactive':
        batches = batches.filter(is_active=False)
    elif status == 'out_of_stock':
        batches = batches.filter(quantity=0)
    
    # Expiry filtering
    today = timezone.now().date()
    if expiry_status == 'expired':
        batches = batches.filter(expiry_date__lt=today)
    elif expiry_status == 'expiring_soon':
        thirty_days_later = today + timedelta(days=30)
        batches = batches.filter(
            expiry_date__gte=today,
            expiry_date__lte=thirty_days_later
        )
    elif expiry_status == 'valid':
        batches = batches.filter(expiry_date__gt=today)
    elif expiry_status == 'no_expiry':
        batches = batches.filter(expiry_date__isnull=True)
    
    # Search functionality
    search_query = request.GET.get('search')
    if search_query:
        batches = batches.filter(
            Q(batch_number__icontains=search_query) |
            Q(product__name__icontains=search_query) |
            Q(product__barcode__icontains=search_query)
        )
    
    # Calculate statistics
    total_batches = batches.count()
    total_quantity = batches.aggregate(total=Sum('quantity'))['total'] or Decimal('0')
    total_value = batches.aggregate(
        total=Sum(F('quantity') * F('purchase_price'))
    )['total'] or Decimal('0')
    
    # Expiry statistics
    expired_count = Batch.objects.filter(
        expiry_date__lt=today,
        quantity__gt=0
    ).count()
    
    expiring_soon_count = Batch.objects.filter(
        expiry_date__gte=today,
        expiry_date__lte=today + timedelta(days=30),
        quantity__gt=0
    ).count()
    
    # Get filter dropdown data
    products = Product.objects.filter(is_active=True).order_by('name')
    suppliers = Supplier.objects.all().order_by('name')
    
    # Export functionality
    export_format = request.GET.get('export')
    if export_format:
        return export_batches(export_format, batches)
    
    # Pagination
    paginator = Paginator(batches, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'batches': page_obj,
        'products': products,
        'suppliers': suppliers,
        'total_batches': total_batches,
        'total_quantity': total_quantity,
        'total_value': total_value,
        'expired_count': expired_count,
        'expiring_soon_count': expiring_soon_count,
        'selected_product': product_id,
        'selected_supplier': supplier_id,
        'selected_status': status,
        'selected_expiry': expiry_status,
        'search_query': search_query,
        'today': today,
    }
    return render(request, 'pos/batch_list.html', context)


def export_batches(export_format, queryset):
    """Export batches to CSV or Excel"""
    if export_format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="batches_{timezone.now().date()}.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            'Batch Number', 'Product', 'Category', 'Supplier',
            'Quantity', 'Purchase Price', 'Selling Price',
            'Total Value', 'Expiry Date', 'Days to Expiry',
            'Date Received', 'Purchase Invoice', 'Status'
        ])
        
        today = timezone.now().date()
        for batch in queryset:
            days_to_expiry = None
            if batch.expiry_date:
                days_to_expiry = (batch.expiry_date - today).days
            
            writer.writerow([
                batch.batch_number,
                batch.product.name if batch.product else 'N/A',
                batch.product.category.name if batch.product and batch.product.category else 'N/A',
                batch.purchase.supplier.name if batch.purchase and batch.purchase.supplier else 'N/A',
                float(batch.quantity),
                float(batch.purchase_price),
                float(batch.selling_price),
                float(batch.quantity * batch.purchase_price),
                batch.expiry_date.strftime('%Y-%m-%d') if batch.expiry_date else 'N/A',
                days_to_expiry if days_to_expiry is not None else 'N/A',
                batch.date_received.strftime('%Y-%m-%d'),
                batch.purchase.invoice_number if batch.purchase else 'N/A',
                'Active' if batch.is_active and batch.quantity > 0 else 'Inactive',
            ])
        
        return response
    
    return HttpResponse('Invalid export format', status=400)

@login_required
def add_batch(request):
    if request.method == 'POST':
        form = BatchForm(request.POST)
        if form.is_valid():
            batch = form.save()
            
            # Update product quantity
            product = batch.product
            product.quantity += batch.quantity
            product.save()
            
            return redirect('batch_list')
    else:
        form = BatchForm()
    
    products = Product.objects.filter(is_active=True).order_by('name')
    
    context = {
        'form': form,
        'products': products
    }
    return render(request, 'pos/add_batch.html', context)

@login_required
def edit_batch(request, pk):
    batch = get_object_or_404(Batch, pk=pk)
    
    if request.method == 'POST':
        original_quantity = batch.quantity
        form = BatchForm(request.POST, instance=batch)
        
        if form.is_valid():
            batch = form.save()
            
            # Update product quantity with the difference
            product = batch.product
            product.quantity += (batch.quantity - original_quantity)
            product.save()
            
            return redirect('batch_list')
    else:
        form = BatchForm(instance=batch)
    
    context = {
        'form': form,
        'batch': batch
    }
    return render(request, 'pos/edit_batch.html', context)

# ============== Customer Views ==============
@login_required
def customer_list(request):
    customers = Customer.objects.all().order_by('name')
    
    # Search functionality
    search_query = request.GET.get('search')
    if search_query:
        customers = customers.filter(
            Q(name__icontains=search_query) |
            Q(phone__icontains=search_query) |
            Q(email__icontains=search_query))
    
    # Filter by balance status
    balance_filter = request.GET.get('balance')
    if balance_filter == 'credit':
        customers = customers.filter(balance__gt=0)
    elif balance_filter == 'zero':
        customers = customers.filter(balance=0)
    
    # Export functionality
    export_format = request.GET.get('export')
    if export_format:
        return generate_customer_export(export_format, customers)
    
    # Pagination
    paginator = Paginator(customers, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    customer_purchases = Sale.objects.filter(
        is_completed=True
    ).values(
        'customer__name'
    ).annotate(
        total_spent=Sum('total'),
        visit_count=Count('id'),
        last_visit=Max('date')
    ).order_by('-total_spent')
    
    context = {
        'customers': page_obj,
        'search_query': search_query,
        'balance_filter': balance_filter,
        'customer_purchases': customer_purchases
    }
    return render(request, 'pos/customer_list.html', context)

@login_required
def add_customer(request):
    if request.method == 'POST':
        form = CustomerForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('customer_list')
    else:
        form = CustomerForm()
    
    context = {
        'form': form
    }
    return render(request, 'pos/add_customer.html', context)

@login_required
def edit_customer(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    
    if request.method == 'POST':
        form = CustomerForm(request.POST, instance=customer)
        if form.is_valid():
            form.save()
            return redirect('customer_list')
    else:
        form = CustomerForm(instance=customer)
    
    context = {
        'form': form,
        'customer': customer
    }
    return render(request, 'pos/edit_customer.html', context)

@login_required
def customer_detail(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    sales = Sale.objects.filter(customer=customer).order_by('-date')
    
    # Pagination for sales
    paginator = Paginator(sales, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'customer': customer,
        'sales': page_obj
    }
    return render(request, 'pos/customer_detail.html', context)

@login_required
def customer_payment(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    
    if request.method == 'POST':
        amount = Decimal(request.POST.get('amount', 0))
        payment_method = request.POST.get('payment_method')
        reference = request.POST.get('reference', '')
        
        if amount > 0:
            with transaction.atomic():
                # Create a negative sale to record the payment
                payment_sale = Sale.objects.create(
                    customer=customer,
                    user=request.user,
                    subtotal=0,
                    total=amount,
                    payment_method=payment_method,
                    amount_paid=amount,
                    balance=-amount,
                    is_completed=True,
                    is_payment=True,
                    reference=reference
                )
                
                # Update customer balance
                customer.balance -= amount
                customer.save()
                
                return redirect('customer_detail', pk=pk)
    
    context = {
        'customer': customer
    }
    return render(request, 'pos/customer_payment.html', context)

def generate_customer_export(format, queryset):
    if format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="customers.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            'Name', 'Phone', 'Email', 'Address', 
            'Credit Limit', 'Balance'
        ])
        
        for customer in queryset:
            writer.writerow([
                customer.name,
                customer.phone,
                customer.email,
                customer.address,
                customer.credit_limit,
                customer.balance
            ])
    
        return response
    
    return HttpResponse('Invalid export format', status=400)

# ============== Sales Reports Views ==============
# views.py - Update the sales_report view

@login_required
def sales_report(request):
    sales = Sale.objects.filter(is_completed=True).order_by('-date')
    
    # Date filtering
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if start_date:
        sales = sales.filter(date__date__gte=start_date)
    if end_date:
        sales = sales.filter(date__date__lte=end_date)
    
    # Payment method filter
    payment_method = request.GET.get('payment_method')
    if payment_method:
        sales = sales.filter(payment_method=payment_method)
    
    # Sale type filter
    sale_type = request.GET.get('sale_type')
    if sale_type:
        sales = sales.filter(sale_type=sale_type)
    
    # Credit status filter
    credit_filter = request.GET.get('credit')
    if credit_filter == 'credit':
        sales = sales.filter(is_credit=True)
    elif credit_filter == 'paid':
        sales = sales.filter(is_credit=False)
    
    # Customer filter
    customer_id = request.GET.get('customer')
    if customer_id:
        sales = sales.filter(customer_id=customer_id)
    
    # Calculate summary stats
    total_sales = sales.aggregate(total=Sum('total'))['total'] or 0
    total_items = sales.aggregate(total=Sum('items__quantity'))['total'] or 0
    
    # Calculate payment method totals for the summary cards
    cash_sales = sales.filter(payment_method='cash').aggregate(total=Sum('total'))['total'] or 0
    mpesa_sales = sales.filter(payment_method='mpesa').aggregate(total=Sum('total'))['total'] or 0
    credit_sales = sales.filter(is_credit=True).aggregate(total=Sum('total'))['total'] or 0
    
    payment_methods = Sale.PAYMENT_METHODS
    payment_totals = []
    
    for method in payment_methods:
        total = sales.filter(payment_method=method[0]).aggregate(
            total=Sum('total'))['total'] or 0
        payment_totals.append({
            'method': method[1],
            'total': total
        })
    
    # Get all customers for filter dropdown
    customers = Customer.objects.all()

    # Pagination
    paginator = Paginator(sales, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'sales': page_obj,
        'customers': customers,
        'total_sales': total_sales,
        'total_items': total_items,
        'cash_sales': cash_sales,
        'mpesa_sales': mpesa_sales,
        'credit_sales': credit_sales,
        'payment_totals': payment_totals,
        'start_date': start_date,
        'end_date': end_date,
        'payment_method': payment_method,
        'sale_type': sale_type,
        'credit_filter': credit_filter,
        'selected_customer': customer_id,
    }
    return render(request, 'pos/sales_report.html', context)

@login_required
def daily_sales_summary(request):
    today = timezone.now().date()
    
    # Get sales for today
    sales = Sale.objects.filter(
        date__date=today,
        is_completed=True
    )
    
    # Calculate totals by payment method using the payment_method field
    # Since you don't have separate amount fields, we need to calculate based on payment_method
    payment_totals = sales.values('payment_method').annotate(
        total_amount=Sum('total')
    )
    
    # Initialize totals
    cash_sales = Decimal('0.00')
    mpesa_sales = Decimal('0.00')
    card_sales = Decimal('0.00')
    cheque_sales = Decimal('0.00')
    credit_sales = Decimal('0.00')
    
    # Calculate totals for each payment method
    for payment in payment_totals:
        method = payment['payment_method']
        amount = payment['total_amount'] or Decimal('0.00')
        
        if method == 'cash':
            cash_sales = amount
        elif method == 'mpesa':
            mpesa_sales = amount
        elif method == 'card':
            card_sales = amount
        elif method == 'cheque':
            cheque_sales = amount
    
    # Calculate credit sales separately
    credit_sales = sales.filter(is_credit=True).aggregate(
        total=Sum('total')
    )['total'] or Decimal('0.00')
    
    total_sales = sales.aggregate(total=Sum('total'))['total'] or Decimal('0.00')
    
    # Get expenses and purchases for the day
    expenses = Expense.objects.filter(date=today)  # Expense.date is already a DateField
    purchases = Purchase.objects.filter(date__date=today) 
    
    total_expenses = expenses.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    total_purchases = purchases.aggregate(total=Sum('total'))['total'] or Decimal('0.00')
    
    # Calculate net cash (excluding credit sales)
    net_cash = (cash_sales + mpesa_sales + card_sales + cheque_sales) - total_expenses - total_purchases
    
    context = {
        'date': today,
        'cash_sales': cash_sales,
        'mpesa_sales': mpesa_sales,
        'card_sales': card_sales,
        'cheque_sales': cheque_sales,
        'credit_sales': credit_sales,
        'total_sales': total_sales,
        'total_expenses': total_expenses,
        'total_purchases': total_purchases,
        'net_cash': net_cash
    }
    return render(request, 'pos/daily_sales_summary.html', context)

def generate_sales_export(format, queryset):
    if format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="sales_report.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            'Sale Number', 'Date', 'Customer', 'Items', 
            'Subtotal', 'Discount', 'Total', 
            'Payment Method', 'Amount Paid', 'Balance',
            'Cash Amount', 'Mpesa Amount', 'Card Amount', 'Cheque Amount'
        ])
        
        for sale in queryset:
            writer.writerow([
                sale.sale_number,
                sale.date.strftime('%Y-%m-%d %H:%M'),
                sale.customer.name if sale.customer else 'Walk-in',
                sale.items.aggregate(total=Sum('quantity'))['total'] or 0,
                sale.subtotal,
                sale.discount_amount,
                sale.total,
                sale.get_payment_method_display(),
                sale.amount_paid,
                sale.balance,
                sale.cash_amount,
                sale.mpesa_amount,
                sale.card_amount,
                sale.cheque_amount
            ])
        
        return response
    
    return HttpResponse('Invalid export format', status=400)

# ============== Purchase Views ==============


from django.db.models import Sum, Count, Q, ExpressionWrapper, DecimalField, Value
from django.db.models.functions import Coalesce

@login_required
def purchase_list(request):
    # Get all purchases
    purchases = Purchase.objects.all().order_by('-date')
    
    # Apply filters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    supplier_id = request.GET.get('supplier')
    is_return_filter = request.GET.get('is_return')
    return_status = request.GET.get('return_status')
    payment_status_filter = request.GET.get('payment_status')
    
    if start_date:
        purchases = purchases.filter(date__date__gte=start_date)
    if end_date:
        purchases = purchases.filter(date__date__lte=end_date)
    if supplier_id:
        purchases = purchases.filter(supplier_id=supplier_id)
    if is_return_filter:
        purchases = purchases.filter(is_return=(is_return_filter == 'true'))
    if return_status:
        purchases = purchases.filter(return_status=return_status)
    
    # First annotate with total_paid calculation
    purchases = purchases.annotate(
        annotated_total_paid=Coalesce(
            Sum('payments__amount'),
            Value(Decimal('0.0000')),
            output_field=DecimalField(max_digits=10, decimal_places=4)
        )
    )
    
    # Then annotate with balance_due calculation
    purchases = purchases.annotate(
        annotated_balance_due=ExpressionWrapper(
            F('total') - F('annotated_total_paid'),
            output_field=DecimalField(max_digits=10, decimal_places=4)
        )
    )
    
    # Apply payment status filter for non-returns
    if payment_status_filter and is_return_filter != 'true':
        if payment_status_filter == 'paid':
            purchases = purchases.filter(
                Q(is_paid=True) | 
                Q(annotated_total_paid__gte=F('total'))
            )
        elif payment_status_filter == 'partial':
            purchases = purchases.filter(
                is_paid=False,
                annotated_total_paid__gt=0,
                annotated_total_paid__lt=F('total')
            )
        elif payment_status_filter == 'unpaid':
            purchases = purchases.filter(
                is_paid=False,
                annotated_total_paid=0
            )
    
    # Format purchase data for template
    formatted_purchases = []
    for purchase in purchases:
        # Calculate payment status based on annotations
        if purchase.is_return:
            payment_status = None
        else:
            # Use the annotated values instead of properties
            annotated_total_paid = purchase.annotated_total_paid or Decimal('0.0000')
            if annotated_total_paid >= purchase.total:
                payment_status = 'Paid'
            elif annotated_total_paid > 0:
                payment_status = 'Partial'
            else:
                payment_status = 'Unpaid'
        
        # Get the balance due from annotation
        annotated_balance_due = purchase.annotated_balance_due or Decimal('0.0000')
        
        formatted_purchases.append({
            'id': purchase.id,
            'invoice_number': purchase.invoice_number,
            'date': purchase.date,
            'supplier_name': purchase.supplier.name,
            'item_count': purchase.item_count,
            'total': format_currency(purchase.total),
            'total_paid': format_currency(purchase.annotated_total_paid),
            'balance_due': format_currency(annotated_balance_due),
            'is_return': purchase.is_return,
            'return_status': purchase.return_status,
            'is_paid': purchase.is_paid,
            'payment_status': payment_status,
            'raw_total': purchase.total,
            'raw_total_paid': purchase.annotated_total_paid,
            'raw_balance_due': annotated_balance_due,
        })
    
    # Calculate summary stats - use filtered purchases but exclude returns
    purchases_for_stats = purchases.filter(is_return=False)
    
    total_purchases = purchases_for_stats.aggregate(
        total=Sum('total'))['total'] or Decimal('0.0000')
    
    total_returns = Purchase.objects.filter(
        is_return=True
    ).aggregate(total=Sum('total'))['total'] or Decimal('0.0000')
    
    total_items = purchases_for_stats.aggregate(
        total=Sum('item_count'))['total'] or 0
    
    # Calculate unpaid balance using annotations
    unpaid_balance = purchases_for_stats.aggregate(
        total_unpaid=Sum('annotated_balance_due')
    )['total_unpaid'] or Decimal('0.0000')
    
    # Get suppliers for filter dropdown
    suppliers = Supplier.objects.all()
    
    # Pagination
    paginator = Paginator(formatted_purchases, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'purchases': page_obj,
        'suppliers': suppliers,
        'total_purchases': format_currency(total_purchases),
        'total_returns': format_currency(total_returns),
        'total_items': total_items,
        'unpaid_balance': format_currency(unpaid_balance),
        'start_date': start_date,
        'end_date': end_date,
        'selected_supplier': supplier_id,
        'is_return_filter': is_return_filter,
        'return_status': return_status,
        'selected_payment_status': payment_status_filter,
        'raw_purchases': purchases,
    }
    return render(request, 'pos/purchase_list.html', context)



from django.db.models import Q
from django.http import JsonResponse
import json
from django.contrib import messages
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.db import transaction
from .models import Supplier, Product, Purchase, PurchaseItem
from .forms import PurchaseForm

from django.shortcuts import render, redirect
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.views.decorators.http import require_GET, require_POST
import json
from .models import Product, Supplier, Purchase, PurchaseItem
from .forms import PurchaseForm

# views.py - Update the add_purchase view to properly create batches

# views.py - Update the add_purchase view

# views.py - Update the add_purchase view

# views.py - Fixed add_purchase view

# views.py - Update add_purchase view

@login_required
def add_purchase(request):
    if request.method == 'POST':
        # Handle form submission
        supplier_id = request.POST.get('supplier')
        invoice_number = request.POST.get('invoice_number')
        items_data = []
        
        # Collect items data from POST
        for key, value in request.POST.items():
            if key.startswith('items['):
                import re
                match = re.match(r'items\[(\d+)\]\[(\w+)\]', key)
                if match:
                    index = int(match.group(1))
                    field = match.group(2)
                    
                    if len(items_data) <= index:
                        items_data.append({})
                    
                    if field in ['product_id', 'quantity', 'purchase_price', 'selling_price', 'batch_number', 'expiry_date']:
                        items_data[index][field] = value
        
        try:
            with transaction.atomic():
                # Validate supplier
                if not supplier_id:
                    raise ValueError("Please select a supplier")
                
                supplier = Supplier.objects.get(id=supplier_id)
                
                # Calculate totals
                rounded_total = Decimal('0.00')
                valid_items = []
                
                for item_data in items_data:
                    product_id = item_data.get('product_id')
                    if not product_id:
                        continue  # Skip empty items
                    
                    quantity = Decimal(item_data.get('quantity', '0')).quantize(Decimal('0.00001'))
                    purchase_price = Decimal(item_data.get('purchase_price', '0')).quantize(Decimal('0.0001'))
                    
                    if quantity <= 0 or purchase_price <= 0:
                        continue
                    
                    selling_price = None
                    if item_data.get('selling_price'):
                        selling_price = Decimal(item_data.get('selling_price')).quantize(Decimal('0.01'))
                    
                    batch_number = item_data.get('batch_number', '').strip()
                    expiry_date = item_data.get('expiry_date', '')
                    
                    item_total = quantity * purchase_price
                    rounded_item_total = round_to_nearest(item_total)
                    
                    valid_items.append({
                        'product_id': product_id,
                        'quantity': quantity,
                        'purchase_price': purchase_price,
                        'selling_price': selling_price,
                        'batch_number': batch_number,
                        'expiry_date': expiry_date,
                        'item_total': item_total,
                        'rounded_total': rounded_item_total
                    })
                    
                    rounded_total += rounded_item_total
                
                if not valid_items:
                    raise ValueError("Please add at least one valid item")
                
                # Generate invoice number if not provided
                if not invoice_number:
                    date_str = timezone.now().strftime('%Y%m%d')
                    last_invoice = Purchase.objects.filter(
                        invoice_number__startswith=f"INV-{date_str}-"
                    ).order_by('-invoice_number').first()
                    
                    if last_invoice:
                        try:
                            last_num = int(last_invoice.invoice_number.split('-')[-1])
                            next_num = last_num + 1
                        except:
                            next_num = 1
                    else:
                        next_num = 1
                    
                    invoice_number = f"INV-{date_str}-{next_num:04d}"
                
                # Create purchase with ROUNDED total
                purchase = Purchase.objects.create(
                    supplier=supplier,
                    invoice_number=invoice_number,
                    user=request.user,
                    subtotal=rounded_total,
                    total=rounded_total,
                    tax=Decimal('0.0000'),
                    discount=Decimal('0.0000'),
                    item_count=len(valid_items),
                    is_paid=False,
                    
                )
                
                # Create purchase items and batches (each item gets its own batch)
                for item in valid_items:
                    product = Product.objects.get(id=item['product_id'])
                    
                    # Generate batch number if not provided
                    batch_number = item['batch_number']
                    if not batch_number:
                        batch_number = f"BATCH-{timezone.now().strftime('%Y%m%d')}-{product.id}-{purchase.id}"
                    
                    # Create batch FIRST (without purchase_item)
                    batch = Batch.objects.create(
                        product=product,
                        batch_number=batch_number,
                        quantity=item['quantity'],
                        expiry_date=item['expiry_date'] if item['expiry_date'] else None,
                        purchase_price=item['purchase_price'],
                        selling_price=item['selling_price'] if item['selling_price'] else product.selling_price,
                        wholesale_price=product.wholesale_price,
                        date_received=timezone.now().date(),
                        purchase=purchase,
                        purchase_item=None,  # Set to None initially
                        is_active=True
                    )
                    
                    # Create purchase item with batch reference
                    purchase_item = PurchaseItem.objects.create(
                        purchase=purchase,
                        product=product,
                        batch=batch,  # Set batch here
                        quantity=item['quantity'],
                        price=item['purchase_price'],
                        total=item['item_total'],
                        remaining_quantity=item['quantity'],
                        is_fully_sold=False
                    )
                    
                    # Now update the batch with the purchase_item reference
                    batch.purchase_item = purchase_item
                    batch.save()
                    
                    # Update product total quantity
                    product.quantity += item['quantity']
                    product.save()
                
                # For AJAX requests, return JSON with redirect to invoice
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': True,
                        'purchase_id': purchase.id,
                        'message': f'Purchase #{purchase.invoice_number} created successfully!',
                        'redirect_url': reverse('purchase_invoice', args=[purchase.id])
                    })
                
                # For regular form submission, redirect to invoice
                messages.success(request, f'Purchase #{purchase.invoice_number} created successfully!')
                return redirect('purchase_invoice', pk=purchase.id)
                
        except Supplier.DoesNotExist:
            error_msg = "Selected supplier does not exist"
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'message': error_msg}, status=400)
            messages.error(request, error_msg)
            
        except Product.DoesNotExist as e:
            error_msg = f"Product not found: {str(e)}"
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'message': error_msg}, status=400)
            messages.error(request, error_msg)
            
        except ValueError as e:
            error_msg = str(e)
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'message': error_msg}, status=400)
            messages.error(request, error_msg)
            
        except Exception as e:
            error_msg = f'Error creating purchase: {str(e)}'
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'message': error_msg}, status=500)
            messages.error(request, error_msg)
    
    # GET request - show form
    # Remove .filter(is_active=True) since Supplier doesn't have is_active field
    suppliers = Supplier.objects.all()  # Get all suppliers
    invoice_number = generate_invoice_number()
    products = Product.objects.filter(is_active=True).order_by('name')
    
    context = {
        'suppliers': suppliers,
        'invoice_number': invoice_number,
        'products': products,
        'selected_supplier': None
    }
    return render(request, 'pos/add_purchase.html', context)
# views.py - Add this new view for purchase invoice printing



def generate_invoice_number():
    date_str = timezone.now().strftime('%Y%m%d')
    last_invoice = Purchase.objects.filter(
        invoice_number__startswith=f"INV-{date_str}-"
    ).order_by('-invoice_number').first()
    
    if last_invoice:
        try:
            last_num = int(last_invoice.invoice_number.split('-')[-1])
            next_num = last_num + 1
        except:
            next_num = 1
    else:
        next_num = 1
    
    return f"INV-{date_str}-{next_num:04d}"

# views.py - Add purchase invoice view

@login_required
def purchase_invoice(request, pk):
    """View purchase invoice"""
    purchase = get_object_or_404(
        Purchase.objects.select_related('supplier', 'user'),
        pk=pk
    )
    
    # Get all items with their batch details
    items = PurchaseItem.objects.filter(
        purchase=purchase
    ).select_related('product', 'batch')
    
    # Get company info
    company = Company.objects.first()
    
    # Calculate totals
    total_items = items.count()
    total_quantity = items.aggregate(total=Sum('quantity'))['total'] or Decimal('0')
    
    # Get payment info if any
    payments = purchase.payments.all() if hasattr(purchase, 'payments') else []
    
    context = {
        'purchase': purchase,
        'items': items,
        'company': company,
        'total_items': total_items,
        'total_quantity': total_quantity,
        'payments': payments,
        'today': timezone.now(),
    }
    
    return render(request, 'pos/purchase_invoice.html', context)


@login_required
def print_purchase_invoice(request, pk):
    """Print purchase invoice (optimized for printing)"""
    purchase = get_object_or_404(
        Purchase.objects.select_related('supplier', 'user'),
        pk=pk
    )
    
    items = PurchaseItem.objects.filter(
        purchase=purchase
    ).select_related('product', 'batch')
    
    company = Company.objects.first()
    
    context = {
        'purchase': purchase,
        'items': items,
        'company': company,
        'print_mode': True,
        'today': timezone.now(),
    }
    
    return render(request, 'pos/print_purchase_invoice.html', context)


def round_to_nearest(value):
    """Round decimal to nearest whole number"""
    if isinstance(value, Decimal):
        return value.quantize(Decimal('1'), rounding=ROUND_HALF_UP)
    try:
        decimal_value = Decimal(str(value))
        return decimal_value.quantize(Decimal('1'), rounding=ROUND_HALF_UP)
    except:
        return Decimal('0')

def generate_invoice_number():
    date_str = timezone.now().strftime('%Y%m%d')
    last_invoice = Purchase.objects.filter(
        invoice_number__startswith=f"INV-{date_str}-"
    ).order_by('-invoice_number').first()
    
    if last_invoice:
        try:
            last_num = int(last_invoice.invoice_number.split('-')[-1])
            next_num = last_num + 1
        except:
            next_num = 1
    else:
        next_num = 1
    
    return f"INV-{date_str}-{next_num:04d}"

from django.views.decorators.http import require_POST
from django.http import JsonResponse

@login_required
@require_POST
def delete_purchase(request, pk):
    """Delete a purchase"""
    purchase = get_object_or_404(Purchase, pk=pk)
    
    try:
        with transaction.atomic():
            # If it's a completed purchase, restore stock quantities
            if not purchase.is_return:
                for item in purchase.items.all():
                    product = item.product
                    product.quantity -= item.quantity
                    if product.quantity < 0:
                        product.quantity = Decimal('0')
                    product.save()
                    
                    if item.batch:
                        batch = item.batch
                        batch.quantity -= item.quantity
                        if batch.quantity < 0:
                            batch.quantity = Decimal('0')
                        batch.save()
            
            # Delete the purchase
            purchase.delete()
            
            messages.success(request, f'Purchase #{purchase.invoice_number} has been deleted successfully!')
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': 'Purchase deleted successfully!'
                })
            return redirect('purchase_list')
    
    except Exception as e:
        error_message = f'Error deleting purchase: {str(e)}'
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'message': error_message
            }, status=500)
        messages.error(request, error_message)
        return redirect('purchase_list')


# views.py - Update edit_purchase view with proper batch handling

@login_required
def edit_purchase(request, pk):
    """Edit an existing purchase with proper batch management"""
    purchase = get_object_or_404(Purchase.objects.select_related('supplier'), pk=pk)
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Get form data
                supplier_id = request.POST.get('supplier')
                invoice_number = request.POST.get('invoice_number')
                notes = request.POST.get('notes', '')
                payment_method = request.POST.get('payment_method', '')
                is_paid = request.POST.get('is_paid') == 'on'
                
                # Validate supplier
                if not supplier_id:
                    raise ValueError("Please select a supplier")
                
                supplier = Supplier.objects.get(id=supplier_id)
                
                # Collect items data
                items_data = []
                i = 0
                while f'items[{i}][product_id]' in request.POST:
                    product_id = request.POST.get(f'items[{i}][product_id]')
                    if product_id:  # Only process if product_id exists
                        quantity = Decimal(request.POST.get(f'items[{i}][quantity]', '0'))
                        purchase_price = Decimal(request.POST.get(f'items[{i}][price]', '0'))
                        selling_price = request.POST.get(f'items[{i}][selling_price]')
                        batch_number = request.POST.get(f'items[{i}][batch_number]', '').strip()
                        expiry_date = request.POST.get(f'items[{i}][expiry_date]', '')
                        item_id = request.POST.get(f'items[{i}][item_id]')
                        
                        if quantity > 0 and purchase_price > 0:
                            items_data.append({
                                'product_id': product_id,
                                'quantity': quantity,
                                'purchase_price': purchase_price,
                                'selling_price': Decimal(selling_price) if selling_price else None,
                                'batch_number': batch_number,
                                'expiry_date': expiry_date if expiry_date else None,
                                'item_id': int(item_id) if item_id else None,
                            })
                    i += 1
                
                if not items_data:
                    raise ValueError("Please add at least one valid item")
                
                # Calculate new total
                new_total = Decimal('0.00')
                for item in items_data:
                    item_total = item['quantity'] * item['purchase_price']
                    rounded_item_total = round_to_nearest(item_total)
                    new_total += rounded_item_total
                
                # STEP 1: RESTORE ORIGINAL STOCK FROM ALL EXISTING ITEMS
                for old_item in purchase.items.all():
                    if old_item.product:
                        # Restore product quantity
                        old_item.product.quantity -= old_item.quantity
                        if old_item.product.quantity < 0:
                            old_item.product.quantity = Decimal('0')
                        old_item.product.save()
                        
                        # Handle batch
                        if old_item.batch:
                            batch = old_item.batch
                            
                            # Check if this batch is used in any sales
                            sales_using_batch = SaleItem.objects.filter(
                                batch=batch,
                                sale__is_completed=True
                            ).exists()
                            
                            if sales_using_batch:
                                # Batch has been used in sales, don't delete it
                                # Just adjust its quantity
                                batch.quantity -= old_item.quantity
                                if batch.quantity < 0:
                                    batch.quantity = Decimal('0')
                                batch.save()
                            else:
                                # Batch not used in any sales, can be deleted
                                if old_item.id and PurchaseItem.objects.filter(batch=batch).exclude(id=old_item.id).exists():
                                    # Another purchase item references this batch
                                    batch.quantity -= old_item.quantity
                                    if batch.quantity < 0:
                                        batch.quantity = Decimal('0')
                                    batch.save()
                                else:
                                    # Safe to delete the batch
                                    batch.delete()
                
                # STEP 2: DELETE OLD PURCHASE ITEMS
                purchase.items.all().delete()
                
                # STEP 3: UPDATE PURCHASE HEADER
                purchase.supplier = supplier
                purchase.invoice_number = invoice_number
                purchase.subtotal = new_total
                purchase.total = new_total
                purchase.tax = Decimal('0.00')
                purchase.discount = Decimal('0.00')
                purchase.is_paid = is_paid
                purchase.payment_method = payment_method
                purchase.notes = notes
                purchase.save()
                
                # STEP 4: CREATE NEW PURCHASE ITEMS AND BATCHES
                item_count = 0
                for item_data in items_data:
                    try:
                        product = Product.objects.get(id=item_data['product_id'])
                        quantity = item_data['quantity']
                        purchase_price = item_data['purchase_price']
                        
                        # Generate batch number if not provided
                        batch_number = item_data['batch_number']
                        if not batch_number:
                            batch_number = f"BATCH-{timezone.now().strftime('%Y%m%d')}-{product.id}-{purchase.id}"
                        
                        # Create purchase item first (without batch)
                        purchase_item = PurchaseItem.objects.create(
                            purchase=purchase,
                            product=product,
                            quantity=quantity,
                            price=purchase_price,
                            total=quantity * purchase_price,
                            remaining_quantity=quantity,
                            is_fully_sold=False
                        )
                        
                        # Determine selling price for batch
                        selling_price = item_data['selling_price']
                        if not selling_price:
                            selling_price = product.selling_price
                        
                        # Check if batch already exists
                        existing_batch = Batch.objects.filter(
                            product=product,
                            batch_number=batch_number
                        ).first()
                        
                        if existing_batch:
                            # Update existing batch
                            batch = existing_batch
                            batch.quantity += quantity
                            batch.purchase_price = purchase_price
                            batch.selling_price = selling_price
                            batch.wholesale_price = product.wholesale_price
                            if item_data['expiry_date']:
                                batch.expiry_date = item_data['expiry_date']
                            batch.save()
                        else:
                            # Create new batch
                            batch = Batch.objects.create(
                                product=product,
                                batch_number=batch_number,
                                quantity=quantity,
                                expiry_date=item_data['expiry_date'] if item_data['expiry_date'] else None,
                                purchase_price=purchase_price,
                                selling_price=selling_price,
                                wholesale_price=product.wholesale_price,
                                date_received=timezone.now().date(),
                                purchase=purchase,
                                purchase_item=purchase_item,
                                is_active=True
                            )
                        
                        # Link batch to purchase item
                        purchase_item.batch = batch
                        purchase_item.save()
                        
                        # Update product quantity
                        product.quantity += quantity
                        product.save()
                        
                        item_count += 1
                        
                    except Product.DoesNotExist:
                        continue
                
                # Update purchase item count
                purchase.item_count = item_count
                purchase.save()
                
                # Log activity
                UserActivityLog.objects.create(
                    user=request.user,
                    action_type='update',
                    model_name='Purchase',
                    object_id=str(purchase.id),
                    description=f'Updated purchase #{purchase.invoice_number} with {item_count} items',
                    ip_address=request.META.get('REMOTE_ADDR'),
                    user_agent=request.META.get('HTTP_USER_AGENT', '')
                )
                
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': True,
                        'message': f'Purchase #{purchase.invoice_number} updated successfully!',
                        'redirect_url': reverse('view_purchase', args=[purchase.id])
                    })
                
                messages.success(request, f'Purchase #{purchase.invoice_number} updated successfully!')
                return redirect('view_purchase', pk=purchase.id)
                
        except Supplier.DoesNotExist:
            error_msg = "Selected supplier does not exist"
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'message': error_msg}, status=400)
            messages.error(request, error_msg)
            
        except Product.DoesNotExist as e:
            error_msg = f"Product not found: {str(e)}"
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'message': error_msg}, status=400)
            messages.error(request, error_msg)
            
        except ValueError as e:
            error_msg = str(e)
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'message': error_msg}, status=400)
            messages.error(request, error_msg)
            
        except Exception as e:
            error_msg = f'Error updating purchase: {str(e)}'
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'message': error_msg}, status=500)
            messages.error(request, error_msg)
    
    # GET request - display form with existing data
    suppliers = Supplier.objects.all().order_by('name')
    products = Product.objects.filter(is_active=True).order_by('name')
    
    # Prepare items data for template
    items = []
    for item in purchase.items.all().select_related('product', 'batch'):
        items.append({
            'item_id': item.id,
            'product_id': item.product.id if item.product else None,
            'product_name': item.product.name if item.product else 'Unknown',
            'product_barcode': item.product.barcode if item.product else '',
            'quantity': float(item.quantity),
            'price': float(item.price),
            'selling_price': float(item.batch.selling_price) if item.batch and item.batch.selling_price else float(item.product.selling_price) if item.product else 0,
            'batch_number': item.batch.batch_number if item.batch else '',
            'expiry_date': item.batch.expiry_date.strftime('%Y-%m-%d') if item.batch and item.batch.expiry_date else '',
            'total': float(item.total),
        })
    
    context = {
        'purchase': purchase,
        'suppliers': suppliers,
        'products': products,
        'items_json': json.dumps(items),
        'selected_supplier': purchase.supplier,
        'invoice_number': purchase.invoice_number,
        'is_paid': purchase.is_paid,
        'payment_method': purchase.payment_method,
        'notes': purchase.notes,
        'today': timezone.now().date(),
    }
    
    return render(request, 'pos/edit_purchase.html', context)


from django.shortcuts import get_object_or_404, render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.http import JsonResponse, Http404  # Added Http404 import
from django.urls import reverse
from django.utils import timezone
from decimal import Decimal
from django.core.exceptions import ValidationError
from .models import Purchase, Product, Supplier, PurchaseItem, Batch
from .forms import PurchaseForm  # Make sure to import your form


from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db import transaction
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from .models import Purchase, PendingPurchase

from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.views.decorators.http import require_POST

from .models import Purchase, PendingPurchase, PurchaseItem, Batch
import json

from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from django.core.serializers.json import DjangoJSONEncoder
import json

@login_required
@csrf_exempt  # Temporarily for testing, remove in production
@require_POST
@login_required
@login_required
@require_POST
def save_purchase_as_pending(request):
    try:
        data = json.loads(request.body)
        
        # Extract data
        supplier_id = data.get('supplier')
        invoice_number = data.get('invoice_number')
        items = data.get('items', [])
        
        if not supplier_id:
            return JsonResponse({
                'success': False,
                'message': 'Supplier is required'
            }, status=400)
        
        if not items:
            return JsonResponse({
                'success': False,
                'message': 'At least one item is required'
            }, status=400)
        
        # Calculate ROUNDED totals
        subtotal = Decimal('0.00')
        item_details = []
        
        for item in items:
            product_id = item.get('product_id')
            quantity = Decimal(str(item.get('quantity', 0)))
            price = Decimal(str(item.get('price', 0)))
            
            # Calculate item total
            item_total = quantity * price
            # Round to nearest whole number
            rounded_item_total = round_to_nearest(item_total)
            
            # Convert Decimal to string for JSON serialization
            item_details.append({
                'product_id': product_id,
                'quantity': str(quantity),  # Convert to string
                'price': str(price),        # Convert to string
                'total': str(item_total),   # Convert to string
                'rounded_total': str(rounded_item_total),  # Convert to string
                'batch_number': item.get('batch_number', ''),
                'expiry_date': item.get('expiry_date', '')
            })
            
            subtotal += rounded_item_total  # Use rounded total for subtotal
        
        # Generate invoice number if not provided
        if not invoice_number:
            date_str = timezone.now().strftime('%Y%m%d')
            today_pending = PendingPurchase.objects.filter(
                created_at__date=timezone.now().date()
            ).count()
            next_num = today_pending + 1
            invoice_number = f"PEND-{date_str}-{next_num:04d}"
        
        # Create pending purchase with ROUNDED total
        pending_purchase = PendingPurchase.objects.create(
            user=request.user,
            supplier_id=supplier_id,
            invoice_number=invoice_number,
            subtotal=subtotal,
            total=subtotal,
            tax=Decimal('0.0000'),
            discount=Decimal('0.0000'),
            data={
                'items': item_details,
                'notes': data.get('notes', ''),
                'subtotal': str(subtotal),  # Store as string
                'total': str(subtotal),     # Store as string
                'payment_method': data.get('payment_method', ''),
                'is_paid': data.get('is_paid', False)
            }
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Purchase saved as pending successfully',
            'pending_purchase_id': pending_purchase.id,
            'redirect_url': reverse('pending_purchases')
        })
        
    except json.JSONDecodeError as e:
        return JsonResponse({
            'success': False,
            'message': f'Invalid JSON data: {str(e)}'
        }, status=400)
    except Exception as e:
        import traceback
        traceback.print_exc()  # Print full traceback for debugging
        return JsonResponse({
            'success': False,
            'message': f'Error saving pending purchase: {str(e)}'
        }, status=500)

from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction
from django.views.decorators.http import require_POST, require_GET
from django.core.paginator import Paginator
import json
from decimal import Decimal

from .models import PendingPurchase, Product, Supplier, Purchase, PurchaseItem, Batch
from .forms import PurchaseForm

@login_required
def pending_purchases_list(request):
    """List all pending purchases"""
    pending_purchases = PendingPurchase.objects.filter(
        user=request.user, 
        status='pending'
    ).select_related('supplier').order_by('-created_at')
    
    # Pagination
    paginator = Paginator(pending_purchases, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'pending_purchases': page_obj,
    }
    return render(request, 'pos/pending_purchases_list.html', context)


# views.py
# views.py - Update edit_pending_purchase view

@login_required
def edit_pending_purchase(request, pk):
    """Edit a pending purchase with batch management"""
    pending_purchase = get_object_or_404(PendingPurchase, pk=pk, user=request.user, status='pending')
    
    if request.method == 'POST':
        try:
            # Parse JSON data
            data = json.loads(request.body)
            
            with transaction.atomic():
                # Validate supplier
                supplier_id = data.get('supplier')
                if not supplier_id:
                    return JsonResponse({
                        'success': False,
                        'message': 'Supplier is required'
                    }, status=400)
                
                try:
                    supplier = Supplier.objects.get(id=supplier_id)
                except Supplier.DoesNotExist:
                    return JsonResponse({
                        'success': False,
                        'message': 'Supplier not found'
                    }, status=404)
                
                # Validate items
                if 'items' not in data or not data['items']:
                    return JsonResponse({
                        'success': False,
                        'message': 'At least one item is required'
                    }, status=400)
                
                # Process items with batch information
                items_data = []
                subtotal = Decimal('0.00')
                
                for item in data['items']:
                    try:
                        product = Product.objects.get(id=item['product_id'])
                        quantity = Decimal(str(item['quantity'])).quantize(Decimal('0.00001'))
                        price = Decimal(str(item['price']))
                        total = quantity * price
                        
                        # Store batch information if provided
                        batch_number = item.get('batch_number', '').strip()
                        expiry_date = item.get('expiry_date', '')
                        
                        items_data.append({
                            'product_id': product.id,
                            'product_name': product.name,
                            'product_barcode': product.barcode,
                            'quantity': float(quantity),
                            'price': float(price),
                            'batch_number': batch_number,
                            'expiry_date': expiry_date,
                            'selling_price': float(product.selling_price),
                            'wholesale_price': float(product.wholesale_price),
                            'total': float(total)
                        })
                        
                        subtotal += total
                    except Product.DoesNotExist:
                        return JsonResponse({
                            'success': False,
                            'message': f'Product with ID {item["product_id"]} not found'
                        }, status=404)
                    except (KeyError, ValueError) as e:
                        return JsonResponse({
                            'success': False,
                            'message': f'Invalid item data: {str(e)}'
                        }, status=400)
                
                # Update pending purchase with batch-aware data
                pending_purchase.supplier = supplier
                pending_purchase.invoice_number = data.get('invoice_number', '')
                pending_purchase.data = {
                    'supplier_id': supplier.id,
                    'supplier_name': supplier.name,
                    'invoice_number': data.get('invoice_number', ''),
                    'items': items_data,
                    'subtotal': float(subtotal),
                    'total': float(subtotal),
                    'is_paid': data.get('is_paid', False),
                    'payment_method': data.get('payment_method', ''),
                    'notes': data.get('notes', ''),
                    'batch_managed': True  # Flag to indicate this uses batch management
                }
                pending_purchase.save()
                
                # Log activity
                UserActivityLog.objects.create(
                    user=request.user,
                    action_type='update',
                    model_name='PendingPurchase',
                    object_id=str(pending_purchase.id),
                    description=f'Updated pending purchase for {supplier.name} with {len(items_data)} items (batch managed)',
                    ip_address=request.META.get('REMOTE_ADDR'),
                    user_agent=request.META.get('HTTP_USER_AGENT', '')
                )
                
                return JsonResponse({
                    'success': True,
                    'message': 'Pending purchase updated successfully with batch information',
                    'redirect_url': reverse('view_pending_purchase', args=[pending_purchase.id])
                })
                
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'message': 'Invalid JSON data'
            }, status=400)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error updating pending purchase: {str(e)}'
            }, status=500)
    
    # GET request - prepare data for the form
    suppliers = Supplier.objects.all().order_by('name')
    products = Product.objects.filter(is_active=True).order_by('name')
    
    # Prepare items data with batch information
    items = pending_purchase.data.get('items', [])
    
    # Enhance items with additional product info if needed
    enhanced_items = []
    for item in items:
        try:
            product = Product.objects.get(id=item.get('product_id'))
            enhanced_items.append({
                'product_id': item.get('product_id'),
                'product_name': item.get('product_name', product.name),
                'product_barcode': product.barcode,
                'quantity': item.get('quantity', 0),
                'price': item.get('price', 0),
                'batch_number': item.get('batch_number', ''),
                'expiry_date': item.get('expiry_date', ''),
                'selling_price': item.get('selling_price', float(product.selling_price)),
            })
        except Product.DoesNotExist:
            # Keep original item data if product no longer exists
            enhanced_items.append(item)
    
    context = {
        'pending_purchase': pending_purchase,
        'suppliers': suppliers,
        'products': products,
        'items_json': json.dumps(enhanced_items),
        'selected_supplier': pending_purchase.supplier,
        'invoice_number': pending_purchase.data.get('invoice_number', ''),
        'is_paid': pending_purchase.data.get('is_paid', False),
        'payment_method': pending_purchase.data.get('payment_method', ''),
        'notes': pending_purchase.data.get('notes', ''),
        'subtotal': pending_purchase.data.get('subtotal', 0),
        'total': pending_purchase.data.get('total', 0),
        'is_batch_managed': pending_purchase.data.get('batch_managed', False),
    }
    
    return render(request, 'pos/edit_pending_purchase.html', context)



@login_required
@login_required
@require_POST
@login_required
@require_POST
# views.py - Update complete_pending_purchase view

@login_required
@require_POST
def complete_pending_purchase(request, pk):
    """Convert a pending purchase to a completed purchase with batch management"""
    try:
        # Get the pending purchase
        pending_purchase = PendingPurchase.objects.get(
            pk=pk, 
            user=request.user,
            status='pending'
        )
        
        with transaction.atomic():
            data = pending_purchase.data
            
            # Check if this is batch managed
            is_batch_managed = data.get('batch_managed', False)
            
            # Calculate ROUNDED totals from pending items
            rounded_total = Decimal('0.00')
            items_list = data.get('items', [])
            
            for item_data in items_list:
                quantity = Decimal(str(item_data.get('quantity', 0))).quantize(Decimal('0.00001'))
                price = Decimal(str(item_data.get('price', 0)))
                item_total = quantity * price
                rounded_item_total = round_to_nearest(item_total)
                rounded_total += rounded_item_total
            
            # Generate unique invoice number
            def generate_invoice_number():
                date_str = timezone.now().strftime('%Y%m%d')
                last_invoice = Purchase.objects.filter(
                    invoice_number__startswith=f"INV-{date_str}-"
                ).order_by('-invoice_number').first()
                
                if last_invoice:
                    try:
                        last_num = int(last_invoice.invoice_number.split('-')[-1])
                        next_num = last_num + 1
                    except:
                        next_num = 1
                else:
                    next_num = 1
                
                return f"INV-{date_str}-{next_num:04d}"
            
            # Get original invoice number from pending purchase or generate new
            pending_invoice = pending_purchase.invoice_number or ''
            if pending_invoice.startswith('PEND-') or Purchase.objects.filter(invoice_number=pending_invoice).exists():
                invoice_number = generate_invoice_number()
            else:
                invoice_number = pending_invoice
            
            # Create the actual purchase with ROUNDED total
            purchase = Purchase.objects.create(
                supplier=pending_purchase.supplier,
                invoice_number=invoice_number,
                user=request.user,
                subtotal=rounded_total,
                total=rounded_total,
                tax=Decimal('0.0000'),
                discount=Decimal('0.0000'),
                is_paid=data.get('is_paid', False),
                payment_method=data.get('payment_method', 'cash'),
                notes=data.get('notes', '')
            )
            
            # Create purchase items and batches with proper tracking
            item_count = 0
            created_batches = []
            
            for item_data in items_list:
                product_id = item_data.get('product_id')
                quantity = Decimal(str(item_data.get('quantity', 0))).quantize(Decimal('0.00001'))
                price = Decimal(str(item_data.get('price', 0)))
                batch_number = item_data.get('batch_number', '').strip()
                expiry_date = item_data.get('expiry_date', '')
                selling_price = Decimal(str(item_data.get('selling_price', 0))) if item_data.get('selling_price') else None
                
                if product_id and quantity > 0:
                    try:
                        product = Product.objects.get(id=product_id)
                        item_total = quantity * price
                        
                        # Create purchase item
                        purchase_item = PurchaseItem.objects.create(
                            purchase=purchase,
                            product=product,
                            quantity=quantity,
                            price=price,
                            total=item_total,
                            remaining_quantity=quantity,
                            is_fully_sold=False
                        )
                        
                        # Handle batch creation/updating
                        if batch_number and batch_number.strip():
                            # Check if batch already exists for this product
                            existing_batch = Batch.objects.filter(
                                product=product, 
                                batch_number=batch_number.strip()
                            ).first()
                            
                            if existing_batch:
                                batch = existing_batch
                                batch.quantity += quantity
                                batch.purchase_price = price
                                if selling_price:
                                    batch.selling_price = selling_price
                                if expiry_date:
                                    batch.expiry_date = expiry_date
                                batch.save()
                            else:
                                # Create new batch
                                batch = Batch.objects.create(
                                    product=product,
                                    batch_number=batch_number.strip(),
                                    quantity=quantity,
                                    expiry_date=expiry_date if expiry_date else None,
                                    purchase_price=price,
                                    selling_price=selling_price or product.selling_price,
                                    wholesale_price=product.wholesale_price,
                                    date_received=timezone.now().date(),
                                    purchase=purchase,
                                    purchase_item=purchase_item,
                                    is_active=True
                                )
                            
                            # Link batch to purchase item
                            purchase_item.batch = batch
                            purchase_item.save()
                            created_batches.append(batch)
                        
                        # Update product stock
                        product.quantity += quantity
                        product.save()
                        
                        item_count += 1
                        
                    except Product.DoesNotExist:
                        # Skip products that no longer exist
                        continue
            
            # Update purchase item count
            purchase.item_count = item_count
            purchase.save()
            
            # Log activity with batch info
            batch_count = len(created_batches)
            UserActivityLog.objects.create(
                user=request.user,
                action_type='create',
                model_name='Purchase',
                object_id=str(purchase.id),
                description=f'Completed pending purchase #{purchase.invoice_number} with {item_count} items and {batch_count} batches',
                ip_address=request.META.get('REMOTE_ADDR'),
                user_agent=request.META.get('HTTP_USER_AGENT', '')
            )
            
            # Delete pending purchase
            pending_purchase.delete()
            
            # Return JSON response for AJAX
            return JsonResponse({
                'success': True,
                'message': f'Purchase #{purchase.invoice_number} completed successfully with batch tracking!',
                'purchase_id': purchase.id,
                'redirect_url': reverse('view_purchase', args=[purchase.id]),
                'batch_count': batch_count,
                'item_count': item_count
            })
            
    except PendingPurchase.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Pending purchase not found or already completed.'
        }, status=404)
    
    except IntegrityError as e:
        if 'invoice_number' in str(e):
            return JsonResponse({
                'success': False,
                'message': 'Invoice number already exists. Please try again.'
            }, status=400)
        return JsonResponse({
            'success': False,
            'message': f'Database error: {str(e)}'
        }, status=500)
        
    except Exception as e:
        # Log the full error for debugging
        print(f"Error completing purchase: {str(e)}")
        import traceback
        traceback.print_exc()
        
        return JsonResponse({
            'success': False,
            'message': f'Error completing purchase: {str(e)}'
        }, status=500)

@login_required
@require_POST
def delete_pending_purchase(request, pk):
    """Delete a pending purchase"""
    pending_purchase = get_object_or_404(PendingPurchase, pk=pk, user=request.user, status='pending')
    
    try:
        pending_purchase.delete()
        messages.success(request, 'Pending purchase deleted successfully!')
    except Exception as e:
        messages.error(request, f'Error deleting pending purchase: {str(e)}')
    
    return redirect('pending_purchases_list')


@login_required
def pending_purchases(request):
    pending_purchases = PendingPurchase.objects.filter(status='pending').order_by('-created_at')
    
    # Pagination
    page = request.GET.get('page', 1)
    paginator = Paginator(pending_purchases, 25)
    
    try:
        pending_purchases = paginator.page(page)
    except PageNotAnInteger:
        pending_purchases = paginator.page(1)
    except EmptyPage:
        pending_purchases = paginator.page(paginator.num_pages)
    
    context = {
        'pending_purchases': pending_purchases,
    }
    return render(request, 'pos/pending_purchases.html', context)

@login_required
def view_pending_purchase(request, pk):
    """View details of a specific pending purchase"""
    pending_purchase = get_object_or_404(PendingPurchase, pk=pk, user=request.user)
    
    # Get product details for display
    product_ids = [item.get('product_id') for item in pending_purchase.data.get('items', [])]
    products = Product.objects.filter(id__in=product_ids).in_bulk()
    
    # Enhance items with product details
    items = []
    for item in pending_purchase.data.get('items', []):
        product = products.get(int(item.get('product_id'))) if item.get('product_id') else None
        if product:
            items.append({
                'product': product,
                'quantity': item.get('quantity', 0),
                'price': item.get('price', 0),
                'batch_number': item.get('batch_number', ''),
                'expiry_date': item.get('expiry_date', ''),
                'total': Decimal(item.get('quantity', 0)) * Decimal(item.get('price', 0))
            })
    
    context = {
        'pending_purchase': pending_purchase,
        'items': items,
        'subtotal': pending_purchase.data.get('subtotal', 0),
        'total': pending_purchase.data.get('total', 0),
    }
    return render(request, 'pos/view_pending_purchase.html', context)




@login_required
def get_purchase_items_api(request, purchase_id):
    # API endpoint to get purchase items in JSON format
    purchase = get_object_or_404(Purchase, id=purchase_id)
    items = purchase.items.all().values(
        'id', 'product__id', 'product__name', 'quantity', 
        'price', 'total', 'batch__batch_number', 'batch__expiry_date'
    )
    return JsonResponse(list(items), safe=False)

@login_required
def export_purchases(request):
    # Export purchases to CSV
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="purchases_{}.csv"'.format(
        datetime.now().strftime('%Y%m%d_%H%M%S')
    )
    
    writer = csv.writer(response)
    writer.writerow([
        'Invoice Number', 'Date', 'Supplier', 'Items Count', 
        'Total Amount', 'Payment Status', 'Created By'
    ])
    
    purchases = Purchase.objects.filter(is_return=False).select_related('supplier', 'created_by')
    for purchase in purchases:
        writer.writerow([
            purchase.invoice_number,
            purchase.date.strftime('%Y-%m-%d %H:%M'),
            purchase.supplier.name,
            purchase.item_count,
            purchase.total,
            'Paid' if purchase.is_paid else 'Unpaid',
            purchase.created_by.get_full_name() or purchase.created_by.username
        ])
    
    return response

@login_required
def export_pending_purchases(request):
    # Export pending purchases to CSV
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="pending_purchases_{}.csv"'.format(
        datetime.now().strftime('%Y%m%d_%H%M%S')
    )
    
    writer = csv.writer(response)
    writer.writerow([
        'ID', 'Supplier', 'Invoice Number', 'Items Count', 
        'Total Amount', 'Created At', 'Status', 'Created By'
    ])
    
    pending_purchases = PendingPurchase.objects.select_related('supplier', 'user')
    for pp in pending_purchases:
        writer.writerow([
            pp.id,
            pp.supplier.name,
            pp.data.get('invoice_number', ''),
            len(pp.data.get('items', [])),
            pp.data.get('total', '0.00'),
            pp.created_at.strftime('%Y-%m-%d %H:%M'),
            pp.get_status_display(),
            pp.user.get_full_name() or pp.user.username
        ])
    
    return response

@login_required
def export_supplier_returns(request):
    # Export supplier returns to CSV
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="supplier_returns_{}.csv"'.format(
        datetime.now().strftime('%Y%m%d_%H%M%S')
    )
    
    writer = csv.writer(response)
    writer.writerow([
        'Return Number', 'Date', 'Original Invoice', 'Supplier', 
        'Items Count', 'Total Amount', 'Status', 'Processed By'
    ])
    
    returns = Purchase.objects.filter(is_return=True).select_related('supplier', 'processed_by')
    for ret in returns:
        writer.writerow([
            ret.invoice_number,
            ret.date.strftime('%Y-%m-%d %H:%M'),
            ret.original_invoice,
            ret.supplier.name,
            ret.item_count,
            ret.total,
            ret.get_return_status_display(),
            ret.processed_by.get_full_name() if ret.processed_by else ''
        ])
    
    return response

@login_required
def cancel_pending_purchase(request, pk):
    pending_purchase = get_object_or_404(PendingPurchase, pk=pk, status='pending')
    
    try:
        pending_purchase.status = 'canceled'
        pending_purchase.save()
        messages.success(request, 'Pending purchase canceled successfully')
    except Exception as e:
        messages.error(request, f'Error canceling pending purchase: {str(e)}')
    
    return redirect('pending_purchases')


@login_required
def search_suppliers(request):
    search_term = request.GET.get('search', '').strip()
    
    if len(search_term) < 2:
        return JsonResponse([], safe=False)
    
    suppliers = Supplier.objects.filter(
        Q(name__icontains=search_term) |
        Q(contact_person__icontains=search_term) |
        Q(phone__icontains=search_term) |
        Q(email__icontains=search_term)
    ).distinct()[:10]
    
    results = [{
        'id': s.id,
        'text': f"{s.name} - {s.contact_person or ''} ({s.phone or ''})"
    } for s in suppliers]
    
    return JsonResponse(results, safe=False)


from django.views.decorators.http import require_GET
import json

from django.http import JsonResponse
from django.views.decorators.http import require_GET
from django.db.models import Q
import json

@login_required


@login_required
def view_purchase(request, pk):
    purchase = get_object_or_404(Purchase, pk=pk)
    return render(request, 'pos/view_purchase.html', {'purchase': purchase})

# views.py - Update record_payment view
@login_required
def record_payment(request, pk):
    purchase = get_object_or_404(Purchase, pk=pk)
    
    # Get payment details
    total_paid = purchase.total_paid
    balance_due = purchase.balance_due
    
    if request.method == 'POST':
        try:
            amount = Decimal(request.POST.get('amount', '0.00'))
            payment_method = request.POST.get('payment_method')
            reference = request.POST.get('reference', '')
            payment_date = request.POST.get('date') or timezone.now().date()
            notes = request.POST.get('notes', '')
            
            # Validate amount
            if amount <= 0:
                messages.error(request, 'Amount must be greater than 0')
            elif amount > (balance_due + purchase.supplier.balance):
                messages.error(request, f'Amount cannot exceed total due of KSh {(balance_due + purchase.supplier.balance):.2f}')
            else:
                with transaction.atomic():
                    # Check if we should also settle supplier balance
                    supplier = purchase.supplier
                    
                    if amount > balance_due:
                        # Pay more than purchase balance - apply to supplier balance
                        extra_amount = amount - balance_due
                        
                        if extra_amount > supplier.balance:
                            extra_amount = supplier.balance
                        
                        # Pay purchase balance first
                        purchase_payment = SupplierPayment.objects.create(
                            supplier=supplier,
                            purchase=purchase,
                            amount=balance_due,
                            date=payment_date,
                            payment_method=payment_method,
                            reference=reference,
                            user=request.user,
                            notes=notes,
                            is_partial=(balance_due < purchase.total),
                            remaining_balance=Decimal('0.0000')
                        )
                        
                        # Pay supplier balance with remaining
                        if extra_amount > 0:
                            SupplierPayment.objects.create(
                                supplier=supplier,
                                purchase=None,  # No specific purchase
                                amount=extra_amount,
                                date=payment_date,
                                payment_method=payment_method,
                                reference=f"Balance payment - {reference}",
                                user=request.user,
                                notes=f"Settled supplier balance - {notes}",
                                is_partial=False,
                                remaining_balance=supplier.balance - extra_amount
                            )
                            
                            # Update supplier balance
                            supplier.balance -= extra_amount
                            supplier.save()
                    
                    else:
                        # Pay only purchase amount (partial or full)
                        is_partial_payment = amount < balance_due
                        remaining_balance = balance_due - amount
                        
                        payment = SupplierPayment.objects.create(
                            supplier=supplier,
                            purchase=purchase,
                            amount=amount,
                            date=payment_date,
                            payment_method=payment_method,
                            reference=reference,
                            user=request.user,
                            notes=notes,
                            is_partial=is_partial_payment,
                            remaining_balance=remaining_balance
                        )
                    
                    # Update purchase payment status
                    purchase.update_payment_status()
                    
                    # Check if we need to update supplier balance (if payment is for specific purchase only)
                    if amount <= balance_due:
                        # This payment doesn't affect supplier balance directly
                        pass
                    
                    messages.success(request, f'Payment of KSh {amount:.2f} recorded successfully!')
                    return redirect('view_purchase', pk=purchase.id)
                    
        except Exception as e:
            messages.error(request, f'Error processing payment: {str(e)}')
    
    # Get supplier balance
    supplier_balance = purchase.supplier.balance
    
    context = {
        'purchase': purchase,
        'total_paid': total_paid,
        'balance_due': balance_due,
        'supplier_balance': supplier_balance,
        'total_due': balance_due + supplier_balance,  # Combined due
    }
    return render(request, 'pos/record_payment.html', context)


# views.py - Add supplier_payment_summary view
@login_required
def supplier_payment_summary(request, supplier_id=None):
    """View all payments for a supplier with balances"""
    suppliers = Supplier.objects.all().order_by('name')
    
    # Get specific supplier or all
    if supplier_id:
        supplier = get_object_or_404(Supplier, id=supplier_id)
        payments = SupplierPayment.objects.filter(supplier=supplier).order_by('-date')
        purchases = Purchase.objects.filter(supplier=supplier).order_by('-date')
    else:
        supplier = None
        payments = SupplierPayment.objects.all().order_by('-date')
        purchases = Purchase.objects.none()
    
    # Date filtering
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if start_date:
        payments = payments.filter(date__gte=start_date)
    if end_date:
        payments = payments.filter(date__lte=end_date)
    
    # Calculate summary
    total_payments = payments.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    
    # Calculate supplier balances
    supplier_balances = []
    for s in suppliers:
        supplier_payments = SupplierPayment.objects.filter(supplier=s)
        total_paid = supplier_payments.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        
        # Get purchases total
        supplier_purchases = Purchase.objects.filter(supplier=s, is_return=False)
        total_purchases = supplier_purchases.aggregate(total=Sum('total'))['total'] or Decimal('0.00')
        
        # Calculate balance (purchases - payments)
        balance = total_purchases - total_paid
        
        supplier_balances.append({
            'supplier': s,
            'total_purchases': total_purchases,
            'total_payments': total_paid,
            'balance': balance,
            'payment_status': 'Paid' if balance <= 0 else 'Partial' if total_paid > 0 else 'Unpaid'
        })
    
    context = {
        'supplier': supplier,
        'suppliers': suppliers,
        'payments': payments,
        'purchases': purchases,
        'supplier_balances': supplier_balances,
        'total_payments': total_payments,
        'start_date': start_date,
        'end_date': end_date,
        'selected_supplier': supplier_id,
    }
    
    return render(request, 'pos/supplier_payment_summary.html', context)

def generate_purchases_export(format, queryset):
    if format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="purchases_report.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            'Invoice', 'Date', 'Supplier', 'Items', 
            'Subtotal', 'Tax', 'Discount', 'Total',
            'Payment Status', 'Is Return'
        ])
        
        for purchase in queryset:
            writer.writerow([
                purchase.invoice_number,
                purchase.date.strftime('%Y-%m-%d %H:%M'),
                purchase.supplier.name,
                purchase.items.aggregate(total=Sum('quantity'))['total'] or 0,
                purchase.subtotal,
                purchase.tax,
                purchase.discount,
                purchase.total,
                'Paid' if purchase.is_paid else 'Unpaid',
                'Yes' if purchase.is_return else 'No'
            ])
        
        return response
    
    return HttpResponse('Invalid export format', status=400)

@login_required
def create_purchase_return(request):
    if request.method == 'POST':
        form = PurchaseReturnForm(request.POST)
        if form.is_valid():
            return_purchase = form.save(commit=False)
            return_purchase.is_return = True
            return_purchase.user = request.user
            return_purchase.save()
            return redirect('purchase_list')
    else:
        form = PurchaseReturnForm()
    
    # Get all purchases that can be returned (not already returns)
    purchases = Purchase.objects.filter(is_return=False).order_by('-date')
    
    return render(request, 'pos/create_purchase_return.html', {
        'form': form,
        'purchases': purchases
    })


# ============== Supplier Views ==============
@login_required
def supplier_list(request):
    suppliers = Supplier.objects.all().order_by('name')
    
    # Search functionality
    search_query = request.GET.get('search')
    if search_query:
        suppliers = suppliers.filter(
            Q(name__icontains=search_query) |
            Q(contact_person__icontains=search_query) |
            Q(phone__icontains=search_query))
    
    # Filter by balance status
    balance_filter = request.GET.get('balance')
    if balance_filter == 'credit':
        suppliers = suppliers.filter(balance__gt=0)
    elif balance_filter == 'zero':
        suppliers = suppliers.filter(balance=0)
    
    # Export functionality
    export_format = request.GET.get('export')
    if export_format:
        return generate_suppliers_export(export_format, suppliers)
    
    context = {
        'suppliers': suppliers,
        'search_query': search_query,
        'balance_filter': balance_filter
    }
    

    supplier_purchases = Purchase.objects.filter(
        is_return=False
    ).values(
        'supplier__name'
    ).annotate(
        total_spent=Sum('total'),
        order_count=Count('id'),
        last_order=Max('date')
    ).order_by('-total_spent')
    
    context.update({
        'supplier_purchases': supplier_purchases
    })
    return render(request, 'pos/supplier_list.html', context)

@login_required
def add_supplier(request):
    if request.method == 'POST':
        form = SupplierForm(request.POST)
        if form.is_valid():
            supplier = form.save()
            messages.success(request, f'Supplier "{supplier.name}" added successfully!')
            return redirect('supplier_list')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = SupplierForm()
    
    context = {
        'form': form,
        'title': 'Add New Supplier'
    }
    return render(request, 'pos/add_supplier.html', context)
@login_required
def edit_supplier(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)
    
    if request.method == 'POST':
        form = SupplierForm(request.POST, instance=supplier)
        if form.is_valid():
            form.save()
            return redirect('supplier_list')
    else:
        form = SupplierForm(instance=supplier)
    
    context = {
        'form': form,
        'supplier': supplier
    }
    return render(request, 'pos/edit_supplier.html', context)

@login_required
def supplier_detail(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)
    purchases = Purchase.objects.filter(supplier=supplier).order_by('-date')
    payments = SupplierPayment.objects.filter(supplier=supplier).order_by('-date')
    returns = Return.objects.filter(supplier=supplier).order_by('-date')
    
    # Pagination for purchases and payments
    purchase_paginator = Paginator(purchases, 10)
    purchase_page = request.GET.get('purchase_page')
    purchase_page_obj = purchase_paginator.get_page(purchase_page)
    
    payment_paginator = Paginator(payments, 10)
    payment_page = request.GET.get('payment_page')
    payment_page_obj = payment_paginator.get_page(payment_page)
    
    return_paginator = Paginator(returns, 10)
    return_page = request.GET.get('return_page')
    return_page_obj = return_paginator.get_page(return_page)
    
    context = {
        'supplier': supplier,
        'purchases': purchase_page_obj,
        'payments': payment_page_obj,
        'returns': return_page_obj
    }
    return render(request, 'pos/supplier_detail.html', context)

def generate_suppliers_export(format, queryset):
    if format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="suppliers.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            'Name', 'Contact Person', 'Phone', 'Email', 
            'Address', 'Balance'
        ])
        
        for supplier in queryset:
            writer.writerow([
                supplier.name,
                supplier.contact_person,
                supplier.phone,
                supplier.email,
                supplier.address,
                supplier.balance
            ])
        
        return response
    
    return HttpResponse('Invalid export format', status=400)

@login_required
def customer_payment_report(request):
    # Get filter parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    search_query = request.GET.get('search', '')
    customer_id = request.GET.get('customer')
    
    # Start with all payments
    payments = CustomerPayment.objects.all().select_related('customer', 'sale')
    
    # Apply date filters only if dates are provided
    if start_date:
        payments = payments.filter(date__gte=start_date)
    if end_date:
        payments = payments.filter(date__lte=end_date)
    
    # Apply customer filter if provided
    if customer_id:
        payments = payments.filter(customer_id=customer_id)
    
    # Apply search filter if provided
    if search_query:
        payments = payments.filter(
            Q(customer__name__icontains=search_query) |
            Q(sale__sale_number__icontains=search_query) |
            Q(reference__icontains=search_query)
        )
    
    # Group by payment method
    by_method = payments.values('payment_method').annotate(
        total=Sum('amount'),
        count=Count('id')
    )
    
    # Group by customer
    by_customer = payments.values('customer__name').annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('-total')
    
    # Calculate total amount and average payment
    total_amount = payments.aggregate(total=Sum('amount'))['total'] or 0
    average_payment = total_amount / len(payments) if payments.exists() else 0
    
    # Get all customers for filter dropdown
    customers = Customer.objects.all().order_by('name')
    
    context = {
        'payments': payments,
        'by_method': by_method,
        'by_customer': by_customer,
        'start_date': start_date,
        'end_date': end_date,
        'search_query': search_query,
        'total_amount': total_amount,
        'average_payment': average_payment,
        'customers': customers,  # Add customers to context
        'selected_customer': customer_id,  # Add selected customer
    }
    return render(request, 'pos/customer_payment_report.html', context)

@login_required
def supplier_payments_report(request):
    # Date filtering
    start_date = request.GET.get('start_date') or (timezone.now() - timedelta(days=90)).date()
    end_date = request.GET.get('end_date') or timezone.now().date()
    
    # Get payments with filters
    payments = SupplierPayment.objects.filter(
        date__range=[start_date, end_date]
    ).select_related('supplier', 'purchase', 'user')
    
    # Apply additional filters
    supplier_id = request.GET.get('supplier')
    if supplier_id:
        payments = payments.filter(supplier_id=supplier_id)
    
    payment_method = request.GET.get('payment_method')
    if payment_method:
        payments = payments.filter(payment_method=payment_method)
    
    # Format payments data
    formatted_payments = []
    for payment in payments:
        formatted_payments.append({
            'id': payment.id,
            'date': payment.date,
            'supplier': payment.supplier,
            'purchase': payment.purchase,
            'payment_method': payment.payment_method,
            'amount': format_currency(payment.amount),  # Format with 4 decimal places
            'user': payment.user,
            'notes': payment.notes,
            'reference': payment.reference,
            'raw_amount': payment.amount,  # Keep raw for calculations
        })
    
    # Calculate totals
    totals = {
        'total_amount': payments.aggregate(total=Sum('amount'))['total'] or Decimal('0.0000'),
        'total_count': payments.count(),
        'avg_amount': payments.aggregate(avg=Avg('amount'))['avg'] or Decimal('0.0000'),
    }
    
    # Get payment method summary
    method_summary = payments.values('payment_method').annotate(
        total_amount=Sum('amount'),
        count=Count('id')
    ).order_by('-total_amount')
    
    # Format method summary
    formatted_method_summary = []
    for method in method_summary:
        formatted_method_summary.append({
            'payment_method': method['payment_method'],
            'total_amount': format_currency(method['total_amount'] or Decimal('0.0000')),
            'count': method['count'],
            'raw_total_amount': method['total_amount'] or Decimal('0.0000'),
        })
    
    # Get supplier summary
    supplier_summary = payments.values('supplier__name').annotate(
        total_payments=Sum('amount'),
        payment_count=Count('id')
    ).order_by('-total_payments')
    
    # Format supplier summary
    formatted_supplier_summary = []
    for supplier in supplier_summary:
        formatted_supplier_summary.append({
            'supplier__name': supplier['supplier__name'],
            'total_payments': format_currency(supplier['total_payments'] or Decimal('0.0000')),
            'payment_count': supplier['payment_count'],
        })
    
    # Get all suppliers for filter dropdown
    suppliers = Supplier.objects.all()
    
    # Payment method choices
    payment_methods = SupplierPayment.PAYMENT_METHODS
    
    context = {
        'payments': formatted_payments,
        'totals': {
            'total_amount': format_currency(totals['total_amount']),
            'total_count': totals['total_count'],
            'avg_amount': format_currency(totals['avg_amount']),
        },
        'method_summary': formatted_method_summary,
        'supplier_summary': formatted_supplier_summary,
        'suppliers': suppliers,
        'payment_methods': payment_methods,
        'start_date': start_date,
        'end_date': end_date,
        'selected_supplier': supplier_id,
        'selected_payment_method': payment_method,
        'today': timezone.now().date(),
    }
    
    return render(request, 'pos/reports/supplier_payments_report.html', context)


# views.py - Add these imports
from django.db.models import (
    Sum, Count, Avg, Max, Min, F, Q, ExpressionWrapper, 
    DecimalField, Value, Case, When, CharField, IntegerField
)
from django.db.models.functions import TruncDate, TruncMonth, TruncYear, Coalesce, Extract
from datetime import datetime, timedelta
import pandas as pd
from io import BytesIO
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')  # Use non-interactive backend
import base64

# ============== REPORTS MODULE ==============

@login_required
def reports_dashboard(request):
    """Main reports dashboard"""
    context = {
        'today': timezone.now().date(),
        'reports': [
            {'name': 'Sales by Product', 'url': 'sales_by_product_report', 'icon': 'fa-chart-bar', 'color': 'blue'},
            {'name': 'Purchase by Product', 'url': 'purchase_by_product_report', 'icon': 'fa-shopping-cart', 'color': 'green'},
            {'name': 'Daily Sales Report', 'url': 'daily_sales_report', 'icon': 'fa-calendar-day', 'color': 'orange'},
            {'name': 'Sales Report', 'url': 'sales_report', 'icon': 'fa-file-invoice-dollar', 'color': 'purple'},
            {'name': 'Profit by Product', 'url': 'profit_by_product_report', 'icon': 'fa-money-bill-wave', 'color': 'teal'},
            {'name': 'Profit Margin Report', 'url': 'profit_margin_report', 'icon': 'fa-percentage', 'color': 'red'},
            {'name': 'Customer Payments', 'url': 'customer_payment_report', 'icon': 'fa-user-check', 'color': 'indigo'},
            {'name': 'Supplier Payments', 'url': 'supplier_payment_report', 'icon': 'fa-truck', 'color': 'pink'},
            {'name': 'Daily Opening Stock', 'url': 'daily_opening_stock_report', 'icon': 'fa-box-open', 'color': 'yellow'},
            {'name': 'Inventory Valuation', 'url': 'stock_value_report', 'icon': 'fa-coins', 'color': 'cyan'},
            {'name': 'Top Selling Products', 'url': 'top_selling_products_report', 'icon': 'fa-star', 'color': 'amber'},
            {'name': 'Slow Moving Products', 'url': 'slow_moving_products_report', 'icon': 'fa-clock', 'color': 'gray'},
            {'name': 'Customer Sales Analysis', 'url': 'customer_sales_analysis', 'icon': 'fa-users', 'color': 'lime'},
            {'name': 'Supplier Purchase Analysis', 'url': 'supplier_purchase_analysis', 'icon': 'fa-industry', 'color': 'deep-purple'},
            {'name': 'Weekly Sales & Profit Report', 'url': 'weekly_sales_profit_report', 'icon': 'fa-chart-bar', 'color': 'primary', 'description': 'All products sold with quantity, revenue & profits'},
            {'name': 'Product Performance Analysis', 'url': 'product_performance_report', 'icon': 'fa-chart-line', 'color': 'primary', 'description': 'Quantity, revenue & profit analysis'},
        ]
    }
    return render(request, 'pos/reports_dashboard.html', context)

# ============== SALES REPORTS ==============

@login_required
def sales_by_product_report(request):
    """Generate sales by product report"""
    try:
        # Get date range from request
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
        category_id = request.GET.get('category')
        
        # Parse dates
        if start_date_str and end_date_str:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        else:
            # Default to last 30 days
            end_date = timezone.now().date()
            start_date = end_date - timedelta(days=30)
            start_date_str = start_date.strftime('%Y-%m-%d')
            end_date_str = end_date.strftime('%Y-%m-%d')
        
        # Base query
        sale_items = SaleItem.objects.filter(
            sale__date__date__range=[start_date, end_date],
            sale__is_completed=True
        ).select_related('product', 'product__category')
        
        # Filter by category if selected
        if category_id:
            sale_items = sale_items.filter(product__category_id=category_id)
        
        # Group by product
        product_data = {}
        for item in sale_items:
            if item.product:
                product_id = item.product.id
                if product_id not in product_data:
                    product_data[product_id] = {
                        'product': item.product,
                        'total_quantity': Decimal('0'),
                        'total_sales': Decimal('0'),
                        'total_cost': Decimal('0'),
                        'sale_count': 0,
                    }
                
                product_data[product_id]['total_quantity'] += item.quantity
                product_data[product_id]['total_sales'] += item.total
                product_data[product_id]['total_cost'] += (item.product.purchase_price * item.quantity)
                product_data[product_id]['sale_count'] += 1
        
        # Calculate profit and margin
        product_list = []
        for data in product_data.values():
            revenue = data['total_sales']
            cost = data['total_cost']
            profit = revenue - cost
            
            # Calculate profit margin
            profit_margin = (profit / revenue * 100) if revenue > 0 else Decimal('0')
            
            product_list.append({
                'product__id': data['product'].id,
                'product__name': data['product'].name,
                'product__barcode': data['product'].barcode,
                'product__category__name': data['product'].category.name if data['product'].category else 'Uncategorized',
                'total_quantity': data['total_quantity'],
                'total_sales': revenue,
                'total_cost': cost,
                'total_profit': profit,
                'profit_margin': profit_margin,
                'sale_count': data['sale_count'],
            })
        
        # Sort by total sales
        product_list.sort(key=lambda x: x['total_sales'], reverse=True)
        
        # Calculate summary
        total_quantity = sum(item['total_quantity'] for item in product_list)
        total_sales = sum(item['total_sales'] for item in product_list)
        total_cost = sum(item['total_cost'] for item in product_list)
        total_profit = sum(item['total_profit'] for item in product_list)
        
        avg_profit_margin = (total_profit / total_sales * 100) if total_sales > 0 else Decimal('0')
        
        # Prepare chart data (top 10 products)
        chart_data = []
        for item in product_list[:10]:
            chart_data.append({
                'product': item['product__name'],
                'sales': float(item['total_sales']),
                'quantity': float(item['total_quantity']),
            })
        
        context = {
            'product_sales': product_list,
            'summary': {
                'total_products': len(product_list),
                'total_quantity': total_quantity,
                'total_sales': total_sales,
                'total_cost': total_cost,
                'total_profit': total_profit,
                'avg_profit_margin': avg_profit_margin,
            },
            'categories': Category.objects.all(),
            'selected_category': category_id,
            'start_date': start_date_str,
            'end_date': end_date_str,
            'chart_data': json.dumps(chart_data),
            'today': timezone.now().date(),
        }
        
        # Export logic
        export_format = request.GET.get('export')
        if export_format:
            return export_report(export_format, product_list, start_date_str, end_date_str, 'sales_by_product')
        
        return render(request, 'pos/reports/sales_by_product.html', context)
        
    except Exception as e:
        messages.error(request, f"Error generating report: {str(e)}")
        return redirect('reports_dashboard')

@login_required
def purchase_by_product_report(request):
    """Purchase analysis by product"""
    # Date filtering
    start_date = request.GET.get('start_date', (timezone.now() - timedelta(days=90)).strftime('%Y-%m-%d'))
    end_date = request.GET.get('end_date', timezone.now().strftime('%Y-%m-%d'))
    
    # Get purchases data
    purchases = Purchase.objects.filter(
        date__date__range=[start_date, end_date],
        is_return=False
    )
    
    # Apply additional filters
    supplier_id = request.GET.get('supplier')
    if supplier_id:
        purchases = purchases.filter(supplier_id=supplier_id)
    
    category_id = request.GET.get('category')
    if category_id:
        purchases = purchases.filter(items__product__category_id=category_id)
    
    # Get purchase items with product details
    purchase_items = PurchaseItem.objects.filter(
        purchase__in=purchases
    ).select_related('product', 'product__category', 'purchase__supplier')
    
    # Group by product
    product_purchases = purchase_items.values(
        'product__id',
        'product__name',
        'product__category__name',
        'purchase__supplier__name'
    ).annotate(
        total_quantity=Sum('quantity'),
        total_amount=Sum('total'),
        avg_price=Avg('price'),
        purchase_count=Count('purchase', distinct=True),
        supplier_count=Count('purchase__supplier', distinct=True)
    ).order_by('-total_amount')
    
    # Calculate summary statistics
    summary = product_purchases.aggregate(
        total_products=Count('product__id', distinct=True),
        total_quantity=Sum('total_quantity'),
        total_amount=Sum('total_amount'),
        avg_price=Avg('avg_price')
    )
    
    # Get filters data
    suppliers = Supplier.objects.all()
    categories = Category.objects.all()
    
    context = {
        'product_purchases': product_purchases,
        'summary': summary,
        'suppliers': suppliers,
        'categories': categories,
        'start_date': start_date,
        'end_date': end_date,
        'selected_supplier': supplier_id,
        'selected_category': category_id,
    }
    return render(request, 'pos/reports/purchase_by_product.html', context)

# ============== PAYMENT REPORTS ==============

@login_required
def customer_payment_report(request):
    """Customer payments report"""
    # Date filtering
    start_date = request.GET.get('start_date', (timezone.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
    end_date = request.GET.get('end_date', timezone.now().strftime('%Y-%m-%d'))
    
    # Get payments
    payments = CustomerPayment.objects.filter(
        date__range=[start_date, end_date]
    ).select_related('customer', 'sale', 'user')
    
    # Apply filters
    customer_id = request.GET.get('customer')
    if customer_id:
        payments = payments.filter(customer_id=customer_id)
    
    payment_method = request.GET.get('payment_method')
    if payment_method:
        payments = payments.filter(payment_method=payment_method)
    
    # Group by customer
    customer_summary = payments.values(
        'customer__name',
        'customer__phone'
    ).annotate(
        total_payments=Sum('amount'),
        payment_count=Count('id'),
        avg_payment=Avg('amount'),
        last_payment=Max('date')
    ).order_by('-total_payments')
    
    # Group by payment method
    method_summary = payments.values('payment_method').annotate(
        total_amount=Sum('amount'),
        count=Count('id')
    ).order_by('-total_amount')
    
    # Group by date (for trend)
    daily_summary = payments.annotate(
        payment_date=TruncDate('date')
    ).values('payment_date').annotate(
        total_amount=Sum('amount'),
        count=Count('id')
    ).order_by('payment_date')
    
    # Calculate totals
    totals = payments.aggregate(
        total_amount=Sum('amount'),
        total_count=Count('id'),
        avg_amount=Avg('amount')
    )
    
    # Get customers for filter
    customers = Customer.objects.all()
    
    context = {
        'payments': payments,
        'customer_summary': customer_summary,
        'method_summary': method_summary,
        'daily_summary': daily_summary,
        'totals': totals,
        'customers': customers,
        'payment_methods': CustomerPayment.PAYMENT_METHODS,
        'start_date': start_date,
        'end_date': end_date,
        'selected_customer': customer_id,
        'selected_payment_method': payment_method,
    }
    return render(request, 'pos/reports/customer_payments.html', context)

@login_required
def supplier_payment_report(request):
    """Supplier payments report"""
    # Date filtering
    start_date = request.GET.get('start_date', (timezone.now() - timedelta(days=90)).strftime('%Y-%m-%d'))
    end_date = request.GET.get('end_date', timezone.now().strftime('%Y-%m-%d'))
    
    # Get payments
    payments = SupplierPayment.objects.filter(
        date__range=[start_date, end_date]
    ).select_related('supplier', 'purchase', 'user')
    
    # Apply filters
    supplier_id = request.GET.get('supplier')
    if supplier_id:
        payments = payments.filter(supplier_id=supplier_id)
    
    payment_method = request.GET.get('payment_method')
    if payment_method:
        payments = payments.filter(payment_method=payment_method)
    
    # Group by supplier
    supplier_summary = payments.values(
        'supplier__name',
        'supplier__phone'
    ).annotate(
        total_payments=Sum('amount'),
        payment_count=Count('id'),
        avg_payment=Avg('amount'),
        last_payment=Max('date')
    ).order_by('-total_payments')
    
    # Group by payment method
    method_summary = payments.values('payment_method').annotate(
        total_amount=Sum('amount'),
        count=Count('id')
    ).order_by('-total_amount')
    
    # Calculate totals
    totals = payments.aggregate(
        total_amount=Sum('amount'),
        total_count=Count('id'),
        avg_amount=Avg('amount')
    )
    
    # Get suppliers for filter
    suppliers = Supplier.objects.all()
    
    context = {
        'payments': payments,
        'supplier_summary': supplier_summary,
        'method_summary': method_summary,
        'totals': totals,
        'suppliers': suppliers,
        'payment_methods': SupplierPayment.PAYMENT_METHODS,
        'start_date': start_date,
        'end_date': end_date,
        'selected_supplier': supplier_id,
        'selected_payment_method': payment_method,
    }
    return render(request, 'pos/reports/supplier_payments.html', context)

# ============== PROFIT REPORTS ==============

@login_required
def profit_by_product_report(request):
    """Profit analysis by product"""
    # Date filtering
    start_date = request.GET.get('start_date', (timezone.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
    end_date = request.GET.get('end_date', timezone.now().strftime('%Y-%m-%d'))
    
    # Get sales data
    sales = Sale.objects.filter(
        date__date__range=[start_date, end_date],
        is_completed=True
    )
    
    # Apply filters
    category_id = request.GET.get('category')
    if category_id:
        sales = sales.filter(items__product__category_id=category_id)
    
    # Get sale items with profit calculations
    sale_items = SaleItem.objects.filter(
        sale__in=sales
    ).select_related('product', 'product__category')
    
    # Calculate profit by product
    product_profits = []
    all_items = sale_items.select_related('product', 'product__category')
    
    # Group by product
    product_groups = {}
    for item in all_items:
        product_id = item.product.id
        if product_id not in product_groups:
            product_groups[product_id] = {
                'product': item.product,
                'category': item.product.category.name if item.product.category else 'Uncategorized',
                'total_quantity': Decimal('0'),
                'total_sales': Decimal('0'),
                'total_cost': Decimal('0'),
                'total_profit': Decimal('0'),
                'sale_count': 0
            }
        
        quantity = item.quantity
        sales_amount = item.total
        cost = item.product.purchase_price * quantity
        profit = sales_amount - cost
        
        product_groups[product_id]['total_quantity'] += quantity
        product_groups[product_id]['total_sales'] += sales_amount
        product_groups[product_id]['total_cost'] += cost
        product_groups[product_id]['total_profit'] += profit
        product_groups[product_id]['sale_count'] += 1
    
    # Convert to list and calculate margins
    for product_id, data in product_groups.items():
        profit_margin = (data['total_profit'] / data['total_sales'] * 100) if data['total_sales'] > 0 else 0
        avg_profit_per_unit = data['total_profit'] / data['total_quantity'] if data['total_quantity'] > 0 else 0
        
        product_profits.append({
            'product': data['product'],
            'category': data['category'],
            'total_quantity': data['total_quantity'],
            'total_sales': data['total_sales'],
            'total_cost': data['total_cost'],
            'total_profit': data['total_profit'],
            'profit_margin': profit_margin,
            'avg_profit_per_unit': avg_profit_per_unit,
            'sale_count': data['sale_count']
        })
    
    # Sort by total profit
    product_profits.sort(key=lambda x: x['total_profit'], reverse=True)
    
    # Calculate summary
    summary = {
        'total_products': len(product_profits),
        'total_sales': sum(p['total_sales'] for p in product_profits),
        'total_cost': sum(p['total_cost'] for p in product_profits),
        'total_profit': sum(p['total_profit'] for p in product_profits),
        'avg_profit_margin': sum(p['profit_margin'] for p in product_profits) / len(product_profits) if product_profits else 0
    }
    
    # Get categories for filter
    categories = Category.objects.all()
    
    context = {
        'product_profits': product_profits,
        'summary': summary,
        'categories': categories,
        'start_date': start_date,
        'end_date': end_date,
        'selected_category': category_id,
    }
    return render(request, 'pos/reports/profit_by_product.html', context)

# ============== DAILY REPORTS ==============

@login_required
def daily_opening_stock_report(request):
    """Generate daily opening stock report - Shows actual inventory available at START of selected day"""
    try:
        # Get selected date
        date_str = request.GET.get('date')
        if date_str:
            try:
                selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                selected_date = timezone.now().date()
                date_str = selected_date.strftime('%Y-%m-%d')
                messages.warning(request, "Invalid date format. Using today's date.")
        else:
            selected_date = timezone.now().date()
            date_str = selected_date.strftime('%Y-%m-%d')
        
        # IMPORTANT: For opening stock, we need to know what was available at START of day
        # This means we need to consider:
        # 1. Stock at end of previous day
        # 2. Or stock at beginning of selected day
        
        # Calculate the next day for date range queries
        next_day = selected_date + timedelta(days=1)
        
        # Get all active products
        products = Product.objects.filter(is_active=True).select_related('category', 'supplier')
        
        report_data = []
        total_quantity = Decimal('0')
        total_value = Decimal('0')
        total_products = 0
        low_stock_count = 0
        out_of_stock_count = 0
        
        for product in products:
            # METHOD 1: If you have StockJournal entries, use them to calculate opening stock
            # This is the most accurate method
            
            # First, get the initial quantity (when product was created)
            base_quantity = product.quantity
            
            # Get all stock movements BEFORE the selected date
            stock_movements_before = StockJournalItem.objects.filter(
                product=product,
                journal__date__date__lt=selected_date
            ).aggregate(
                total_in=Sum('quantity', filter=Q(movement_type='in')),
                total_out=Sum('quantity', filter=Q(movement_type__in=['out', 'missing', 'damaged', 'broken', 'expired'])),
                total_adjustment=Sum('quantity', filter=Q(movement_type='adjustment'))
            )
            
            # Get all sales BEFORE the selected date
            sales_before = SaleItem.objects.filter(
                product=product,
                sale__date__date__lt=selected_date,
                sale__is_completed=True
            ).aggregate(total_sold=Sum('quantity'))['total_sold'] or Decimal('0')
            
            # Get all purchases BEFORE the selected date
            purchases_before = PurchaseItem.objects.filter(
                product=product,
                purchase__date__date__lt=selected_date,
                purchase__is_return=False
            ).aggregate(total_purchased=Sum('quantity'))['total_purchased'] or Decimal('0')
            
            # Calculate OPENING STOCK (stock at beginning of selected day)
            # Formula: Opening Stock = Current Stock + Sales Today - Purchases Today - Adjustments Today
            # But we need to work backwards from current stock
            
            # Get today's movements (for calculation if needed)
            stock_movements_today = StockJournalItem.objects.filter(
                product=product,
                journal__date__date=selected_date
            ).aggregate(
                total_in=Sum('quantity', filter=Q(movement_type='in')),
                total_out=Sum('quantity', filter=Q(movement_type__in=['out', 'missing', 'damaged', 'broken', 'expired']))
            )
            
            # Get today's sales
            sales_today = SaleItem.objects.filter(
                product=product,
                sale__date__date=selected_date,
                sale__is_completed=True
            ).aggregate(total_sold=Sum('quantity'))['total_sold'] or Decimal('0')
            
            # Get today's purchases
            purchases_today = PurchaseItem.objects.filter(
                product=product,
                purchase__date__date=selected_date,
                purchase__is_return=False
            ).aggregate(total_purchased=Sum('quantity'))['total_purchased'] or Decimal('0')
            
            # CALCULATE ACTUAL OPENING STOCK:
            # Opening Stock = Current Stock + Sales Today - Purchases Today + Stock Out Today - Stock In Today
            # This works backwards from current stock
            
            stock_in_today = stock_movements_today['total_in'] or Decimal('0')
            stock_out_today = stock_movements_today['total_out'] or Decimal('0')
            
            opening_stock = (
                product.quantity  # Current stock
                + sales_today     # Add back what was sold today
                - purchases_today # Subtract what was purchased today
                + stock_out_today # Add back stock that moved out today (damaged, expired, etc.)
                - stock_in_today  # Subtract stock that moved in today
            )
            
            # Ensure opening stock is not negative
            opening_stock = max(opening_stock, Decimal('0'))
            
            # Calculate stock value using purchase price
            stock_value = opening_stock * product.purchase_price
            
            # Determine stock status based on opening stock
            if opening_stock <= 0:
                stock_status = 'out_of_stock'
                out_of_stock_count += 1
            elif opening_stock <= product.reorder_level:
                stock_status = 'low'
                low_stock_count += 1
            else:
                stock_status = 'normal'
            
            # Add to totals
            total_quantity += opening_stock
            total_value += stock_value
            total_products += 1
            
            # Prepare data for display
            report_data.append({
                'product': product,
                'opening_stock': opening_stock,  # ACTUAL opening stock
                'current_stock': product.quantity,  # For comparison
                'stock_change': product.quantity - opening_stock,  # How much changed during the day
                'unit_cost': product.purchase_price,
                'stock_value': stock_value,
                'stock_status': stock_status,
                'reorder_level': product.reorder_level,
                'category': product.category.name if product.category else 'Uncategorized',
                'supplier': product.supplier.name if product.supplier else 'No Supplier',
                'sales_today': sales_today,
                'purchases_today': purchases_today,
                'stock_movements_today': stock_in_today - stock_out_today,
            })
        
        # Sort by stock value (most valuable items first)
        report_data.sort(key=lambda x: x['stock_value'], reverse=True)
        
        # Calculate summary statistics
        if total_products > 0:
            avg_stock_value = total_value / total_products
            avg_quantity = total_quantity / total_products
            
            # Calculate stock health
            in_stock_count = total_products - out_of_stock_count
            healthy_stock_count = total_products - (low_stock_count + out_of_stock_count)
            
            if total_products > 0:
                stock_health_percentage = (healthy_stock_count / total_products) * 100
            else:
                stock_health_percentage = Decimal('0')
        else:
            avg_stock_value = Decimal('0')
            avg_quantity = Decimal('0')
            stock_health_percentage = Decimal('0')
        
        # Get products that need reordering
        reorder_items = [item for item in report_data if item['stock_status'] in ['low', 'out_of_stock']]
        
        context = {
            'report_data': report_data,
            'summary': {
                'date': selected_date,
                'total_products': total_products,
                'total_quantity': total_quantity,
                'total_value': total_value,
                'avg_stock_value': avg_stock_value,
                'avg_quantity': avg_quantity,
                'low_stock_count': low_stock_count,
                'out_of_stock_count': out_of_stock_count,
                'in_stock_count': total_products - out_of_stock_count,
                'healthy_stock_count': total_products - (low_stock_count + out_of_stock_count),
                'stock_health_percentage': stock_health_percentage,
                'reorder_items_count': len(reorder_items),
            },
            'selected_date': date_str,
            'today': timezone.now().date(),
            'report_title': f"Opening Stock Report - {selected_date.strftime('%B %d, %Y')}",
        }
        
        # Export functionality
        export_format = request.GET.get('export')
        if export_format:
            return export_opening_stock_report(export_format, report_data, selected_date, summary=context['summary'])
        
        return render(request, 'pos/reports/daily_opening_stock.html', context)
        
    except Exception as e:
        print(f"ERROR in daily_opening_stock_report: {str(e)}")
        import traceback
        traceback.print_exc()
        
        messages.error(request, f"Error generating opening stock report: {str(e)}")
        
        # Fallback: Show simple opening stock (current stock) if detailed calculation fails
        products = Product.objects.filter(is_active=True)
        simple_report = []
        total_qty = Decimal('0')
        total_val = Decimal('0')
        
        for product in products:
            opening_stock = product.quantity  # Fallback to current stock
            stock_value = opening_stock * product.purchase_price
            total_qty += opening_stock
            total_val += stock_value
            
            simple_report.append({
                'product': product,
                'opening_stock': opening_stock,
                'stock_value': stock_value,
                'unit_cost': product.purchase_price,
                'category': product.category.name if product.category else 'Uncategorized',
            })
        
        context = {
            'report_data': simple_report,
            'summary': {
                'date': timezone.now().date(),
                'total_products': len(simple_report),
                'total_quantity': total_qty,
                'total_value': total_val,
                'avg_stock_value': total_val / len(simple_report) if simple_report else Decimal('0'),
                'low_stock_count': 0,
                'out_of_stock_count': 0,
                'stock_health_percentage': 0,
            },
            'selected_date': timezone.now().date().strftime('%Y-%m-%d'),
            'today': timezone.now().date(),
            'report_title': f"Opening Stock Report (Fallback) - {timezone.now().date().strftime('%B %d, %Y')}",
            'error': str(e),
        }
        
        return render(request, 'pos/reports/daily_opening_stock.html', context)


def export_opening_stock_report(format_type, report_data, date, summary):
    """Export opening stock report to CSV or PDF"""
    if format_type == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="opening_stock_{date}.csv"'
        
        writer = csv.writer(response)
        # Write header
        writer.writerow([
            'Product Name', 'Barcode/SKU', 'Category', 'Supplier',
            'Opening Stock Quantity', 'Unit Cost', 'Stock Value',
            'Reorder Level', 'Stock Status', 'Last Updated'
        ])
        
        # Write data rows
        for item in report_data:
            writer.writerow([
                item['product'].name,
                item['product'].barcode or 'N/A',
                item['category'],
                item['supplier'],
                float(item['opening_stock']),
                float(item['unit_cost']),
                float(item['stock_value']),
                item['reorder_level'],
                item['stock_status'].replace('_', ' ').title(),
                item['product'].updated_at.strftime('%Y-%m-%d %H:%M') if hasattr(item['product'], 'updated_at') else 'N/A'
            ])
        
        # Write summary
        writer.writerow([])  # Empty row
        writer.writerow(['SUMMARY'])
        writer.writerow(['Total Products', summary['total_products']])
        writer.writerow(['Total Quantity', float(summary['total_quantity'])])
        writer.writerow(['Total Value', f"KES {float(summary['total_value']):.2f}"])
        writer.writerow(['Low Stock Items', summary['low_stock_count']])
        writer.writerow(['Out of Stock Items', summary['out_of_stock_count']])
        writer.writerow(['Report Date', date.strftime('%Y-%m-%d')])
        writer.writerow(['Generated On', timezone.now().strftime('%Y-%m-%d %H:%M:%S')])
        
        return response
    
    # You can add PDF export here if needed
    return HttpResponseBadRequest("Export format not supported")
        

def export_opening_stock(export_format, report_data, date):
    """Export opening stock report"""
    if export_format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="opening_stock_{date}.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            'Product Name', 'SKU', 'Category',
            'Opening Stock', 'Stock In', 'Stock Out', 'Closing Stock',
            'Status', 'Opening Value', 'Closing Value', 'Value Change'
        ])
        
        for item in report_data:
            writer.writerow([
                item['product'].name,
                item['product'].barcode or '',
                item['product'].category.name if item['product'].category else '',
                float(item['opening_stock']),
                float(item['stock_in']),
                float(item['stock_out']),
                float(item['closing_stock']),
                item['stock_status'].capitalize(),
                float(item['opening_value']),
                float(item['closing_value']),
                float(item['value_change']),
            ])
        
        return response
    
    elif export_format == 'pdf':
        # You can implement PDF export using reportlab or weasyprint
        pass
    
    return HttpResponseBadRequest("Invalid export format")

# ============== ADDITIONAL VALUABLE REPORTS ==============

@login_required
def top_selling_products_report(request):
    """Top selling products report"""
    # Date filtering
    start_date = request.GET.get('start_date', (timezone.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
    end_date = request.GET.get('end_date', timezone.now().strftime('%Y-%m-%d'))
    
    # Get sales data
    sale_items = SaleItem.objects.filter(
        sale__date__date__range=[start_date, end_date],
        sale__is_completed=True
    ).select_related('product', 'product__category')
    
    # Group by product for top sellers
    top_by_quantity = sale_items.values(
        'product__id',
        'product__name',
        'product__category__name'
    ).annotate(
        total_quantity=Sum('quantity'),
        total_sales=Sum('total'),
        sale_count=Count('sale', distinct=True)
    ).order_by('-total_quantity')[:20]
    
    top_by_revenue = sale_items.values(
        'product__id',
        'product__name',
        'product__category__name'
    ).annotate(
        total_quantity=Sum('quantity'),
        total_sales=Sum('total'),
        sale_count=Count('sale', distinct=True)
    ).order_by('-total_sales')[:20]
    
    # Get categories for filter
    categories = Category.objects.all()
    
    context = {
        'top_by_quantity': top_by_quantity,
        'top_by_revenue': top_by_revenue,
        'categories': categories,
        'start_date': start_date,
        'end_date': end_date,
    }
    return render(request, 'pos/reports/top_selling_products.html', context)

@login_required
def slow_moving_products_report(request):
    """Slow moving/obsolete products report"""
    # Get date threshold (products not sold in last X days)
    days_threshold = int(request.GET.get('days', 30))
    threshold_date = timezone.now() - timedelta(days=days_threshold)
    
    # Get all active products
    products = Product.objects.filter(is_active=True).select_related('category')
    
    # Get last sale date for each product
    slow_moving_products = []
    for product in products:
        last_sale_item = SaleItem.objects.filter(
            product=product,
            sale__is_completed=True
        ).order_by('-sale__date').first()
        
        last_sale_date = last_sale_item.sale.date if last_sale_item else None
        
        days_since_sale = (timezone.now().date() - last_sale_date.date()).days if last_sale_date else None
        
        stock_value = product.quantity * product.purchase_price
        
        slow_moving_products.append({
            'product': product,
            'current_stock': product.quantity,
            'last_sale_date': last_sale_date,
            'days_since_sale': days_since_sale,
            'stock_value': stock_value,
            'reorder_level': product.reorder_level,
            'is_slow_moving': days_since_sale and days_since_sale > days_threshold,
            'is_excess_stock': product.quantity > (product.reorder_level * 3)  # More than 3x reorder level
        })
    
    # Filter and sort
    filtered_products = [p for p in slow_moving_products if p['is_slow_moving'] or p['is_excess_stock']]
    filtered_products.sort(key=lambda x: x['stock_value'], reverse=True)
    
    # Calculate summary
    total_slow_moving = len([p for p in filtered_products if p['is_slow_moving']])
    total_excess_stock = len([p for p in filtered_products if p['is_excess_stock']])
    total_stock_value = sum(p['stock_value'] for p in filtered_products)
    
    context = {
        'products': filtered_products,
        'total_slow_moving': total_slow_moving,
        'total_excess_stock': total_excess_stock,
        'total_stock_value': total_stock_value,
        'days_threshold': days_threshold,
    }
    return render(request, 'pos/reports/slow_moving_products.html', context)

@login_required
def customer_sales_analysis(request):
    """Customer sales and profitability analysis"""
    try:
        # Date filtering
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
        
        # Parse dates
        if start_date_str and end_date_str:
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            except ValueError:
                # Default to last 90 days
                end_date = timezone.now().date()
                start_date = end_date - timedelta(days=90)
                start_date_str = start_date.strftime('%Y-%m-%d')
                end_date_str = end_date.strftime('%Y-%m-%d')
        else:
            # Default to last 90 days
            end_date = timezone.now().date()
            start_date = end_date - timedelta(days=90)
            start_date_str = start_date.strftime('%Y-%m-%d')
            end_date_str = end_date.strftime('%Y-%m-%d')
        
        # Customer filter
        customer_id = request.GET.get('customer')
        selected_customer = None
        
        # Get sales data
        sales = Sale.objects.filter(
            date__date__range=[start_date, end_date],
            is_completed=True
        ).select_related('customer')
        
        # Apply customer filter if provided and valid
        if customer_id and customer_id.isdigit():
            try:
                selected_customer = Customer.objects.get(id=int(customer_id))
                sales = sales.filter(customer_id=int(customer_id))
            except (Customer.DoesNotExist, ValueError):
                # If customer doesn't exist, ignore the filter
                customer_id = None
        
        # Group by customer
        customer_analysis = sales.values(
            'customer__id',
            'customer__name',
            'customer__phone',
            'customer__balance'
        ).annotate(
            total_sales=Sum('total'),
            sale_count=Count('id'),
            avg_sale_value=Avg('total'),
            last_purchase=Max('date')
        ).order_by('-total_sales')
        
        # Calculate total items and profitability for each customer
        customer_list = []
        for customer in customer_analysis:
            # Get customer sales
            if customer['customer__id']:
                customer_sales = sales.filter(customer_id=customer['customer__id'])
            else:
                # Handle sales without customer (walk-in)
                customer_sales = sales.filter(customer__isnull=True)
            
            # Get sale items for this customer
            sale_items = SaleItem.objects.filter(
                sale__in=customer_sales
            ).select_related('product')
            
            # Calculate totals
            total_items = 0
            total_cost = Decimal('0.00')
            total_revenue = Decimal('0.00')
            
            for item in sale_items:
                if item.product:
                    quantity = item.quantity or Decimal('0')
                    total_items += quantity
                    
                    # Calculate cost and revenue
                    cost_per_unit = item.product.purchase_price or Decimal('0')
                    total_cost += quantity * cost_per_unit
                    total_revenue += item.total or Decimal('0')
            
            # Calculate profit
            total_profit = total_revenue - total_cost
            
            # Calculate profit margin
            profit_margin = Decimal('0.00')
            if total_revenue > 0:
                profit_margin = (total_profit / total_revenue) * 100
            
            # Add credit/cash sales breakdown
            credit_sales = customer_sales.filter(is_credit=True).aggregate(
                total=Sum('total')
            )['total'] or Decimal('0.00')
            
            cash_sales = customer_sales.filter(is_credit=False).aggregate(
                total=Sum('total')
            )['total'] or Decimal('0.00')
            
            customer_list.append({
                'customer__id': customer['customer__id'],
                'customer__name': customer['customer__name'] or "Walk-in Customer",
                'customer__phone': customer['customer__phone'],
                'customer__balance': customer['customer__balance'],
                'total_sales': customer['total_sales'] or Decimal('0.00'),
                'sale_count': customer['sale_count'] or 0,
                'avg_sale_value': customer['avg_sale_value'] or Decimal('0.00'),
                'total_items': total_items,
                'total_cost': total_cost,
                'total_revenue': total_revenue,
                'total_profit': total_profit,
                'profit_margin': profit_margin,
                'credit_sales': credit_sales,
                'cash_sales': cash_sales,
                'last_purchase': customer['last_purchase'],
            })
        
        # Calculate summary statistics
        total_customers = len(customer_list)
        total_sales_sum = sum(c['total_sales'] for c in customer_list)
        total_profit_sum = sum(c['total_profit'] for c in customer_list)
        
        avg_profit_margin = Decimal('0.00')
        if total_sales_sum > 0:
            avg_profit_margin = (total_profit_sum / total_sales_sum) * 100
        
        active_customers = len([c for c in customer_list if c['sale_count'] > 0])
        repeat_customers = len([c for c in customer_list if c['sale_count'] > 1])
        
        # Calculate average sales per customer
        avg_sales_per_customer = Decimal('0.00')
        if total_customers > 0:
            avg_sales_per_customer = total_sales_sum / total_customers
        
        summary = {
            'total_customers': total_customers,
            'total_sales': total_sales_sum,
            'total_profit': total_profit_sum,
            'avg_profit_margin': avg_profit_margin,
            'avg_sales_per_customer': avg_sales_per_customer,
            'active_customers': active_customers,
            'repeat_customers': repeat_customers,
        }
        
        # Get all customers for filter dropdown
        all_customers = Customer.objects.all().order_by('name')
        
        # Calculate days difference for display
        days_diff = (end_date - start_date).days + 1
        
        context = {
            'customer_analysis': customer_list,
            'summary': summary,
            'start_date': start_date_str,
            'end_date': end_date_str,
            'customers': all_customers,
            'selected_customer': customer_id,
            'selected_customer_obj': selected_customer,
            'days_diff': days_diff,
            'today': timezone.now(),  # Changed to datetime for template compatibility
        }
        
        # Export functionality
        export_format = request.GET.get('export')
        if export_format:
            return export_customer_analysis(export_format, customer_list, start_date_str, end_date_str, summary)
        
        return render(request, 'pos/reports/customer_sales_analysis.html', context)
        
    except Exception as e:
        print(f"Error in customer_sales_analysis: {str(e)}")
        import traceback
        traceback.print_exc()
        
        messages.error(request, f"Error generating customer analysis report: {str(e)}")
        return redirect('reports_dashboard')

def export_customer_analysis(export_format, report_data, start_date, end_date, summary):
    """Export customer analysis to CSV or Excel"""
    if export_format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="customer_analysis_{start_date}_to_{end_date}.csv"'
        
        writer = csv.writer(response)
        
        # Write header
        writer.writerow([
            'Customer Name', 'Phone', 'Balance', 'Total Sales (KES)', 
            'Number of Purchases', 'Total Items', 'Average Sale Value (KES)',
            'Total Cost (KES)', 'Total Revenue (KES)', 'Total Profit (KES)',
            'Profit Margin %', 'Credit Sales (KES)', 'Cash Sales (KES)',
            'Last Purchase Date'
        ])
        
        # Write data rows
        for customer in report_data:
            writer.writerow([
                customer['customer__name'],
                customer['customer__phone'] or '',
                float(customer['customer__balance'] or 0),
                float(customer['total_sales'] or 0),
                customer['sale_count'],
                float(customer['total_items'] or 0),
                float(customer['avg_sale_value'] or 0),
                float(customer['total_cost'] or 0),
                float(customer['total_revenue'] or 0),
                float(customer['total_profit'] or 0),
                float(customer['profit_margin'] or 0),
                float(customer['credit_sales'] or 0),
                float(customer['cash_sales'] or 0),
                customer['last_purchase'].strftime('%Y-%m-%d') if customer['last_purchase'] else '',
            ])
        
        # Write summary
        writer.writerow([])
        writer.writerow(['SUMMARY'])
        writer.writerow(['Report Period', f'{start_date} to {end_date}'])
        writer.writerow(['Total Customers', summary['total_customers']])
        writer.writerow(['Total Sales (KES)', float(summary['total_sales'] or 0)])
        writer.writerow(['Total Profit (KES)', float(summary['total_profit'] or 0)])
        writer.writerow(['Average Profit Margin', f"{summary['avg_profit_margin']:.2f}%"])
        writer.writerow(['Active Customers', summary['active_customers']])
        writer.writerow(['Repeat Customers', summary['repeat_customers']])
        writer.writerow(['Generated On', timezone.now().strftime('%Y-%m-%d %H:%M:%S')])
        
        return response
    
    elif export_format == 'excel':
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="customer_analysis_{start_date}_to_{end_date}.xlsx"'
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Customer Analysis"
        
        # Add headers
        headers = [
            'Customer Name', 'Phone', 'Balance', 'Total Sales (KES)', 
            'Number of Purchases', 'Total Items', 'Average Sale Value (KES)',
            'Total Cost (KES)', 'Total Revenue (KES)', 'Total Profit (KES)',
            'Profit Margin %', 'Credit Sales (KES)', 'Cash Sales (KES)',
            'Last Purchase Date'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
        
        # Add data
        row_num = 2
        for customer in report_data:
            ws.cell(row=row_num, column=1, value=customer['customer__name'])
            ws.cell(row=row_num, column=2, value=customer['customer__phone'] or '')
            ws.cell(row=row_num, column=3, value=float(customer['customer__balance'] or 0))
            ws.cell(row=row_num, column=4, value=float(customer['total_sales'] or 0))
            ws.cell(row=row_num, column=5, value=customer['sale_count'])
            ws.cell(row=row_num, column=6, value=float(customer['total_items'] or 0))
            ws.cell(row=row_num, column=7, value=float(customer['avg_sale_value'] or 0))
            ws.cell(row=row_num, column=8, value=float(customer['total_cost'] or 0))
            ws.cell(row=row_num, column=9, value=float(customer['total_revenue'] or 0))
            
            # Profit with conditional formatting
            profit_cell = ws.cell(row=row_num, column=10, value=float(customer['total_profit'] or 0))
            if customer['total_profit'] >= 0:
                profit_cell.font = Font(color='006100', bold=True)
            else:
                profit_cell.font = Font(color='FF0000', bold=True)
            
            # Margin with conditional formatting
            margin_cell = ws.cell(row=row_num, column=11, value=float(customer['profit_margin'] or 0))
            if customer['profit_margin'] >= 20:
                margin_cell.font = Font(color='006100', bold=True)
            elif customer['profit_margin'] >= 10:
                margin_cell.font = Font(color='FFA500', bold=True)
            else:
                margin_cell.font = Font(color='FF0000', bold=True)
            
            ws.cell(row=row_num, column=12, value=float(customer['credit_sales'] or 0))
            ws.cell(row=row_num, column=13, value=float(customer['cash_sales'] or 0))
            ws.cell(row=row_num, column=14, value=customer['last_purchase'].strftime('%Y-%m-%d') if customer['last_purchase'] else '')
            
            row_num += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add summary sheet
        ws2 = wb.create_sheet(title="Summary")
        ws2.cell(row=1, column=1, value="CUSTOMER ANALYSIS SUMMARY").font = Font(bold=True, size=14)
        
        summary_data = [
            ("Report Period", f"{start_date} to {end_date}"),
            ("Total Customers", summary['total_customers']),
            ("Total Sales (KES)", float(summary['total_sales'] or 0)),
            ("Total Profit (KES)", float(summary['total_profit'] or 0)),
            ("Average Profit Margin", f"{summary['avg_profit_margin']:.2f}%"),
            ("Average Sales per Customer", f"KES {float(summary['avg_sales_per_customer'] or 0):.2f}"),
            ("Active Customers", summary['active_customers']),
            ("Repeat Customers", summary['repeat_customers']),
            ("Generated On", timezone.now().strftime('%Y-%m-%d %H:%M:%S')),
        ]
        
        for i, (label, value) in enumerate(summary_data, 3):
            ws2.cell(row=i, column=1, value=label).font = Font(bold=True)
            ws2.cell(row=i, column=2, value=value)
        
        wb.save(response)
        return response
    
    return HttpResponse('Invalid export format', status=400)

@login_required
def supplier_purchase_analysis(request):
    # Date filtering
    start_date = request.GET.get('start_date') or (timezone.now() - timedelta(days=90)).date()
    end_date = request.GET.get('end_date') or timezone.now().date()
    
    # Get purchases in date range
    purchases = Purchase.objects.filter(
        date__range=[start_date, end_date],
        is_return=False  # Exclude returns
    ).select_related('supplier')
    
    # Get supplier payments
    supplier_payments = SupplierPayment.objects.filter(
        date__range=[start_date, end_date]
    ).select_related('supplier')
    
    # Get all suppliers for analysis
    suppliers = Supplier.objects.all()
    
    supplier_analysis = []
    
    for supplier in suppliers:
        # Get purchases for this supplier
        supplier_purchases = purchases.filter(supplier=supplier)
        
        # Get payments for this supplier
        supplier_payment_records = supplier_payments.filter(supplier=supplier)
        
        # Calculate purchase statistics
        purchase_count = supplier_purchases.count()
        total_purchases = supplier_purchases.aggregate(total=Sum('total'))['total'] or Decimal('0.0000')
        total_items = supplier_purchases.aggregate(total=Sum('item_count'))['total'] or 0
        
        # Get product count (unique products purchased)
        product_count = PurchaseItem.objects.filter(
            purchase__in=supplier_purchases
        ).values('product').distinct().count()
        
        # Calculate payment statistics
        total_payments = supplier_payment_records.aggregate(total=Sum('amount'))['total'] or Decimal('0.0000')
        
        # Calculate balance due (purchases - payments)
        balance_due = total_purchases - total_payments
        
        # Calculate payment ratio (avoid division by zero)
        if total_purchases > 0:
            payment_ratio = (total_payments / total_purchases) * 100
        else:
            payment_ratio = Decimal('0.0000') if total_payments > 0 else Decimal('0.0000')
        
        # Get last purchase date
        last_purchase = supplier_purchases.order_by('-date').first()
        last_purchase_date = last_purchase.date if last_purchase else None
        
        # Calculate average purchase value (avoid division by zero)
        if purchase_count > 0:
            avg_purchase_value = total_purchases / purchase_count
        else:
            avg_purchase_value = Decimal('0.0000')
        
        # Only include suppliers with activity in the period
        if purchase_count > 0 or total_payments > 0:
            supplier_analysis.append({
                'supplier__name': supplier.name,
                'supplier__phone': supplier.phone,
                'supplier__balance': supplier.balance,
                'total_purchases': total_purchases,
                'purchase_count': purchase_count,
                'total_items': total_items,
                'product_count': product_count,
                'total_payments': total_payments,
                'balance_due': balance_due,
                'payment_ratio': payment_ratio,
                'last_purchase': last_purchase_date,
                'avg_purchase_value': avg_purchase_value,
                'supplier_id': supplier.id,
            })
    
    # Sort by total purchases descending
    supplier_analysis.sort(key=lambda x: x['total_purchases'], reverse=True)
    
    # Calculate summary statistics
    total_suppliers = len(supplier_analysis)
    active_suppliers = len([s for s in supplier_analysis if s['purchase_count'] > 0])
    
    # Calculate totals safely
    total_purchases_sum = Decimal('0.0000')
    total_payments_sum = Decimal('0.0000')
    total_balance_sum = Decimal('0.0000')
    
    for supplier in supplier_analysis:
        total_purchases_sum += supplier['total_purchases']
        total_payments_sum += supplier['total_payments']
        total_balance_sum += supplier['balance_due']
    
    # Calculate average payment ratio (only for suppliers with purchases)
    suppliers_with_purchases = [s for s in supplier_analysis if s['total_purchases'] > 0]
    if suppliers_with_purchases:
        avg_payment_ratio = sum(s['payment_ratio'] for s in suppliers_with_purchases) / len(suppliers_with_purchases)
    else:
        avg_payment_ratio = Decimal('0.0000')
    
    # Calculate good, average, and poor suppliers
    good_suppliers = len([s for s in supplier_analysis if s['payment_ratio'] > 80])
    average_suppliers = len([s for s in supplier_analysis if 50 <= s['payment_ratio'] <= 80])
    poor_suppliers = len([s for s in supplier_analysis if s['payment_ratio'] < 50])
    
    # Format the data for template
    formatted_supplier_analysis = []
    for supplier in supplier_analysis:
        formatted_supplier_analysis.append({
            'supplier__name': supplier['supplier__name'],
            'supplier__phone': supplier['supplier__phone'],
            'supplier__balance': format_currency(supplier['supplier__balance']),
            'total_purchases': format_currency(supplier['total_purchases']),
            'purchase_count': supplier['purchase_count'],
            'total_items': supplier['total_items'],
            'product_count': supplier['product_count'],
            'total_payments': format_currency(supplier['total_payments']),
            'balance_due': format_currency(supplier['balance_due']),
            'payment_ratio': supplier['payment_ratio'],
            'last_purchase': supplier['last_purchase'],
            'avg_purchase_value': format_currency(supplier['avg_purchase_value']),
            'supplier_id': supplier['supplier_id'],
            'raw_total_purchases': supplier['total_purchases'],  # Keep raw for calculations
            'raw_total_payments': supplier['total_payments'],
            'raw_payment_ratio': supplier['payment_ratio'],
        })
    
    summary = {
        'total_suppliers': total_suppliers,
        'active_suppliers': active_suppliers,
        'total_purchases': format_currency(total_purchases_sum),
        'total_payments': format_currency(total_payments_sum),
        'total_balance': format_currency(total_balance_sum),
        'avg_payment_ratio': avg_payment_ratio,
    }
    
    context = {
        'supplier_analysis': formatted_supplier_analysis,
        'summary': summary,
        'good_suppliers': good_suppliers,
        'average_suppliers': average_suppliers,
        'poor_suppliers': poor_suppliers,
        'start_date': start_date,
        'end_date': end_date,
        'today': timezone.now().date(),
    }
    
    return render(request, 'pos/reports/supplier_purchase_analysis.html', context)

# ============== EXPORT FUNCTIONS ==============

@login_required
def export_report(request, report_type):
    """Export report to Excel"""
    # Map report types to functions
    report_mapping = {
        'sales_by_product': sales_by_product_report,
        'purchase_by_product': purchase_by_product_report,
        'customer_payments': customer_payment_report,
        'supplier_payments': supplier_payment_report,
        'profit_by_product': profit_by_product_report,
        # Add other report mappings
    }
    
    if report_type not in report_mapping:
        messages.error(request, 'Invalid report type')
        return redirect('reports_dashboard')
    
    # Get report data (we'll implement proper Excel export)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{report_type}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = report_type.replace('_', ' ').title()
    
    # Add headers and data based on report type
    # (Implementation depends on specific report structure)
    
    wb.save(response)
    return response

# ============== HELPER FUNCTIONS ==============

def generate_chart_image(data, chart_type='bar', title=''):
    """Generate chart image from data"""
    plt.figure(figsize=(10, 6))
    
    if chart_type == 'bar':
        labels = [item['label'] for item in data]
        values = [item['value'] for item in data]
        plt.bar(labels, values)
    elif chart_type == 'line':
        dates = [item['date'] for item in data]
        values = [item['value'] for item in data]
        plt.plot(dates, values, marker='o')
    elif chart_type == 'pie':
        labels = [item['label'] for item in data]
        values = [item['value'] for item in data]
        plt.pie(values, labels=labels, autopct='%1.1f%%')
    
    plt.title(title)
    plt.xticks(rotation=45)
    plt.tight_layout()
    
    # Save to buffer
    buffer = BytesIO()
    plt.savefig(buffer, format='png')
    plt.close()
    buffer.seek(0)
    
    # Encode to base64
    image_png = buffer.getvalue()
    buffer.close()
    graphic = base64.b64encode(image_png)
    graphic = graphic.decode('utf-8')
    
    return graphic




    
from .models import SupplierReturn
from .forms import SupplierReturnForm

def supplier_returns(request):
    if request.method == 'POST':
        form = SupplierReturnForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('supplier_returns_list')
    else:
        form = SupplierReturnForm()
    
    returns = SupplierReturn.objects.all().order_by('-return_date')
    return render(request, 'pos/supplier_returns.html', {
        'form': form,
        'returns': returns
    })

def process_return(request, return_id):
    supplier_return = get_object_or_404(SupplierReturn, id=return_id)
    if request.method == 'POST':
        new_status = request.POST.get('status')
        supplier_return.status = new_status
        supplier_return.save()
        
        # If approved, update inventory
        if new_status == 'approved':
            product = supplier_return.product
            product.quantity += supplier_return.quantity
            product.save()
            
        return redirect('supplier_returns_list')
    
    return render(request, 'pos/process_return.html', {
        'return': supplier_return
    })


# ============== Expense Views ==============
@login_required
def expense_list(request):
    # Base queryset
    expenses = Expense.objects.all().order_by('-date')
    
    # Date filtering - FIXED
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if start_date:
        try:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            expenses = expenses.filter(date__gte=start_date)
        except ValueError:
            pass
    
    if end_date:
        try:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            expenses = expenses.filter(date__lte=end_date)
        except ValueError:
            pass
    
    # Category filter - FIXED
    category = request.GET.get('category')
    if category:
        expenses = expenses.filter(category=category)
    
    # Calculate totals - FIXED
    total_expenses = expenses.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    
    # Calculate average daily expense
    if start_date and end_date:
        days = (end_date - start_date).days + 1
        avg_daily_expense = total_expenses / days if days > 0 else Decimal('0.00')
    else:
        # Default to last 30 days
        thirty_days_ago = timezone.now().date() - timedelta(days=30)
        recent_expenses = Expense.objects.filter(date__gte=thirty_days_ago)
        monthly_total = recent_expenses.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        avg_daily_expense = monthly_total / 30
    
    # This month's expenses
    this_month_start = timezone.now().date().replace(day=1)
    monthly_expense = Expense.objects.filter(
        date__gte=this_month_start
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
    
    # Export functionality
    export_format = request.GET.get('export')
    if export_format == 'csv':
        return generate_expenses_export(expenses)
    
    # Category breakdown for current filter
    expenses_by_category = expenses.values('category').annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('-total')
    
    # Monthly expenses for chart (last 6 months)
    six_months_ago = timezone.now().date() - timedelta(days=180)
    monthly_expenses_chart = expenses.filter(
        date__gte=six_months_ago
    ).annotate(
        month=TruncMonth('date')
    ).values('month').annotate(
        total=Sum('amount')
    ).order_by('month')
    
    # Pagination
    paginator = Paginator(expenses, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'expenses': page_obj,
        'total_expenses': total_expenses,
        'avg_daily_expense': avg_daily_expense,
        'monthly_expense': monthly_expense,
        'start_date': start_date.strftime('%Y-%m-%d') if start_date else '',
        'end_date': end_date.strftime('%Y-%m-%d') if end_date else '',
        'selected_category': category,
        'expenses_by_category': expenses_by_category,
        'monthly_expenses_chart': monthly_expenses_chart,
        'categories': Expense.CATEGORIES,  # CHANGED FROM CATEGORY_CHOICES
    }
    return render(request, 'pos/expense_list.html', context)


@login_required
def add_expense(request):
    if request.method == 'POST':
        formset = ExpenseFormSet(request.POST, queryset=Expense.objects.none())
        bulk_form = BulkExpenseForm(request.POST)
        
        if formset.is_valid():
            instances = formset.save(commit=False)
            total_amount = Decimal('0.00')
            
            # Save all instances
            for instance in instances:
                instance.user = request.user
                if not instance.date:  # Set default date if not provided
                    instance.date = timezone.now().date()
                instance.save()
                total_amount += instance.amount
            
            # Handle deleted forms if any
            for deleted_form in formset.deleted_forms:
                if deleted_form.instance.pk:
                    deleted_form.instance.delete()
            
            messages.success(request, f'Successfully added {len(instances)} expense(s) totaling KSh {total_amount:.2f}')
            return redirect('expense_list')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        formset = ExpenseFormSet(queryset=Expense.objects.none())
        bulk_form = BulkExpenseForm()
    
    context = {
        'formset': formset,
        'bulk_form': bulk_form,
        'title': 'Add Multiple Expenses'
    }
    return render(request, 'pos/add_expense.html', context)

@login_required
def edit_expense(request, pk):
    expense = get_object_or_404(Expense, pk=pk)
    
    if request.method == 'POST':
        form = ExpenseForm(request.POST, instance=expense)
        if form.is_valid():
            form.save()
            return redirect('expense_list')
    else:
        form = ExpenseForm(instance=expense)
    
    context = {
        'form': form,
        'expense': expense
    }
    return render(request, 'pos/edit_expense.html', context)

def generate_expenses_export(format, queryset):
    if format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="expenses.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            'Date', 'Category', 'Description', 'Amount', 'User'
        ])
        
        for expense in queryset:
            writer.writerow([
                expense.date.strftime('%Y-%m-%d'),
                expense.get_category_display(),
                expense.description,
                expense.amount,
                expense.user.username
            ])
        
        return response
    
    return HttpResponse('Invalid export format', status=400)


@login_required
@require_POST
def delete_expense(request, pk):
    expense = get_object_or_404(Expense, pk=pk)
    
    try:
        expense.delete()
        messages.success(request, f'Expense "{expense.description}" has been deleted successfully.')
    except Exception as e:
        messages.error(request, f'Error deleting expense: {str(e)}')
    
    return redirect('expense_list')

# ============== Discount Views ==============
@login_required
def discount_list(request):
    today = timezone.now().date()
    discounts = Discount.objects.all().order_by('-start_date')
    
    # Filter by active status
    status_filter = request.GET.get('status')
    if status_filter == 'active':
        discounts = discounts.filter(
            is_active=True, 
            start_date__lte=today, 
            end_date__gte=today
        )
    elif status_filter == 'upcoming':
        discounts = discounts.filter(start_date__gt=today)
    elif status_filter == 'expired':
        discounts = discounts.filter(end_date__lt=today)
    
    context = {
        'discounts': discounts,
        'status_filter': status_filter
    }
    return render(request, 'pos/discount_list.html', context)


@login_required
def add_discount(request):
    if request.method == 'POST':
        form = DiscountForm(request.POST)
        if form.is_valid():
            discount = form.save()
            
            # Apply discount to selected products/categories
            if discount.discount_type == 'percentage':
                amount = discount.amount / 100
                update_func = lambda p: p.selling_price * (1 - amount)
            else:
                update_func = lambda p: max(p.selling_price - discount.amount, Decimal('0.01'))
            
            if discount.products.exists():
                for product in discount.products.all():
                    product.selling_price = update_func(product)
                    product.save()
            
            if discount.categories.exists():
                for category in discount.categories.all():
                    for product in category.product_set.all():
                        product.selling_price = update_func(product)
                        product.save()
            
            return redirect('discount_list')
    else:
        form = DiscountForm()
    
    products = Product.objects.filter(is_active=True).order_by('name')
    categories = Category.objects.all()
    
    context = {
        'form': form,
        'products': products,
        'categories': categories
    }
    return render(request, 'pos/add_discount.html', context)

@login_required
def toggle_discount_status(request, pk):
    discount = get_object_or_404(Discount, pk=pk)
    discount.is_active = not discount.is_active
    discount.save()
    return redirect('discount_list')

# ============== Company & Pricing Views ==============
@login_required
def company_settings(request):
    company = Company.objects.first()
    
    # Create company if it doesn't exist
    if not company:
        company = Company.objects.create(
            name="Succeed Cereal Hub",
            address="Wataalam, Ruiry",
            phone="+254 74545769",
            email="info@company.com",
            vat_number=""
        )
    
    if request.method == 'POST':
        form = CompanyForm(request.POST, request.FILES, instance=company)
        if form.is_valid():
            form.save()
            messages.success(request, 'Company settings updated successfully!')
            return redirect('company_settings')
    else:
        form = CompanyForm(instance=company)
    
    context = {
        'form': form,
        'company': company
    }
    return render(request, 'pos/company_settings.html', context)

@login_required
def company_pricing(request):
    company_prices = CompanyPrice.objects.all().order_by('company__name', 'product__name')
    
    # Filter by company
    company_id = request.GET.get('company')
    if company_id:
        company_prices = company_prices.filter(company_id=company_id)
    
    companies = Company.objects.all()
    products = Product.objects.filter(is_active=True).order_by('name')
    
    context = {
        'company_prices': company_prices,
        'companies': companies,
        'products': products,
        'selected_company': company_id
    }
    return render(request, 'pos/company_pricing.html', context)

@login_required
def add_company_price(request):
    if request.method == 'POST':
        form = CompanyPriceForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('company_pricing')
    else:
        form = CompanyPriceForm()
    
    companies = Company.objects.all()
    products = Product.objects.filter(is_active=True).order_by('name')
    
    context = {
        'form': form,
        'companies': companies,
        'products': products
    }
    return render(request, 'pos/add_company_price.html', context)

@login_required
def edit_company_price(request, pk):
    company_price = get_object_or_404(CompanyPrice, pk=pk)
    
    if request.method == 'POST':
        form = CompanyPriceForm(request.POST, instance=company_price)
        if form.is_valid():
            form.save()
            return redirect('company_pricing')
    else:
        form = CompanyPriceForm(instance=company_price)
    
    context = {
        'form': form,
        'company_price': company_price
    }
    return render(request, 'pos/edit_company_price.html', context)

# ============== Reports Views ==============
@login_required
def profit_loss_report(request):
    # Date filtering
    start_date = request.GET.get('start_date', (timezone.now() - timedelta(days=30)).strftime('%Y-%m-%d'))
    end_date = request.GET.get('end_date', timezone.now().strftime('%Y-%m-%d'))
    
    # Get sales in date range
    sales = Sale.objects.filter(
        date__date__range=[start_date, end_date],
        is_completed=True
    )
    total_sales = sales.aggregate(total=Sum('total'))['total'] or 0
    
    # Calculate cost of goods sold
    sale_items = SaleItem.objects.filter(sale__in=sales)
    cogs = sale_items.annotate(
        cost=ExpressionWrapper(
            F('product__purchase_price') * F('quantity'),
            output_field=DecimalField()
        )
    ).aggregate(total=Sum('cost'))['total'] or 0
    
    # Calculate gross profit
    gross_profit = total_sales - cogs
    
    # Get expenses in date range
    expenses = Expense.objects.filter(date__range=[start_date, end_date])
    total_expenses = expenses.aggregate(total=Sum('amount'))['total'] or 0
    
    # Group expenses by category
    expenses_by_category = expenses.values('category').annotate(
        total=Sum('amount')
    ).order_by('-total')
    
    # Calculate net profit
    net_profit = gross_profit - total_expenses
    
    # Profit trend data (last 12 months)
    profit_trend = []
    for i in range(12):
        month = timezone.now().date() - timedelta(days=30*i)
        month_start = month.replace(day=1)
        month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        
        month_sales = Sale.objects.filter(
            date__date__range=[month_start, month_end],
            is_completed=True
        ).aggregate(total=Sum('total'))['total'] or 0
        
        month_cogs = SaleItem.objects.filter(
            sale__date__date__range=[month_start, month_end]
        ).annotate(
            cost=ExpressionWrapper(
                F('product__purchase_price') * F('quantity'),
                output_field=DecimalField()
            )
        ).aggregate(total=Sum('cost'))['total'] or 0
        
        month_expenses = Expense.objects.filter(
            date__range=[month_start, month_end]
        ).aggregate(total=Sum('amount'))['total'] or 0
        
        month_profit = (month_sales - month_cogs) - month_expenses
        
        profit_trend.append({
            'month': month_start.strftime('%b %Y'),
            'sales': float(month_sales),
            'cogs': float(month_cogs),
            'expenses': float(month_expenses),
            'profit': float(month_profit)
        })
    
    profit_trend.reverse()  # Show oldest first
    
    context = {
        'total_sales': total_sales,
        'cogs': cogs,
        'gross_profit': gross_profit,
        'total_expenses': total_expenses,
        'net_profit': net_profit,
        'expenses_by_category': expenses_by_category,
        'profit_trend': profit_trend,
        'start_date': start_date,
        'end_date': end_date
    }
    return render(request, 'pos/profit_loss_report.html', context)

@login_required
def stock_value_report(request):
    products = Product.objects.annotate(
        stock_value=ExpressionWrapper(
            F('quantity') * F('selling_price'),
            output_field=DecimalField()
        )
    ).order_by('-stock_value')
    
    # Category filter
    category_id = request.GET.get('category')
    if category_id:
        products = products.filter(category_id=category_id)
    
    # Calculate total stock value
    total_value = products.aggregate(total=Sum('stock_value'))['total'] or 0
    
    # Get all categories for filter dropdown
    categories = Category.objects.all()
    
    context = {
        'products': products,
        'categories': categories,
        'total_value': total_value,
        'selected_category': category_id
    }
    return render(request, 'pos/stock_value_report.html', context)

# ============== Receipt Printing ==============
# views.py
from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.decorators import login_required
from django.template.loader import render_to_string
from django.conf import settings
from .models import Receipt, Sale

@login_required
def view_receipt(request, receipt_id):
    """View a receipt in HTML format"""
    receipt = get_object_or_404(Receipt, id=receipt_id)
    company = Company.objects.first()  # Assuming you have a Company model
    
    context = {
        'receipt': receipt,
        'company': company,
        'sale': receipt.sale,
        'items': receipt.content.get('items', []),
        'payment_details': receipt.content.get('payment_details', {}),
    }
    
    return render(request, 'pos/receipt.html', context)

@login_required
def print_receipt(request, receipt_id):
    """View optimized for printing"""
    receipt = get_object_or_404(Receipt, id=receipt_id)
    company = Company.objects.first()
    
    context = {
        'receipt': receipt,
        'company': company,
        'sale': receipt.sale,
        'items': receipt.content.get('items', []),
        'payment_details': receipt.content.get('payment_details', {}),
        'print_directly': True,  # Flag for template to optimize for printing
    }
    
    return render(request, 'pos/receipt_print.html', context)

@login_required
def generate_receipt(request, sale_id):
    """Generate a receipt for a sale"""
    sale = get_object_or_404(Sale, id=sale_id)
    
    # Generate receipt number (you might want a better system)
    today = timezone.now().date()
    last_receipt = Receipt.objects.filter(
        created_at__date=today
    ).order_by('-id').first()
    
    next_num = 1
    if last_receipt:
        try:
            last_num = int(last_receipt.receipt_number.split('-')[-1])
            next_num = last_num + 1
        except:
            pass
    
    receipt_number = f"RCP-{today.strftime('%Y%m%d')}-{next_num:04d}"
    
    # Prepare receipt content
    receipt_content = {
        'sale_number': sale.sale_number,
        'date': sale.date.strftime('%Y-%m-%d %H:%M:%S'),
        'customer': sale.customer.name if sale.customer else 'Walk-in Customer',
        'items': [{
            'name': item.product.name,
            'quantity': item.quantity,
            'price': str(item.price),
            'total': str(item.total),
            'batch': item.batch.batch_number if item.batch else '',
        } for item in sale.items.all()],
        'subtotal': str(sale.subtotal),
        'discount': str(sale.discount_amount),
        'total': str(sale.total),
        'payment_method': sale.get_payment_method_display(),
        'payment_details': {
            'cash': str(sale.cash_amount),
            'mpesa': str(sale.mpesa_amount),
            'card': str(sale.card_amount),
            'cheque': str(sale.cheque_amount),
        },
        'amount_paid': str(sale.amount_paid),
        'change': str(max(sale.amount_paid - sale.total, Decimal(0))),
        'balance': str(sale.balance),
    }
    
    # Create receipt record
    receipt = Receipt.objects.create(
        receipt_number=receipt_number,
        receipt_type='sale',
        sale=sale,
        content=receipt_content,
        created_by=request.user
    )
    
    return JsonResponse({
        'success': True,
        'receipt_id': receipt.id,
        'receipt_number': receipt.receipt_number,
        'print_url': reverse('print_receipt', args=[receipt.id]),
    })

@login_required
def receipt_history(request):
    """View receipt history"""
    receipts = Receipt.objects.all().order_by('-created_at')
    
    # Filter by date range
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if start_date:
        receipts = receipts.filter(created_at__date__gte=start_date)
    if end_date:
        receipts = receipts.filter(created_at__date__lte=end_date)
    
    # Filter by receipt type
    receipt_type = request.GET.get('type')
    if receipt_type:
        receipts = receipts.filter(receipt_type=receipt_type)
    
    # Filter by printed status
    printed = request.GET.get('printed')
    if printed == 'yes':
        receipts = receipts.filter(is_printed=True)
    elif printed == 'no':
        receipts = receipts.filter(is_printed=False)
    
    # Pagination
    paginator = Paginator(receipts, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'receipts': page_obj,
        'start_date': start_date,
        'end_date': end_date,
        'selected_type': receipt_type,
        'selected_printed': printed,
    }
    
    return render(request, 'pos/receipt_history.html', context)

# Add this to your views.py
import win32print
import win32ui
from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from .models import Sale, Company

import win32print
import win32ui
from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
import json

import pytz
from django.utils import timezone as django_timezone

# views.py - Updated print_receipt_direct with better error handling

@login_required
@require_POST
def print_receipt_direct(request, sale_id):
    """
    Print a receipt directly to the default thermal printer using ESC/POS commands.
    Handles both Windows and Linux environments with fallback to browser print.
    """
    try:
        # Get sale and company info
        sale = get_object_or_404(Sale, pk=sale_id)
        company = Company.objects.first() or Company.objects.create(
            name="Succeed Cereal Hub",
            address="Wataalam, Ruiry",
            phone="+254 74545769",
            email="info@company.com",
            vat_number=""
        )
        
        # Get current time in Nairobi timezone
        try:
            import pytz
            nairobi_tz = pytz.timezone('Africa/Nairobi')
            nairobi_time = django_timezone.now().astimezone(nairobi_tz)
        except ImportError:
            nairobi_time = django_timezone.now()
        
        # Format date and time
        nairobi_date_str = nairobi_time.strftime('%d/%m/%Y')
        nairobi_time_str = nairobi_time.strftime('%H:%M:%S')
        
        # Calculate change if amount paid is greater than total
        change = max(sale.amount_paid - sale.total, Decimal('0.00'))
        
        # 1. Prepare ESC/POS commands for thermal printer
        raw_data = b'\x1B\x40'  # ESC @ - Initialize printer
        
        # Company header
        raw_data += b'\x1B\x61\x01'  # Center alignment
        raw_data += b'\x1D\x21\x11'  # Double height & width
        raw_data += f"{company.name}\n".encode('utf-8')
        raw_data += b'\x1D\x21\x00'  # Normal text
        raw_data += f"{company.address}\n".encode('utf-8')
        raw_data += f"Phone: {company.phone}\n".encode('utf-8')
        if company.vat_number:
            raw_data += f"VAT: {company.vat_number}\n".encode('utf-8')
        raw_data += b'================================\n'
        
        # Sale info
        raw_data += b'\x1B\x21\x10'  # Emphasized
        raw_data += f"RECEIPT: {sale.sale_number}\n".encode('utf-8')
        raw_data += b'\x1B\x21\x00'
        raw_data += f"Date: {nairobi_date_str}\n".encode('utf-8')
        raw_data += f"Time: {nairobi_time_str}\n".encode('utf-8')
        raw_data += f"Customer: {sale.customer.name if sale.customer else 'Walk-in'}\n".encode('utf-8')
        raw_data += f"Sale Type: {'Wholesale' if sale.sale_type == 'wholesale' else 'Retail'}\n".encode('utf-8')
        raw_data += f"Cashier: {sale.user.get_full_name() or sale.user.username}\n".encode('utf-8')
        raw_data += b'--------------------------------\n'
        
        # Items table header
        raw_data += b'\x1B\x21\x08'  # Bold/emphasized
        raw_data += "Item                     Qty  Price   Total\n".encode('utf-8')
        raw_data += b'\x1B\x21\x00'
        
        # Items
        for item in sale.items.all():
            item_name = item.product.name
            if len(item_name) > 18:
                item_name = item_name[:16] + '..'
            item_name = item_name.ljust(22)
            
            qty = f"{item.quantity:.2f}"
            price = f"{item.price:.2f}"
            total = f"{item.total:.2f}"
            
            line = f"{item_name} {qty:>5} {price:>7} {total:>8}\n"
            raw_data += line.encode('utf-8')
        
        raw_data += b'--------------------------------\n'
        
        # Totals section
        raw_data += f"SUBTOTAL:{' ' * 15}KSh {sale.subtotal:>9.2f}\n".encode('utf-8')
        
        if sale.discount_amount > 0:
            raw_data += f"DISCOUNT:{' ' * 15}KSh {sale.discount_amount:>9.2f}\n".encode('utf-8')
        
        raw_data += f"TOTAL:{' ' * 17}KSh {sale.total:>9.2f}\n".encode('utf-8')
        
        # Payment details
        raw_data += f"Payment: {sale.get_payment_method_display()}\n".encode('utf-8')
        
        if sale.payment_method == 'cash':
            raw_data += f"Cash Paid: KSh {sale.amount_paid:.2f}\n".encode('utf-8')
            if change > 0:
                raw_data += f"Change: KSh {change:.2f}\n".encode('utf-8')
        elif sale.payment_method == 'mpesa':
            raw_data += f"M-Pesa: KSh {sale.amount_paid:.2f}\n".encode('utf-8')
            if sale.mpesa_code:
                raw_data += f"Code: {sale.mpesa_code}\n".encode('utf-8')
        elif sale.payment_method == 'multiple':
            if sale.cash_amount > 0:
                raw_data += f"Cash: KSh {sale.cash_amount:.2f}\n".encode('utf-8')
            if sale.mpesa_amount > 0:
                raw_data += f"M-Pesa: KSh {sale.mpesa_amount:.2f}\n".encode('utf-8')
                if sale.mpesa_code:
                    raw_data += f"M-Pesa Code: {sale.mpesa_code}\n".encode('utf-8')
            if sale.card_amount > 0:
                raw_data += f"Card: KSh {sale.card_amount:.2f}\n".encode('utf-8')
            if sale.cheque_amount > 0:
                raw_data += f"Cheque: KSh {sale.cheque_amount:.2f}\n".encode('utf-8')
                if sale.cheque_number:
                    raw_data += f"Cheque #: {sale.cheque_number}\n".encode('utf-8')
            raw_data += f"Total Paid: KSh {sale.amount_paid:.2f}\n".encode('utf-8')
            if change > 0:
                raw_data += f"Change: KSh {change:.2f}\n".encode('utf-8')
        
        if sale.balance > 0:
            raw_data += f"Balance Due: KSh {sale.balance:.2f}\n".encode('utf-8')
        
        raw_data += b'\n'
        
        # Footer
        raw_data += b'\x1B\x61\x01'
        raw_data += f"Time Printed: {nairobi_time_str}\n".encode('utf-8')
        raw_data += "Thank you for your business!\n".encode('utf-8')
        raw_data += "--------------------------------\n".encode('utf-8')
        raw_data += b'\x1B\x21\x08'
        raw_data += "POS by Globe Tech Systems\n".encode('utf-8')
        raw_data += b'\x1B\x21\x00'
        raw_data += "--------------------------------\n".encode('utf-8')
        raw_data += "Goods once sold cannot be returned\n".encode('utf-8')
        raw_data += f"Receipt valid in Kenya only\n".encode('utf-8')
        
        # Paper feed and cut
        raw_data += b'\n\n\n'
        raw_data += b'\x1D\x56\x41\x03'
        
        # 2. Send to printer with better error handling for Waitress
        print_success = False
        printer_name = None
        error_message = None
        
        # Try Windows printing - properly handle missing modules
        try:
            import win32print
            
            printer_name = win32print.GetDefaultPrinter()
            if printer_name:
                hprinter = win32print.OpenPrinter(printer_name)
                try:
                    win32print.StartDocPrinter(hprinter, 1, ("POS Receipt", None, "RAW"))
                    win32print.StartPagePrinter(hprinter)
                    win32print.WritePrinter(hprinter, raw_data)
                    win32print.EndPagePrinter(hprinter)
                    win32print.EndDocPrinter(hprinter)
                    print_success = True
                except Exception as e:
                    error_message = f"Windows write error: {str(e)}"
                finally:
                    try:
                        win32print.ClosePrinter(hprinter)
                    except:
                        pass
            else:
                error_message = "No default printer found"
                
        except ImportError:
            error_message = "Windows printing not available (pywin32 not installed)"
        except Exception as e:
            error_message = f"Windows printing error: {str(e)}"
        
        # If direct printing failed, fall back to browser-based printing
        if not print_success:
            # Use the correct URL name from your urls.py
            return JsonResponse({
                'success': True,
                'use_browser_print': True,
                'message': 'Redirecting to browser print...',
                'print_url': reverse('sale_detail', args=[sale.id]),  # Changed from 'sale_invoice' to 'sale_detail'
                'time_printed': nairobi_time_str,
                'date_printed': nairobi_date_str
            })
        
        return JsonResponse({
            'success': True,
            'message': 'Receipt printed successfully!',
            'printer': printer_name,
            'time_printed': nairobi_time_str,
            'date_printed': nairobi_date_str
        })
    
    except Exception as e:
        print(f"Printing error details: {str(e)}")
        import traceback
        traceback.print_exc()
        
        # Use the correct URL name in the error handler too
        return JsonResponse({
            'success': False,
            'message': f'Printing error: {str(e)}',
            'use_browser_print': True,  # Fallback to browser
            'print_url': reverse('sale_detail', args=[sale_id]) if 'sale_id' in locals() else None
        }, status=200)  # Return 200 even on error to allow browser fallback


@login_required
def thermal_print_receipt(request, sale_id):
    """Generate thermal printer friendly receipt with high contrast"""
    from django.shortcuts import get_object_or_404
    from django.http import HttpResponse
    from django.utils import timezone
    from pos.models import Sale, Company
    
    sale = get_object_or_404(Sale, pk=sale_id)
    company = Company.objects.first()
    
    # Use current time
    current_time = timezone.now().strftime('%d/%m/%Y %H:%M:%S')
    
    # Calculate change
    change = max(sale.amount_paid - sale.total, 0)
    
    html = f'''
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <title>Receipt #{sale.sale_number}</title>
        <style>
            /* Thermal printer optimal settings */
            @page {{
                size: 80mm auto;
                margin: 0;
            }}
            
            body {{ 
                font-family: 'Courier New', monospace, 'Consolas', 'Lucida Console';
                font-size: 12px;
                width: 80mm;
                margin: 0;
                padding: 8px;
                -webkit-print-color-adjust: exact !important;
                print-color-adjust: exact !important;
                color-adjust: exact !important;
            }}
            
            /* Make everything darker/bolder */
            * {{
                -webkit-print-color-adjust: exact;
                print-color-adjust: exact;
            }}
            
            .center {{ text-align: center; }}
            .line {{ 
                border-top: 2px solid #000; 
                margin: 5px 0;
                -webkit-print-color-adjust: exact;
                print-color-adjust: exact;
            }}
            
            .dashed-line {{
                border-top: 1px dashed #000;
                margin: 5px 0;
                -webkit-print-color-adjust: exact;
                print-color-adjust: exact;
            }}
            
            table {{ 
                width: 100%;
                border-collapse: collapse;
            }}
            
            th, td {{ 
                text-align: left; 
                padding: 4px 2px;
                font-weight: bold;
            }}
            
            th {{
                border-bottom: 2px solid #000;
                font-weight: bold;
            }}
            
            .right {{ text-align: right; }}
            .bold {{ font-weight: bold; }}
            .large {{ font-size: 14px; font-weight: bold; }}
            .header-store {{ font-size: 16px; font-weight: bold; margin-bottom: 5px; }}
            
            /* Ensure all text is dark */
            div, span, p, td, th {{
                color: #000000 !important;
                font-weight: bold !important;
            }}
            
            .total-amount {{
                font-size: 14px;
                font-weight: bold;
            }}
            
            /* Barcode styling */
            .barcode {{
                font-family: 'Courier New', monospace;
                font-weight: bold;
                letter-spacing: 1px;
            }}
        </style>
    </head>
    <body>
        <div class="center">
            <div class="header-store">{company.name if company else 'MINIMART POS'}</div>
            <div>{company.address if company else ''}</div>
            <div>Tel: {company.phone if company else ''}</div>
            <div class="dashed-line"></div>
            <div class="bold">RECEIPT</div>
            <div>#{sale.sale_number}</div>
            <div>{current_time}</div>
            <div>Cashier: {sale.user.username if sale.user else ''}</div>
            <div>Customer: {sale.customer.name if sale.customer else 'Walk-in'}</div>
            <div class="dashed-line"></div>
        </div>
        
        <table>
            <thead>
                <tr>
                    <th>Item</th>
                    <th class="right">Qty</th>
                    <th class="right">Price</th>
                    <th class="right">Total</th>
                </tr>
            </thead>
            <tbody>
    '''
    
    for item in sale.items.all():
        item_name = (item.product.name[:20] if item.product else '') + ('...' if len(item.product.name) > 20 else '')
        html += f'''
            <tr>
                <td><span class="bold">{item_name}</span></td>
                <td class="right bold">{item.quantity}</td>
                <td class="right bold">{item.price:.2f}</td>
                <td class="right bold">{item.total:.2f}</td>
            </tr>
        '''
    
    html += f'''
            </tbody>
        </table>
        
        <div class="dashed-line"></div>
        
        <div class="right">
            <div>Subtotal: <span class="bold">{sale.subtotal:.2f}</span></div>
    '''
    
    if hasattr(sale, 'discount_amount') and sale.discount_amount > 0:
        html += f'<div>Discount: -{sale.discount_amount:.2f}</div>'
    
    html += f'''
            <div class="line"></div>
            <div class="total-amount">TOTAL: KSh {sale.total:.2f}</div>
            <div>Paid: {sale.amount_paid:.2f}</div>
    '''
    
    if change > 0:
        html += f'<div class="bold">Change: {change:.2f}</div>'
    
    # Payment method
    payment_method = "Cash"
    if hasattr(sale, 'payment_method'):
        if sale.payment_method == 'mpesa':
            payment_method = "M-Pesa"
        elif sale.payment_method == 'card':
            payment_method = "Card"
        elif sale.payment_method == 'credit':
            payment_method = "Credit"
    
    html += f'''
            <div>Payment: {payment_method}</div>
        </div>
        
        <div class="dashed-line"></div>
        
        <div class="center">
            <div class="bold">THANK YOU!</div>
            <div>Goods once sold cannot be returned</div>
            <div>Receipt required for warranty</div>
            <div class="dashed-line"></div>
            <div class="barcode">{sale.sale_number}</div>
        </div>
        
        <script>
            // Auto-print with better thermal printer support
            window.onload = function() {{
                // Add a small delay to ensure CSS loads
                setTimeout(function() {{
                    window.print();
                }}, 1000);
                
                // Close after printing
                window.onafterprint = function() {{
                    window.close();
                }};
                
                // Fallback close after 10 seconds
                setTimeout(function() {{
                    window.close();
                }}, 15000);
            }};
        </script>
    </body>
    </html>
    '''
    
    return HttpResponse(html)


@login_required
def print_receipt(request, receipt_id):
    """Redirect to thermal print"""
    from django.shortcuts import redirect
    from django.urls import reverse
    from django.contrib import messages
    
    # receipt_id is actually the sale_id
    return redirect('thermal_print_receipt', sale_id=receipt_id)
        
# ============== AJAX & Utility Views ==============
from django.db import models

from django.db.models import Q
from django.views.decorators.http import require_http_methods

@login_required
@require_http_methods(["GET", "POST"])  # Allow both GET and POST
def search_products(request):
    """Search products for purchase (include all products, even with 0 quantity)"""
    # Get query from either GET or POST
    if request.method == 'POST':
        query = request.POST.get('q', '').strip()
    else:  # GET
        query = request.GET.get('q', '').strip()
    
    # Get category filter if provided
    if request.method == 'POST':
        category = request.POST.get('category', 'all')
    else:
        category = request.GET.get('category', 'all')
    
    if query:
        # Base query
        products = Product.objects.filter(
            Q(name__icontains=query) |
            Q(barcode__icontains=query)
        )
        
        # Apply category filter if not 'all'
        if category and category != 'all':
            products = products.filter(category_id=category)
        
        products = products.order_by('name')[:10]
    else:
        products = Product.objects.none()
    
    # Create results including stock info
    results = []
    for product in products:
        # Only include active products for POS
        if not product.is_active:
            continue
            
        results.append({
            'id': product.id,
            'name': product.name,
            'barcode': product.barcode or '',
            'quantity': str(product.quantity),  # Current stock
            'selling_price': str(product.selling_price),
            'purchase_price': float(product.purchase_price),  # Make sure this is included
            'wholesale_price': str(product.wholesale_price) if hasattr(product, 'wholesale_price') else '0',
            'wholesale_min_quantity': str(product.wholesale_min_quantity) if hasattr(product, 'wholesale_min_quantity') else '1',
            'category_name': product.category.name if product.category else '',
            'is_active': product.is_active
        })
    
    return JsonResponse(results, safe=False)

@login_required
@require_POST
def search_customers(request):
    query = request.POST.get('q', '')
    
    if query:
        customers = Customer.objects.filter(
            Q(name__icontains=query) |
            Q(phone__icontains=query)
        ).order_by('name')[:10]
    else:
        customers = Customer.objects.none()
    
    results = [{
        'id': c.id,
        'name': c.name,
        'phone': c.phone or '',
        'balance': str(c.balance)
    } for c in customers]
    
    return JsonResponse(results, safe=False)

@login_required
@require_GET
def get_product_details(request, pk):
    product = get_object_or_404(Product, pk=pk)
    sale_type = request.GET.get('sale_type', 'retail')
    
    price = product.selling_price
    if sale_type == 'wholesale' and hasattr(product, 'wholesale_price'):
        price = product.wholesale_price
    
    data = {
        'id': product.id,
        'name': product.name,
        'price': str(price),
        'quantity': product.quantity,
        'batches': [{
            'id': b.id,
            'batch_number': b.batch_number,
            'quantity': b.quantity,
            'expiry_date': b.expiry_date.strftime('%Y-%m-%d') if b.expiry_date else ''
        } for b in product.batches.filter(quantity__gt=0)]
    }
    
    return JsonResponse(data)

@login_required
@require_GET
def get_customer_details(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    
    data = {
        'id': customer.id,
        'name': customer.name,
        'phone': customer.phone or '',
        'credit_limit': str(customer.credit_limit),
        'balance': str(customer.balance)
    }
    
    return JsonResponse(data)

@login_required
@require_GET
def get_purchase_items(request, purchase_id):
    purchase = get_object_or_404(Purchase, pk=purchase_id)
    items = purchase.items.all()
    
    results = [{
        'id': item.id,
        'product_id': item.product.id,
        'product_name': item.product.name,
        'batch_id': item.batch.id if item.batch else None,
        'batch_number': item.batch.batch_number if item.batch else '',
        'quantity': item.quantity,
        'price': str(item.price),
        'max_returnable': item.quantity - (item.returned_quantity if hasattr(item, 'returned_quantity') else 0)
    } for item in items]
    
    return JsonResponse(results, safe=False)

@login_required
def get_batches_for_product(request):
    product_id = request.GET.get('product_id')
    batches = Batch.objects.filter(product_id=product_id).order_by('-expiry_date')
    
    options = '<option value="">Select Batch</option>'
    for batch in batches:
        expiry_info = f" (Exp: {batch.expiry_date.strftime('%Y-%m-%d')})" if batch.expiry_date else ''
        options += f'<option value="{batch.id}">{batch.batch_number}{expiry_info} - Qty: {batch.quantity}</option>'
    
    return HttpResponse(options)

# ============== Bulk Upload Views ==============
import pandas as pd
from decimal import Decimal
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from .models import Product, Category, Supplier, Customer
import re
from datetime import datetime

@login_required
def bulk_upload(request):
    if request.method == 'POST':
        form = BulkUploadForm(request.POST, request.FILES)
        
        if form.is_valid():
            try:
                file = request.FILES['file']
                model_type = form.cleaned_data['model_type']
                
                # Read the Excel file
                try:
                    df = pd.read_excel(file)
                    # Convert all columns to string to handle mixed types
                    df = df.astype(str)
                except Exception as e:
                    messages.error(request, f'Error reading Excel file: {str(e)}')
                    return redirect('bulk_upload')
                
                records_created = 0
                errors = []
                
                def clean_numeric(value, default=0):
                    """Extract numeric value from strings like '38 pcs'"""
                    if pd.isna(value) or value == 'nan' or value == 'None':
                        return default
                    try:
                        # Extract first numeric part
                        match = re.search(r'(\d+\.?\d*)', str(value))
                        if match:
                            return Decimal(match.group(1))
                        return Decimal(value)
                    except:
                        return default
                
                def clean_text(value):
                    """Clean text fields"""
                    if pd.isna(value) or value == 'nan' or value == 'None':
                        return ''
                    return str(value).strip()
                
                # PRODUCT IMPORT
                if model_type == 'product':
                    required_fields = ['Name', 'Purchase Price', 'Selling Price']
                    for field in required_fields:
                        if field not in df.columns:
                            messages.error(request, f'Missing required column: {field}')
                            return redirect('bulk_upload')
                    
                    for index, row in df.iterrows():
                        try:
                            # Required fields
                            name = clean_text(row.get('Name'))
                            if not name:
                                errors.append(f"Row {index+1}: Missing product name")
                                continue
                                
                            purchase_price = clean_numeric(row.get('Purchase Price'))
                            selling_price = clean_numeric(row.get('Selling Price'))
                            
                            # Optional fields with defaults
                            wholesale_price = clean_numeric(row.get('Wholesale Price', selling_price))
                            quantity = int(clean_numeric(row.get('Qty', 0)))
                            reorder_level = int(clean_numeric(row.get('Reorder Level', 5)))
                            barcode = clean_text(row.get('Barcode', ''))
                            category_name = clean_text(row.get('Category', 'Uncategorized'))
                            
                            # Get or create category
                            category = None
                            if category_name:
                                category, _ = Category.objects.get_or_create(name=category_name)
                            
                            # Handle supplier if provided
                            supplier = None
                            supplier_name = clean_text(row.get('Supplier'))
                            if supplier_name:
                                supplier, _ = Supplier.objects.get_or_create(name=supplier_name)
                            
                            # Create product
                            Product.objects.create(
                                name=name,
                                barcode=barcode,
                                category=category,
                                purchase_price=purchase_price,
                                selling_price=selling_price,
                                wholesale_price=wholesale_price,
                                quantity=quantity,
                                reorder_level=reorder_level,
                                supplier=supplier,
                                is_active=True
                            )
                            records_created += 1
                            
                        except Exception as e:
                            errors.append(f"Row {index+1}: {str(e)}")
                
                # CUSTOMER IMPORT
                elif model_type == 'customer':
                    required_fields = ['Name']
                    for field in required_fields:
                        if field not in df.columns:
                            messages.error(request, f'Missing required column: {field}')
                            return redirect('bulk_upload')
                    
                    for index, row in df.iterrows():
                        try:
                            name = clean_text(row.get('Name'))
                            if not name:
                                errors.append(f"Row {index+1}: Missing customer name")
                                continue
                                
                            Customer.objects.create(
                                name=name,
                                phone=clean_text(row.get('Phone', '')),
                                email=clean_text(row.get('Email', '')),
                                address=clean_text(row.get('Address', '')),
                                credit_limit=clean_numeric(row.get('Credit Limit', 0))
                            )
                            records_created += 1
                            
                        except Exception as e:
                            errors.append(f"Row {index+1}: {str(e)}")
                
                # SUPPLIER IMPORT
                elif model_type == 'supplier':
                    required_fields = ['Name']
                    for field in required_fields:
                        if field not in df.columns:
                            messages.error(request, f'Missing required column: {field}')
                            return redirect('bulk_upload')
                    
                    for index, row in df.iterrows():
                        try:
                            name = clean_text(row.get('Name'))
                            if not name:
                                errors.append(f"Row {index+1}: Missing supplier name")
                                continue
                                
                            Supplier.objects.create(
                                name=name,
                                contact_person=clean_text(row.get('Contact Person', '')),
                                phone=clean_text(row.get('Phone', '')),
                                email=clean_text(row.get('Email', '')),
                                address=clean_text(row.get('Address', '')),
                                balance=0
                            )
                            records_created += 1
                            
                        except Exception as e:
                            errors.append(f"Row {index+1}: {str(e)}")
                
                # Results handling
                if errors:
                    messages.warning(request, f"Completed with {len(errors)} errors. {records_created} records imported successfully.")
                    request.session['bulk_upload_errors'] = errors[:50]  # Limit to 50 errors
                else:
                    messages.success(request, f'Successfully imported {records_created} {model_type} records!')
                
                return redirect('bulk_upload')
            
            except Exception as e:
                messages.error(request, f'System error during import: {str(e)}')
                return redirect('bulk_upload')
        else:
            messages.error(request, 'Invalid form submission')
            return redirect('bulk_upload')
    
    else:
        form = BulkUploadForm()
    
    # Get errors from session if they exist
    errors = request.session.pop('bulk_upload_errors', [])
    
    return render(request, 'pos/bulk_upload.html', {
        'form': form,
        'errors': errors
    })
    
# ============== Authentication Views ==============
from django.contrib.auth import login, logout, authenticate, update_session_auth_hash
from django.contrib.auth.forms import AuthenticationForm, PasswordChangeForm, PasswordResetForm, SetPasswordForm
from django.contrib.auth.tokens import default_token_generator
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.mail import send_mail
from django.shortcuts import render, redirect
from django.template.loader import render_to_string
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes, force_str
from django.conf import settings
from django.db.models.query_utils import Q

def custom_login(request):
    # If user is already authenticated, redirect to POS
    if request.user.is_authenticated:
        return redirect('pos')
    
    error = None
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                # Redirect to the next page if provided, otherwise to POS
                next_page = request.GET.get('next', 'pos')
                return redirect(next_page)
        else:
            # Form is invalid, show error
            error = 'Invalid username or password. Please try again.'
    
    return render(request, 'pos/login.html', {'error': error})

def custom_logout(request):
    logout(request)
    return redirect('custom_login')

@login_required
def password_change(request):
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            # Update session to prevent logout
            update_session_auth_hash(request, user)
            messages.success(request, 'Your password has been changed successfully!')
            return redirect('pos')
        else:
            messages.error(request, 'Please correct the error below.')
    else:
        form = PasswordChangeForm(request.user)
    
    return render(request, 'pos/password_change.html', {'form': form})

def password_reset_request(request):
    if request.method == 'POST':
        form = PasswordResetForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data['email']
            associated_users = User.objects.filter(Q(email=email))
            if associated_users.exists():
                for user in associated_users:
                    subject = "Password Reset Request - POS System"
                    email_template_name = "pos/password_reset_email.html"
                    context = {
                        'email': user.email,
                        'domain': request.get_host(),
                        'site_name': 'POS System',
                        'uid': urlsafe_base64_encode(force_bytes(user.pk)),
                        'user': user,
                        'token': default_token_generator.make_token(user),
                        'protocol': 'https' if request.is_secure() else 'http',
                    }
                    email_content = render_to_string(email_template_name, context)
                    
                    try:
                        send_mail(
                            subject,
                            email_content,
                            settings.DEFAULT_FROM_EMAIL,
                            [user.email],
                            fail_silently=False,
                            html_message=email_content
                        )
                    except Exception as e:
                        messages.error(request, f'Error sending email: {str(e)}')
                        return redirect('password_reset_request')
                
                messages.success(request, 'Password reset link has been sent to your email.')
                return redirect('custom_login')
            else:
                messages.error(request, 'No account found with that email address.')
    else:
        form = PasswordResetForm()
    
    return render(request, 'pos/password_reset.html', {'form': form})

def password_reset_confirm(request, uidb64=None, token=None):
    try:
        uid = force_str(urlsafe_base64_decode(uidb64))
        user = User.objects.get(pk=uid)
    except (TypeError, ValueError, OverflowError, User.DoesNotExist):
        user = None

    if user is not None and default_token_generator.check_token(user, token):
        if request.method == 'POST':
            form = SetPasswordForm(user, request.POST)
            if form.is_valid():
                form.save()
                messages.success(request, 'Your password has been reset successfully. You can now login with your new password.')
                return redirect('custom_login')
            else:
                messages.error(request, 'Please correct the error below.')
        else:
            form = SetPasswordForm(user)
        
        return render(request, 'pos/password_reset_confirm.html', {'form': form})
    else:
        messages.error(request, 'The password reset link is invalid or has expired.')
        return redirect('password_reset_request')
    


# ============== Reports Dashboard ==============
@login_required
def reports_dashboard(request):
    """Main reports dashboard that links to all report types"""
    return render(request, 'pos/reports_dashboard.html')

from django.db.models import Sum, Count, F, Q
from django.db.models.functions import TruncDate
from datetime import timedelta
from decimal import Decimal

@login_required
def dashboard(request):
    # Date range handling
    date_range = request.GET.get('range', '30d')
    
    if date_range == '7d':
        days = 7
        start_date = timezone.now() - timedelta(days=days)
    elif date_range == '90d':
        days = 90
        start_date = timezone.now() - timedelta(days=days)
    elif date_range == 'ytd':
        start_date = timezone.now().replace(month=1, day=1)
    else:
        days = 30
        start_date = timezone.now() - timedelta(days=days)
    
    end_date = timezone.now()
    
    # Today's sales data
    today = timezone.now().date()
    yesterday = today - timedelta(days=1)
    
    today_sales = Sale.objects.filter(
        date__date=today,
        is_completed=True
    ).aggregate(total=Sum('total'))['total'] or Decimal('0')
    
    yesterday_sales = Sale.objects.filter(
        date__date=yesterday,
        is_completed=True
    ).aggregate(total=Sum('total'))['total'] or Decimal('0.01')
    
    today_vs_yesterday = float(((today_sales - yesterday_sales) / yesterday_sales) * 100) if yesterday_sales != 0 else 0
    
    # Monthly sales data
    this_month = timezone.now().date().replace(day=1)
    last_month = (this_month - timedelta(days=1)).replace(day=1)
    
    monthly_sales = Sale.objects.filter(
        date__date__gte=this_month,
        is_completed=True
    ).aggregate(total=Sum('total'))['total'] or Decimal('0')
    
    last_month_sales = Sale.objects.filter(
        date__date__gte=last_month,
        date__date__lt=this_month,
        is_completed=True
    ).aggregate(total=Sum('total'))['total'] or Decimal('0.01')
    
    monthly_vs_last = float(((monthly_sales - last_month_sales) / last_month_sales) * 100) if last_month_sales != 0 else 0
    
    # Low stock items
    low_stock_count = Product.objects.filter(
        quantity__lte=F('reorder_level'),
        is_active=True
    ).count()
    
    # Outstanding credit
    outstanding_credit = Sale.objects.filter(
        is_credit=True,
        balance__gt=0
    ).aggregate(total=Sum('balance'))['total'] or Decimal('0')
    
    # Recent sales
    recent_sales = Sale.objects.filter(
        is_completed=True
    ).order_by('-date')[:10]
    
    # Sales trend data
    sales_trend = Sale.objects.filter(
        date__date__gte=start_date,
        is_completed=True
    ).annotate(
        day=TruncDate('date')
    ).values('day').annotate(
        total=Sum('total')
    ).order_by('day')
    
    sales_trend_labels = [sale['day'].strftime('%b %d') for sale in sales_trend]
    sales_trend_data = [float(sale['total'] or 0) for sale in sales_trend]
    
    # Payment methods breakdown
    payment_methods = Sale.objects.filter(
        date__date__gte=start_date,
        is_completed=True
    ).values('payment_method').annotate(
        total=Sum('total'),
        count=Count('id')
    ).order_by('-total')
    
    payment_method_labels = []
    payment_method_data = []
    payment_method_colors = []
    payment_method_colors_hover = []
    
    colors = {
        'cash': ['#4e73df', '#2e59d9'],
        'mpesa': ['#1cc88a', '#17a673'],
        'card': ['#36b9cc', '#2c9faf'],
        'cheque': ['#f6c23e', '#dda20a'],
        'credit': ['#e74a3b', '#be2617']
    }
    
    for method in payment_methods:
        payment_method_labels.append(method['payment_method'].capitalize())
        payment_method_data.append(float(method['total'] or 0))
        payment_method_colors.append(colors.get(method['payment_method'], ['#858796'])[0])
        payment_method_colors_hover.append(colors.get(method['payment_method'], ['#707384'])[0])
    
    # SIMPLIFIED: Fast moving products (just get top 5 by quantity sold)
    fast_moving_products = Product.objects.filter(
        saleitem__sale__date__date__gte=start_date,
        saleitem__sale__is_completed=True
    ).annotate(
        total_sold=Sum('saleitem__quantity')
    ).order_by('-total_sold')[:5]
    
    # SIMPLIFIED: Slow moving products (products not sold in last 30 days)
    thirty_days_ago = timezone.now() - timedelta(days=30)
    slow_moving_products = Product.objects.filter(
        is_active=True,
        quantity__gt=0
    ).exclude(
        saleitem__sale__date__gte=thirty_days_ago,
        saleitem__sale__is_completed=True
    ).order_by('?')[:5]  # Random 5 products
    
    context = {
        'today_sales': today_sales,
        'today_vs_yesterday': today_vs_yesterday,
        'today_vs_yesterday_abs': abs(today_vs_yesterday),
        'monthly_sales': monthly_sales,
        'monthly_vs_last': monthly_vs_last,
        'monthly_vs_last_abs': abs(monthly_vs_last),
        'low_stock_count': low_stock_count,
        'outstanding_credit': outstanding_credit,
        'recent_sales': recent_sales,
        'sales_trend_labels': sales_trend_labels,
        'sales_trend_data': sales_trend_data,
        'payment_method_labels': payment_method_labels,
        'payment_method_data': payment_method_data,
        'payment_method_colors': payment_method_colors,
        'payment_method_colors_hover': payment_method_colors_hover,
        'payment_methods': [
            {'name': m['payment_method'].capitalize(), 'color': colors.get(m['payment_method'], ['#858796'])[0]}
            for m in payment_methods
        ],
        'fast_moving_products': fast_moving_products,
        'slow_moving_products': slow_moving_products,
    }
    return render(request, 'pos/dashboard.html', context)

from django.http import JsonResponse
from .models import Product, CompanyPrice

def get_product_pricing(request, pk):
    try:
        product = Product.objects.get(pk=pk)
        company_id = request.GET.get('company_id')
        
        if company_id:
            try:
                company_price = CompanyPrice.objects.get(
                    product=product, 
                    company_id=company_id
                )
                price = company_price.price
            except CompanyPrice.DoesNotExist:
                price = product.selling_price
        else:
            price = product.selling_price
        
        response_data = {
            'product_id': product.id,
            'selling_price': str(product.selling_price),
            'wholesale_price': str(product.wholesale_price),
            'wholesale_min_quantity': product.wholesale_min_quantity,
            'final_price': str(price),
            'quantity': product.quantity
        }
        
        return JsonResponse(response_data)
    
    except Product.DoesNotExist:
        return JsonResponse({'error': 'Product not found'}, status=404)
    



from django.shortcuts import render
from django.utils import timezone
from .models import Sale, Purchase, Expense
from datetime import timedelta

def reports(request):
    # Default to last 30 days
    end_date = timezone.now()
    start_date = end_date - timedelta(days=30)
    
    # Get filter parameters from request
    report_type = request.GET.get('type', 'sales')
    date_range = request.GET.get('range', '30d')
    
    # Adjust date range based on selection
    if date_range == '7d':
        start_date = end_date - timedelta(days=7)
    elif date_range == '90d':
        start_date = end_date - timedelta(days=90)
    elif date_range == 'ytd':
        start_date = end_date.replace(month=1, day=1)
    
    # Generate report data based on type
    if report_type == 'sales':
        data = Sale.objects.filter(
            date__range=[start_date, end_date]
        ).order_by('-date')
        title = "Sales Report"
    elif report_type == 'purchases':
        data = Purchase.objects.filter(
            date__range=[start_date, end_date]
        ).order_by('-date')
        title = "Purchases Report"
    elif report_type == 'expenses':
        data = Expense.objects.filter(
            date__range=[start_date, end_date]
        ).order_by('-date')
        title = "Expenses Report"
    else:
        data = []
        title = "Report"
    
    context = {
        'report_type': report_type,
        'date_range': date_range,
        'start_date': start_date,
        'end_date': end_date,
        'data': data,
        'title': title
    }
    
    return render(request, 'pos/reports.html', context)


from django.shortcuts import render
from django.utils import timezone
from django.db.models import Sum, Count, Q
from .models import Sale, Expense, Purchase, Customer
from datetime import datetime, timedelta
import csv
from django.http import HttpResponse

# views.py
from django.shortcuts import render, redirect
from django.contrib import messages
from django.db.models import Sum, Count, Q
from .models import Sale, Customer
from datetime import datetime, timedelta
import csv
from django.http import HttpResponse
from django.core.paginator import Paginator



@login_required
def daily_sales_report(request):
    # Date filter - default to today
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    try:
        date_from = datetime.strptime(date_from, '%Y-%m-%d').date() if date_from else None
        date_to = datetime.strptime(date_to, '%Y-%m-%d').date() if date_to else None
    except:
        date_from = date_to = None
    
    if not date_from and not date_to:
        date_to = timezone.now().date()
        date_from = date_to
    
    # Base queryset with date filter
    sales = Sale.objects.filter(
        date__date__gte=date_from,
        date__date__lte=date_to,
        is_completed=True
    ).select_related('customer', 'user').prefetch_related('items')
    
    # Apply additional filters
    payment_method = request.GET.get('payment_method')
    if payment_method:
        if payment_method == 'credit':
            sales = sales.filter(is_credit=True)
        else:
            sales = sales.filter(payment_method=payment_method)
    
    customer_id = request.GET.get('customer')
    if customer_id:
        sales = sales.filter(customer_id=customer_id)
    
    min_amount = request.GET.get('min_amount')
    if min_amount:
        try:
            sales = sales.filter(total__gte=Decimal(min_amount))
        except (ValueError, InvalidOperation):
            pass
    
    # Pagination
    paginator = Paginator(sales.order_by('-date'), 25)  # Show 25 sales per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Calculate totals from the PAGINATED sales data (what's actually displayed)
    total_sales = Decimal('0.00')
    total_items = 0
    transaction_count = page_obj.paginator.count  # Total transactions across all pages
    
    # Calculate totals from displayed page only for the stats
    for sale in page_obj:
        total_sales += sale.total
        total_items += sale.items.count()
    
    # Calculate average transaction
    average_transaction = total_sales / transaction_count if transaction_count > 0 else Decimal('0.00')
    
    # Get payment method breakdown from ALL sales (not just displayed page)
    payment_breakdown = sales.values('payment_method').annotate(
        total=Sum('total'),
        count=Count('id')
    )
    
    # Calculate percentages for each payment method
    total_all_sales = sales.aggregate(total=Sum('total'))['total'] or Decimal('0.00')
    for method in payment_breakdown:
        if total_all_sales > 0:
            method['percentage'] = float(method['total'] / total_all_sales * 100)
        else:
            method['percentage'] = 0
    
    # Get top selling products for the day from ALL sales
    top_products = SaleItem.objects.filter(
        sale__in=sales
    ).values(
        'product__name'
    ).annotate(
        quantity_sold=Sum('quantity'),
        revenue=Sum('total')
    ).order_by('-revenue')[:10]
    
    # Get all customers for filter dropdown
    customers = Customer.objects.all().order_by('name')
    
    context = {
        'sales': page_obj,
        'customers': customers,
        'date_from': date_from.strftime('%Y-%m-%d') if date_from else '',
        'date_to': date_to.strftime('%Y-%m-%d') if date_to else '',
        'total_sales': total_sales,
        'total_items': total_items,
        'transaction_count': transaction_count,
        'average_transaction': average_transaction,
        'payment_breakdown': payment_breakdown,
        'top_products': top_products,
        'payment_methods': [
            ('cash', 'Cash'),
            ('mpesa', 'M-Pesa'),
            ('card', 'Card'),
            ('cheque', 'Cheque'),
            ('credit', 'Credit')
        ],
        'selected_payment_method': payment_method,
        'selected_customer': customer_id,
        'selected_min_amount': min_amount
    }
    return render(request, 'pos/daily_sales_report.html', context)


@login_required
def export_daily_sales(request):
    # Get filter parameters from request
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    payment_method = request.GET.get('payment_method')
    customer_id = request.GET.get('customer')
    min_amount = request.GET.get('min_amount')

    # Initialize base queryset
    sales = Sale.objects.filter(is_completed=True)
    
    # Apply date filters only if dates are provided
    if date_from:
        try:
            date_from = datetime.strptime(date_from, '%Y-%m-%d').date()
            sales = sales.filter(date__date__gte=date_from)
        except ValueError:
            pass  # Invalid date format, skip this filter

    if date_to:
        try:
            date_to = datetime.strptime(date_to, '%Y-%m-%d').date()
            sales = sales.filter(date__date__lte=date_to)
        except ValueError:
            pass  # Invalid date format, skip this filter

    # Apply payment method filter
    if payment_method:
        if payment_method == 'credit':
            sales = sales.filter(is_credit=True)
        else:
            sales = sales.filter(payment_method=payment_method)

    # Apply customer filter
    if customer_id:
        sales = sales.filter(customer_id=customer_id)

    # Apply minimum amount filter
    if min_amount:
        try:
            sales = sales.filter(total__gte=float(min_amount))
        except (ValueError, TypeError):
            pass  # Invalid amount, skip this filter

    # Order the results
    sales = sales.order_by('-date').select_related('customer', 'user').prefetch_related('items')

    # Generate filename with date range if available
    filename = "sales_export"
    if date_from and date_to:
        filename += f"_{date_from}_to_{date_to}"
    elif date_from:
        filename += f"_from_{date_from}"
    elif date_to:
        filename += f"_to_{date_to}"
    filename += ".csv"

    # Create CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    writer = csv.writer(response)
    
    # Write headers - removed discount and tax_amount since they don't exist
    writer.writerow([
        'Receipt Number', 'Date', 'Time', 'Customer', 'Customer Phone', 
        'Items Count', 'Total Amount', 'Payment Method', 'Is Credit', 
        'Cashier'  # Removed: 'Discount', 'Tax Amount'
    ])
    
    # Write data rows
    for sale in sales:
        writer.writerow([
            sale.sale_number,
            sale.date.date(),
            sale.date.time(),
            sale.customer.name if sale.customer else 'Walk-in',
            sale.customer.phone if sale.customer else '',
            sale.items.count(),
            sale.total,
            sale.get_payment_method_display(),
            'Yes' if sale.is_credit else 'No',
            sale.user.get_full_name() or sale.user.username,
            # Removed: sale.discount if hasattr(sale, 'discount') else 0,
            # Removed: sale.tax_amount if hasattr(sale, 'tax_amount') else 0
        ])
    
    return response

from .forms import StockJournalForm
from .models import StockJournal

@login_required
def stock_journal_list(request):
    # Get journals with their items and user
    journals = StockJournal.objects.all().select_related('user').prefetch_related('items__product', 'items__batch').order_by('-date')
    
    # Filtering
    product_id = request.GET.get('product')
    movement_type = request.GET.get('movement_type')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if product_id:
        journals = journals.filter(items__product_id=product_id).distinct()
    
    # FIXED: Get movement types from StockJournalItem instead of StockJournal
    if movement_type:
        journals = journals.filter(items__movement_type=movement_type).distinct()
    
    if start_date:
        journals = journals.filter(date__date__gte=start_date)
    if end_date:
        journals = journals.filter(date__date__lte=end_date)
    
    # Export functionality
    export_format = request.GET.get('export')
    if export_format == 'csv':
        return generate_stock_journal_export(journals)
    
    # Pagination
    paginator = Paginator(journals, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    products = Product.objects.filter(is_active=True).order_by('name')
    
    # FIXED: Get movement types from StockJournalItem model
    from .models import StockJournalItem
    
    context = {
        'journals': page_obj,
        'products': products,
        'movement_types': StockJournalItem.MOVEMENT_TYPES,  # Changed to StockJournalItem
        'selected_product': product_id,
        'selected_movement_type': movement_type,
        'start_date': start_date,
        'end_date': end_date,
    }
    return render(request, 'pos/stock_journal_list.html', context)

    
@login_required
def add_stock_journal(request):
    if request.method == 'POST':
        journal_form = StockJournalForm(request.POST)
        formset = StockJournalItemFormSet(request.POST)
        
        if journal_form.is_valid() and formset.is_valid():
            with transaction.atomic():
                # Save the journal
                journal = journal_form.save(commit=False)
                journal.user = request.user
                journal.save()
                
                # Process formset instances
                instances = formset.save(commit=False)
                total_items = 0
                
                for instance in instances:
                    instance.journal = journal
                    product = instance.product
                    movement_type = instance.movement_type
                    
                    # Calculate current stock before changes
                    instance.current_stock = product.quantity
                    
                    # Update stock based on movement type
                    if movement_type in ['out', 'missing', 'damaged', 'broken', 'expired']:
                        # Stock reduction
                        instance.new_stock = max(Decimal('0'), product.quantity - instance.quantity)
                        product.quantity = instance.new_stock
                    elif movement_type == 'in':
                        # Stock addition
                        instance.new_stock = product.quantity + instance.quantity
                        product.quantity = instance.new_stock
                    elif movement_type == 'adjustment':
                        # Direct adjustment
                        instance.new_stock = instance.quantity
                        product.quantity = instance.quantity
                    elif movement_type == 'transfer':
                        # Transfer (typically positive)
                        instance.new_stock = product.quantity + instance.quantity
                        product.quantity = instance.new_stock
                    
                    # Save product changes
                    product.save()
                    
                    # Update batch if specified
                    if instance.batch:
                        batch = instance.batch
                        if movement_type in ['out', 'missing', 'damaged', 'broken', 'expired']:
                            batch.quantity = max(Decimal('0'), batch.quantity - instance.quantity)
                        elif movement_type in ['in', 'transfer']:
                            batch.quantity += instance.quantity
                        elif movement_type == 'adjustment':
                            batch.quantity = instance.quantity
                        batch.save()
                    
                    # Save the instance
                    instance.save()
                    total_items += 1
                
                # Update journal with total items count
                journal.total_items = total_items
                journal.save()
                
                # Delete any forms marked for deletion
                for form in formset.deleted_forms:
                    if form.instance.pk:
                        form.instance.delete()
                
                messages.success(request, f'Stock movement recorded successfully! {total_items} products updated.')
                return redirect('stock_journal_list')
        
        else:
            # Show errors
            messages.error(request, 'Please correct the errors below.')
    
    else:
        journal_form = StockJournalForm()
        formset = StockJournalItemFormSet(queryset=StockJournalItem.objects.none())
    
    # Get all products for the dropdown
    products = Product.objects.filter(is_active=True).order_by('name')
    
    context = {
        'journal_form': journal_form,
        'formset': formset,
        'products': products,
        'title': 'Record Stock Movement'
    }
    return render(request, 'pos/stock_journal_create.html', context)



@login_required
def stock_journal_detail(request, pk):
    journal = get_object_or_404(StockJournal.objects.prefetch_related('items__product', 'items__batch'), pk=pk)
    
    context = {
        'journal': journal,
        'title': f'Stock Movement: {journal.movement_number}'
    }
    return render(request, 'pos/stock_journal_detail.html', context)

def generate_stock_journal_export(queryset):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="stock_journal.csv"'
    
    writer = csv.writer(response)
    writer.writerow([
        'Movement Number', 'Date', 'Movement Type', 
        'Product', 'Batch', 'Quantity', 
        'Current Stock', 'New Stock', 
        'Reference', 'Notes', 'User'
    ])
    
    for journal in queryset:
        # Get all items for this journal
        items = journal.items.all()
        if not items.exists():
            # Write at least the journal info
            writer.writerow([
                journal.movement_number,
                journal.date.strftime('%Y-%m-%d %H:%M'),
                journal.get_movement_type_display(),
                '', '', '', '', '',  # Empty product-related fields
                journal.reference,
                journal.notes,
                journal.user.username if journal.user else ''
            ])
        else:
            for item in items:
                writer.writerow([
                    journal.movement_number,
                    journal.date.strftime('%Y-%m-%d %H:%M'),
                    journal.get_movement_type_display(),
                    item.product.name if item.product else '',
                    item.batch.batch_number if item.batch else '',
                    item.quantity,
                    item.current_stock,
                    item.new_stock,
                    journal.reference,
                    item.notes if item.notes else journal.notes,
                    journal.user.username if journal.user else ''
                ])
    
    return response


from django.http import JsonResponse
from django.template.loader import render_to_string

@login_required
def get_batches_for_product(request):
    product_id = request.GET.get('product_id')
    batches = Batch.objects.filter(product_id=product_id).order_by('-expiry_date')
    
    options = '<option value="">---------</option>'
    for batch in batches:
        options += f'<option value="{batch.id}">{batch.batch_number} (Exp: {batch.expiry_date}) - Qty: {batch.quantity}</option>'
    
    return JsonResponse({'options': options})




import pandas as pd
from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import transaction
from decimal import Decimal
from .forms import ProductImportForm

@login_required
def import_products(request):
    if request.method == 'POST':
        form = ProductImportForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                df = pd.read_excel(request.FILES['file'])
                # Replace NaN values with None
                df = df.where(pd.notnull(df), None)
                
                update_existing = form.cleaned_data['update_existing']
                records_created = 0
                records_updated = 0
                errors = []
                
                with transaction.atomic():
                    for index, row in df.iterrows():
                        try:
                            # Required fields
                            name = row.get('Name')
                            if not name:
                                errors.append(f"Row {index+1}: Missing product name")
                                continue
                                
                            # Get or create category
                            category_name = row.get('Category')
                            category = None
                            if category_name:
                                category, _ = Category.objects.get_or_create(name=category_name)
                            
                            # Get or create supplier
                            supplier_name = row.get('Supplier')
                            supplier = None
                            if supplier_name:
                                supplier, _ = Supplier.objects.get_or_create(name=supplier_name)
                            
                            # Convert prices to Decimal
                            purchase_price = Decimal(str(row.get('Purchase Price', 0))) if row.get('Purchase Price') is not None else Decimal(0)
                            selling_price = Decimal(str(row.get('Selling Price', 0))) if row.get('Selling Price') is not None else Decimal(0)
                            wholesale_price = Decimal(str(row.get('Wholesale Price', selling_price))) if row.get('Wholesale Price') is not None else selling_price
                            
                            # Default values
                            defaults = {
                                'category': category,
                                'supplier': supplier,
                                'purchase_price': purchase_price,
                                'selling_price': selling_price,
                                'wholesale_price': wholesale_price,
                                'wholesale_min_quantity': int(row.get('Wholesale Min Quantity', 0)),
                                'quantity': int(row.get('Quantity', 0)),
                                'reorder_level': int(row.get('Reorder Level', 5)),
                                'barcode': row.get('Barcode', ''),
                                'is_active': bool(row.get('Is Active', True))
                            }
                            
                            if update_existing:
                                product, created = Product.objects.update_or_create(
                                    name=name,
                                    defaults=defaults
                                )
                                if created:
                                    records_created += 1
                                else:
                                    records_updated += 1
                            else:
                                # Only create new products
                                if Product.objects.filter(name=name).exists():
                                    errors.append(f"Row {index+1}: Product '{name}' already exists")
                                    continue
                                    
                                Product.objects.create(name=name, **defaults)
                                records_created += 1
                                
                        except Exception as e:
                            errors.append(f"Row {index+1}: {str(e)}")
                
                if errors:
                    messages.warning(request, f"Import completed with {len(errors)} errors")
                    # You might want to save these errors to display to the user
                else:
                    messages.success(request, f'Successfully imported {records_created} products! {records_updated} updated.' if update_existing else f'Successfully imported {records_created} products!')
                return redirect('product_list')
            
            except Exception as e:
                messages.error(request, f'Error during import: {str(e)}')
    else:
        form = ProductImportForm()
    
    return render(request, 'pos/import_products.html', {'form': form})  



@login_required
def opening_closing_stock_report(request):
    # Date filtering
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date') or timezone.now().date()
    
    if not start_date:
        # Default to beginning of current month
        start_date = timezone.now().date().replace(day=1)
    
    # Get products
    products = Product.objects.filter(is_active=True).order_by('name')
    
    # Calculate stock movements during period
    stock_movements = StockJournal.objects.filter(
        date__date__range=[start_date, end_date]
    ).values('product').annotate(
        stock_in=Sum('quantity', filter=Q(movement_type='in')),
        stock_out=Sum('quantity', filter=Q(movement_type='out'))
    )
    
    movement_dict = {movement['product']: movement for movement in stock_movements}
    
    report_data = []
    total_opening_stock = 0
    total_stock_in = 0
    total_stock_out = 0
    total_closing_stock = 0
    total_opening_value = 0
    total_closing_value = 0
    
    for product in products:
        movements = movement_dict.get(product.id, {})
        opening_stock = product.quantity - (movements.get('stock_in', 0) or 0) + (movements.get('stock_out', 0) or 0)
        closing_stock = product.quantity
        stock_in = movements.get('stock_in', 0) or 0
        stock_out = movements.get('stock_out', 0) or 0
        value_change = (closing_stock - opening_stock) * product.purchase_price
        
        report_data.append({
            'product': product,
            'opening_stock': max(opening_stock, 0),
            'closing_stock': closing_stock,
            'stock_in': stock_in,
            'stock_out': stock_out,
            'value_change': value_change
        })
        
        # Update totals
        total_opening_stock += max(opening_stock, 0)
        total_stock_in += stock_in
        total_stock_out += stock_out
        total_closing_stock += closing_stock
        total_opening_value += max(opening_stock, 0) * product.purchase_price
        total_closing_value += closing_stock * product.purchase_price
    
    context = {
        'report_data': report_data,
        'start_date': start_date,
        'end_date': end_date,
        'total_opening_stock': total_opening_stock,
        'total_stock_in': total_stock_in,
        'total_stock_out': total_stock_out,
        'total_closing_stock': total_closing_stock,
        'total_opening_value': total_opening_value,
        'total_closing_value': total_closing_value,
        'total_value_change': total_closing_value - total_opening_value
    }
    return render(request, 'pos/opening_closing_stock.html', context)




from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
import json

# views.py - Update the load_sale_to_pos function
# views.py - Update the load_sale_to_pos function
@login_required
def load_sale_to_pos(request, pk):
    """Load an existing sale into the POS interface for editing"""
    sale = get_object_or_404(Sale, pk=pk, is_completed=True)
    
    # Helper function to safely convert Decimal to float for JSON serialization
    def decimal_to_json(value):
        if isinstance(value, Decimal):
            return float(value)
        elif value is None:
            return 0.0
        else:
            try:
                return float(value)
            except (ValueError, TypeError):
                return 0.0
    
    # Helper function to get payment amounts safely
    def get_payment_amount(sale_obj, field_name):
        try:
            amount = getattr(sale_obj, field_name, Decimal('0'))
            return decimal_to_json(amount)
        except (AttributeError, ValueError, TypeError):
            return 0.0
    
    # Get payment amounts
    cash_amount = get_payment_amount(sale, 'cash_amount')
    mpesa_amount = get_payment_amount(sale, 'mpesa_amount')
    card_amount = get_payment_amount(sale, 'card_amount')
    cheque_amount = get_payment_amount(sale, 'cheque_amount')
    
    # Prepare sale data in format compatible with POS - CONVERT ALL Decimal TO FLOAT
    sale_data = {
        'customer_id': str(sale.customer.id) if sale.customer else None,
        'sale_type': getattr(sale, 'sale_type', 'retail'),
        
        # Items array - convert all Decimal to float
        'items': [
            {
                'id': str(item.product.id),
                'name': item.product.name,
                'barcode': item.product.barcode or '',
                'quantity': decimal_to_json(item.quantity),
                'price': decimal_to_json(item.price),
                'total': decimal_to_json(item.total),
                'batch_id': str(item.batch.id) if item.batch else None,
                'batch_number': item.batch.batch_number if item.batch else '',
                'discount_amount': decimal_to_json(item.discount_amount or Decimal('0')),
                'discount_percent': decimal_to_json(item.discount_percent or Decimal('0')),
                'wholesale_price': decimal_to_json(item.product.wholesale_price) if hasattr(item.product, 'wholesale_price') else 0.0,
                'min_quantity': decimal_to_json(item.product.wholesale_min_quantity) if hasattr(item.product, 'wholesale_min_quantity') else 1
            }
            for item in sale.items.all().select_related('product', 'batch')
        ],
        
        # Financial totals - convert Decimal to float
        'subtotal': decimal_to_json(sale.subtotal),
        'discount_percent': decimal_to_json(getattr(sale, 'discount_percent', Decimal('0'))),
        'discount_amount': decimal_to_json(getattr(sale, 'discount_amount', Decimal('0'))),
        'total': decimal_to_json(sale.total),
        
        # Payment information
        'payment_method': getattr(sale, 'payment_method', 'cash'),
        'amount_paid': decimal_to_json(sale.amount_paid),
        'balance': decimal_to_json(getattr(sale, 'balance', Decimal('0'))),
        'is_credit': bool(getattr(sale, 'is_credit', False)),
        
        # Payment breakdown - already converted to float
        'payment_details': {
            'cash': cash_amount,
            'mpesa': mpesa_amount,
            'card': card_amount,
            'cheque': cheque_amount
        },
        
        # Additional info
        'mpesa_code': getattr(sale, 'mpesa_code', ''),
        'cheque_number': getattr(sale, 'cheque_number', ''),
        'notes': getattr(sale, 'notes', ''),
        
        # Metadata
        'original_sale_id': str(sale.id),
        'original_sale_number': sale.sale_number,
        'date': sale.date.strftime('%Y-%m-%d %H:%M:%S')
    }
    
    # Store in session (Django automatically converts to JSON for session storage)
    request.session['edit_sale_data'] = sale_data
    request.session['edit_sale_id'] = str(sale.id)
    
    # Clear any pending sales for this user
    PendingSale.objects.filter(user=request.user).delete()
    
    messages.success(request, f'Sale #{sale.sale_number} loaded into POS. You can now edit the items.')
    return redirect('pos')

# views.py - Add this function
@login_required
@require_POST
def clear_edit_session(request):
    """Clear edit sale data from session after loading"""
    if 'edit_sale_data' in request.session:
        del request.session['edit_sale_data']
    if 'edit_sale_id' in request.session:
        del request.session['edit_sale_id']
    return JsonResponse({'success': True})



# views.py - Add this function
# views.py - Update the get_edit_sale_data function
@login_required
@require_GET
def get_edit_sale_data(request, pk):
    """Get sale data for editing in POS"""
    sale = get_object_or_404(Sale, pk=pk, is_completed=True)
    
    # Helper function to safely convert Decimal to float for JSON serialization
    def decimal_to_json(value):
        if isinstance(value, Decimal):
            return float(value)
        elif value is None:
            return 0.0
        else:
            try:
                return float(value)
            except (ValueError, TypeError):
                return 0.0
    
    # Helper function to get payment amounts safely
    def get_payment_amount(sale_obj, field_name):
        try:
            amount = getattr(sale_obj, field_name, Decimal('0'))
            return decimal_to_json(amount)
        except (AttributeError, ValueError, TypeError):
            return 0.0
    
    # Get payment amounts
    cash_amount = get_payment_amount(sale, 'cash_amount')
    mpesa_amount = get_payment_amount(sale, 'mpesa_amount')
    card_amount = get_payment_amount(sale, 'card_amount')
    cheque_amount = get_payment_amount(sale, 'cheque_amount')
    
    # Prepare sale data
    sale_data = {
        'customer_id': str(sale.customer.id) if sale.customer else None,
        'sale_type': getattr(sale, 'sale_type', 'retail'),
        
        # Items array - convert all Decimal to float
        'items': [
            {
                'id': str(item.product.id),
                'name': item.product.name,
                'barcode': item.product.barcode or '',
                'quantity': decimal_to_json(item.quantity),
                'price': decimal_to_json(item.price),
                'total': decimal_to_json(item.total),
                'batch_id': str(item.batch.id) if item.batch else None,
                'batch_number': item.batch.batch_number if item.batch else '',
                'discount_amount': decimal_to_json(item.discount_amount or Decimal('0')),
                'discount_percent': decimal_to_json(item.discount_percent or Decimal('0')),
                'wholesale_price': decimal_to_json(item.product.wholesale_price) if hasattr(item.product, 'wholesale_price') else 0.0,
                'min_quantity': decimal_to_json(item.product.wholesale_min_quantity) if hasattr(item.product, 'wholesale_min_quantity') else 1
            }
            for item in sale.items.all().select_related('product', 'batch')
        ],
        
        # Financial totals
        'subtotal': decimal_to_json(sale.subtotal),
        'discount_percent': decimal_to_json(getattr(sale, 'discount_percent', Decimal('0'))),
        'discount_amount': decimal_to_json(getattr(sale, 'discount_amount', Decimal('0'))),
        'total': decimal_to_json(sale.total),
        
        # Payment information
        'payment_method': getattr(sale, 'payment_method', 'cash'),
        'amount_paid': decimal_to_json(sale.amount_paid),
        'balance': decimal_to_json(getattr(sale, 'balance', Decimal('0'))),
        'is_credit': bool(getattr(sale, 'is_credit', False)),
        
        # Payment breakdown
        'payment_details': {
            'cash': cash_amount,
            'mpesa': mpesa_amount,
            'card': card_amount,
            'cheque': cheque_amount
        },
        
        # Additional info
        'mpesa_code': getattr(sale, 'mpesa_code', ''),
        'cheque_number': getattr(sale, 'cheque_number', ''),
        'notes': getattr(sale, 'notes', ''),
        
        # Metadata
        'original_sale_id': str(sale.id),
        'original_sale_number': sale.sale_number,
        'date': sale.date.strftime('%Y-%m-%d %H:%M:%S')
    }
    
    return JsonResponse({
        'success': True,
        'sale_data': sale_data
    })


from django.shortcuts import get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.db import transaction
from django.http import JsonResponse
from django.contrib import messages
from .models import Sale

@login_required
def delete_sale(request, pk):
    """Delete a sale completely"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    sale = get_object_or_404(Sale, pk=pk)
    
    try:
        with transaction.atomic():
            # Restore stock quantities for all items in the sale
            for item in sale.items.all():
                product = item.product
                product.quantity += item.quantity
                product.save()
                
                # Restore batch quantity if exists
                if item.batch:
                    batch = item.batch
                    batch.quantity += item.quantity
                    batch.save()
            
            # Delete the sale
            sale_number = sale.sale_number
            sale.delete()
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': f'Sale #{sale_number} deleted successfully!'
                })
            else:
                messages.success(request, f'Sale #{sale_number} has been deleted successfully!')
                return redirect('sales_report')
    
    except Exception as e:
        error_message = f'Error deleting sale: {str(e)}'
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'message': error_message
            }, status=500)
        messages.error(request, error_message)
        return redirect('sales_report')





@login_required
def product_performance_report(request):
    """
    Comprehensive product performance analysis with quantity sold, revenue, and profits
    """
    try:
        # Date filtering
        date_range = request.GET.get('range', '7d')
        
        # Set default dates
        end_date = timezone.now().date()
        
        if date_range == '1d':
            start_date = end_date
        elif date_range == '30d':
            start_date = end_date - timedelta(days=29)
        elif date_range == '90d':
            start_date = end_date - timedelta(days=89)
        elif date_range == 'custom':
            try:
                start_date_str = request.GET.get('start_date')
                end_date_str = request.GET.get('end_date')
                if start_date_str:
                    start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                else:
                    start_date = end_date - timedelta(days=6)
                    
                if end_date_str:
                    end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            except (ValueError, TypeError):
                start_date = end_date - timedelta(days=6)
        else:  # Default to 7 days
            start_date = end_date - timedelta(days=6)
        
        # Get sales in date range
        sales = Sale.objects.filter(
            date__date__range=[start_date, end_date],
            is_completed=True
        )
        
        # Apply additional filters
        category_id = request.GET.get('category')
        if category_id:
            sales = sales.filter(items__product__category_id=category_id)
        
        sale_type = request.GET.get('sale_type')
        if sale_type:
            sales = sales.filter(sale_type=sale_type)
        
        # Get all sale items in the date range
        sale_items = SaleItem.objects.filter(
            sale__in=sales
        ).select_related('product', 'product__category')
        
        # Group by product and calculate metrics
        product_performance = {}
        
        for item in sale_items:
            # Skip items without product
            if not item.product:
                continue
            
            product = item.product
            product_id = product.id if product else None
            
            if not product_id:
                continue
            
            if product_id not in product_performance:
                # Safely get category name
                category_name = 'Uncategorized'
                if product.category and hasattr(product.category, 'name'):
                    category_name = product.category.name
                
                product_performance[product_id] = {
                    'product': product,
                    'category': category_name,
                    'total_quantity': Decimal('0.0000'),
                    'total_revenue': Decimal('0.0000'),
                    'total_cost': Decimal('0.0000'),
                    'total_profit': Decimal('0.0000'),
                    'sale_count': 0,
                    'days_sold': set(),
                }
            
            data = product_performance[product_id]
            
            # Safely add quantities and revenue
            try:
                item_quantity = Decimal(str(item.quantity)) if item.quantity else Decimal('0.0000')
                item_total = Decimal(str(item.total)) if item.total else Decimal('0.0000')
                
                data['total_quantity'] += item_quantity
                data['total_revenue'] += item_total
                
                # Calculate cost and profit
                purchase_price = Decimal(str(product.purchase_price)) if product.purchase_price else Decimal('0.0000')
                cost = purchase_price * item_quantity
                data['total_cost'] += cost
                data['total_profit'] += (item_total - cost)
                
                data['sale_count'] += 1
                
                if item.sale and item.sale.date:
                    data['days_sold'].add(item.sale.date.date())
                    
            except (TypeError, ValueError, AttributeError) as e:
                print(f"Error processing item {item.id}: {e}")
                continue
        
        # Calculate additional metrics
        report_data = []
        total_quantity = Decimal('0.0000')
        total_revenue = Decimal('0.0000')
        total_profit = Decimal('0.0000')
        total_products = len(product_performance)
        
        for product_id, data in product_performance.items():
            product = data.get('product')
            if not product:
                continue
            
            # Calculate profit margin safely
            total_revenue_val = data.get('total_revenue', Decimal('0.0000'))
            total_profit_val = data.get('total_profit', Decimal('0.0000'))
            
            if total_revenue_val and total_revenue_val > 0:
                try:
                    profit_margin = (total_profit_val / total_revenue_val) * 100
                except (ZeroDivisionError, TypeError, ValueError):
                    profit_margin = Decimal('0.0000')
            else:
                profit_margin = Decimal('0.0000')
            
            # Calculate average price safely
            total_quantity_val = data.get('total_quantity', Decimal('0.0000'))
            if total_quantity_val and total_quantity_val > 0:
                try:
                    avg_price = total_revenue_val / total_quantity_val
                except (ZeroDivisionError, TypeError, ValueError):
                    avg_price = Decimal('0.0000')
            else:
                avg_price = Decimal('0.0000')
            
            # Days active
            days_active = len(data.get('days_sold', set()))
            
            # Daily average quantity
            days_in_period = max((end_date - start_date).days + 1, 1)
            if days_in_period > 0:
                try:
                    daily_avg_qty = total_quantity_val / days_in_period
                except (TypeError, ValueError, ZeroDivisionError):
                    daily_avg_qty = Decimal('0.0000')
            else:
                daily_avg_qty = Decimal('0.0000')
            
            # Stock turnover rate
            current_stock = Decimal(str(product.quantity)) if product.quantity else Decimal('0.0000')
            if current_stock and current_stock > 0:
                try:
                    turnover_rate = (total_quantity_val / current_stock) * 100
                except (ZeroDivisionError, TypeError, ValueError):
                    turnover_rate = Decimal('0.0000')
            else:
                turnover_rate = Decimal('0.0000')
            
            # Get reorder level
            reorder_level = Decimal(str(product.reorder_level)) if product.reorder_level else Decimal('0.0000')
            
            # Get unit cost
            unit_cost = Decimal(str(product.purchase_price)) if product.purchase_price else Decimal('0.0000')
            
            # Determine stock status
            stock_status = 'Low' if current_stock <= reorder_level else 'Good'
            
            report_data.append({
                'product': product,
                'category': data.get('category', 'Uncategorized'),
                'total_quantity': total_quantity_val,
                'total_revenue': total_revenue_val,
                'total_profit': total_profit_val,
                'total_cost': data.get('total_cost', Decimal('0.0000')),
                'profit_margin': profit_margin,
                'sale_count': data.get('sale_count', 0),
                'avg_price': avg_price,
                'days_active': days_active,
                'daily_avg_qty': daily_avg_qty,
                'turnover_rate': turnover_rate,
                'current_stock': current_stock,
                'stock_status': stock_status,
                'unit_cost': unit_cost,
            })
            
            # Update totals
            total_quantity += total_quantity_val
            total_revenue += total_revenue_val
            total_profit += total_profit_val
        
        # Apply sorting with safe defaults
        sort_by = request.GET.get('sort', 'revenue_desc')
        
        def get_sort_key(item, key, default=Decimal('0.0000')):
            """Safely get sort key with default value"""
            try:
                value = item.get(key, default)
                if isinstance(value, Decimal):
                    return value
                return Decimal(str(value)) if value else Decimal('0.0000')
            except (TypeError, ValueError, AttributeError):
                return Decimal('0.0000')
        
        if sort_by == 'quantity_desc':
            report_data.sort(key=lambda x: get_sort_key(x, 'total_quantity'), reverse=True)
        elif sort_by == 'quantity_asc':
            report_data.sort(key=lambda x: get_sort_key(x, 'total_quantity'))
        elif sort_by == 'revenue_desc':
            report_data.sort(key=lambda x: get_sort_key(x, 'total_revenue'), reverse=True)
        elif sort_by == 'revenue_asc':
            report_data.sort(key=lambda x: get_sort_key(x, 'total_revenue'))
        elif sort_by == 'profit_desc':
            report_data.sort(key=lambda x: get_sort_key(x, 'total_profit'), reverse=True)
        elif sort_by == 'profit_asc':
            report_data.sort(key=lambda x: get_sort_key(x, 'total_profit'))
        elif sort_by == 'margin_desc':
            report_data.sort(key=lambda x: get_sort_key(x, 'profit_margin'), reverse=True)
        elif sort_by == 'margin_asc':
            report_data.sort(key=lambda x: get_sort_key(x, 'profit_margin'))
        else:
            report_data.sort(key=lambda x: get_sort_key(x, 'total_revenue'), reverse=True)
        
        # Calculate summary statistics safely
        if total_products > 0:
            try:
                avg_revenue_per_product = total_revenue / total_products
                avg_profit_per_product = total_profit / total_products
            except (TypeError, ValueError, ZeroDivisionError):
                avg_revenue_per_product = Decimal('0.0000')
                avg_profit_per_product = Decimal('0.0000')
        else:
            avg_revenue_per_product = Decimal('0.0000')
            avg_profit_per_product = Decimal('0.0000')
        
        if total_revenue and total_revenue > 0:
            try:
                avg_margin = (total_profit / total_revenue) * 100
            except (ZeroDivisionError, TypeError, ValueError):
                avg_margin = Decimal('0.0000')
        else:
            avg_margin = Decimal('0.0000')
        
        # Get top performers
        try:
            top_5_by_quantity = sorted(
                [item for item in report_data if get_sort_key(item, 'total_quantity') > 0], 
                key=lambda x: get_sort_key(x, 'total_quantity'), 
                reverse=True
            )[:5]
        except:
            top_5_by_quantity = []
        
        try:
            top_5_by_revenue = sorted(
                [item for item in report_data if get_sort_key(item, 'total_revenue') > 0],
                key=lambda x: get_sort_key(x, 'total_revenue'),
                reverse=True
            )[:5]
        except:
            top_5_by_revenue = []
        
        try:
            top_5_by_profit = sorted(
                [item for item in report_data if get_sort_key(item, 'total_profit') > 0],
                key=lambda x: get_sort_key(x, 'total_profit'),
                reverse=True
            )[:5]
        except:
            top_5_by_profit = []
        
        # Filter for positive margins only
        top_5_by_margin = []
        try:
            for item in report_data:
                margin = get_sort_key(item, 'profit_margin')
                if margin and margin > 0:
                    top_5_by_margin.append(item)
            top_5_by_margin.sort(key=lambda x: get_sort_key(x, 'profit_margin'), reverse=True)
            top_5_by_margin = top_5_by_margin[:5]
        except:
            top_5_by_margin = []
        
        # Low performers (products with revenue but low margin)
        low_performers = []
        try:
            for item in report_data:
                margin = get_sort_key(item, 'profit_margin')
                revenue = get_sort_key(item, 'total_revenue')
                if margin < 10 and revenue > 0:
                    low_performers.append(item)
            low_performers.sort(key=lambda x: get_sort_key(x, 'profit_margin'))
        except:
            low_performers = []
        
        # Get categories for filter dropdown
        try:
            categories = Category.objects.all()
        except:
            categories = []
        
        # Export functionality
        export_format = request.GET.get('export')
        if export_format:
            return export_product_performance(export_format, report_data, start_date, end_date)
        
        # Chart data
        chart_data_quantity = []
        chart_data_revenue = []
        chart_data_profit = []
        
        try:
            # Get valid items for charts
            valid_items = []
            for item in report_data[:10]:  # Top 10 products for chart
                product = item.get('product')
                if product and hasattr(product, 'name'):
                    valid_items.append(item)
            
            for i, item in enumerate(valid_items):
                product = item.get('product')
                if not product:
                    continue
                    
                product_name = getattr(product, 'name', 'Unknown')
                
                chart_data_quantity.append({
                    'product': product_name[:20] + ('...' if len(product_name) > 20 else ''),
                    'quantity': float(get_sort_key(item, 'total_quantity', Decimal('0.0000')))
                })
                
                chart_data_revenue.append({
                    'product': product_name[:20] + ('...' if len(product_name) > 20 else ''),
                    'revenue': float(get_sort_key(item, 'total_revenue', Decimal('0.0000')))
                })
                
                chart_data_profit.append({
                    'product': product_name[:20] + ('...' if len(product_name) > 20 else ''),
                    'profit': float(get_sort_key(item, 'total_profit', Decimal('0.0000')))
                })
        except Exception as e:
            print(f"Error preparing chart data: {e}")
        
        context = {
            'report_data': report_data,
            'total_quantity': total_quantity,
            'total_revenue': total_revenue,
            'total_profit': total_profit,
            'total_products': total_products,
            'avg_revenue_per_product': avg_revenue_per_product,
            'avg_profit_per_product': avg_profit_per_product,
            'avg_margin': avg_margin,
            'top_5_by_quantity': top_5_by_quantity,
            'top_5_by_revenue': top_5_by_revenue,
            'top_5_by_profit': top_5_by_profit,
            'top_5_by_margin': top_5_by_margin,
            'low_performers': low_performers[:10],
            'categories': categories,
            'start_date': start_date.strftime('%Y-%m-%d') if start_date else '',
            'end_date': end_date.strftime('%Y-%m-%d') if end_date else '',
            'selected_range': date_range,
            'selected_category': category_id,
            'selected_sale_type': sale_type,
            'selected_sort': sort_by,
            'sale_types': [('retail', 'Retail'), ('wholesale', 'Wholesale')],
            'sort_options': [
                ('revenue_desc', 'Revenue (High to Low)'),
                ('revenue_asc', 'Revenue (Low to High)'),
                ('quantity_desc', 'Quantity (High to Low)'),
                ('quantity_asc', 'Quantity (Low to High)'),
                ('profit_desc', 'Profit (High to Low)'),
                ('profit_asc', 'Profit (Low to High)'),
                ('margin_desc', 'Profit Margin (High to Low)'),
                ('margin_asc', 'Profit Margin (Low to High)'),
            ],
            'date_ranges': [
                ('1d', 'Today'),
                ('7d', 'Last 7 Days'),
                ('30d', 'Last 30 Days'),
                ('90d', 'Last 90 Days'),
                ('custom', 'Custom Range'),
            ],
            'chart_data_quantity': json.dumps(chart_data_quantity),
            'chart_data_revenue': json.dumps(chart_data_revenue),
            'chart_data_profit': json.dumps(chart_data_profit),
            'today': timezone.now().date(),
        }
        
        return render(request, 'pos/reports/product_performance.html', context)
    
    except Exception as e:
        print(f"Error in product_performance_report: {e}")
        import traceback
        traceback.print_exc()
        
        # Return empty context on error
        context = {
            'report_data': [],
            'total_quantity': Decimal('0.0000'),
            'total_revenue': Decimal('0.0000'),
            'total_profit': Decimal('0.0000'),
            'total_products': 0,
            'avg_revenue_per_product': Decimal('0.0000'),
            'avg_profit_per_product': Decimal('0.0000'),
            'avg_margin': Decimal('0.0000'),
            'top_5_by_quantity': [],
            'top_5_by_revenue': [],
            'top_5_by_profit': [],
            'top_5_by_margin': [],
            'low_performers': [],
            'categories': Category.objects.all(),
            'start_date': timezone.now().date().strftime('%Y-%m-%d'),
            'end_date': timezone.now().date().strftime('%Y-%m-%d'),
            'selected_range': '7d',
            'selected_category': None,
            'selected_sale_type': None,
            'selected_sort': 'revenue_desc',
            'sale_types': [('retail', 'Retail'), ('wholesale', 'Wholesale')],
            'sort_options': [
                ('revenue_desc', 'Revenue (High to Low)'),
                ('revenue_asc', 'Revenue (Low to High)'),
                ('quantity_desc', 'Quantity (High to Low)'),
                ('quantity_asc', 'Quantity (Low to High)'),
                ('profit_desc', 'Profit (High to Low)'),
                ('profit_asc', 'Profit (Low to High)'),
                ('margin_desc', 'Profit Margin (High to Low)'),
                ('margin_asc', 'Profit Margin (Low to High)'),
            ],
            'date_ranges': [
                ('1d', 'Today'),
                ('7d', 'Last 7 Days'),
                ('30d', 'Last 30 Days'),
                ('90d', 'Last 90 Days'),
                ('custom', 'Custom Range'),
            ],
            'chart_data_quantity': json.dumps([]),
            'chart_data_revenue': json.dumps([]),
            'chart_data_profit': json.dumps([]),
            'today': timezone.now().date(),
        }
        
        return render(request, 'pos/reports/product_performance.html', context)


def export_product_performance(export_format, report_data, start_date, end_date):
    """Export product performance report to CSV or Excel"""
    
    if export_format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="product_performance_{start_date}_to_{end_date}.csv"'
        
        writer = csv.writer(response)
        
        # Write header
        writer.writerow([
            'Product Name', 'Barcode', 'Category', 'Current Stock', 
            'Quantity Sold', 'Total Revenue (KES)', 'Total Cost (KES)', 
            'Total Profit (KES)', 'Profit Margin %', 'Avg Selling Price',
            'Number of Sales', 'Days Active', 'Daily Avg Quantity',
            'Stock Status', 'Turnover Rate %'
        ])
        
        # Write data rows with error handling
        for item in report_data:
            try:
                product = item.get('product')
                if not product:
                    continue
                    
                # Safely get values with defaults
                product_name = getattr(product, 'name', 'Unknown')
                barcode = getattr(product, 'barcode', '')
                category = item.get('category', 'Uncategorized')
                current_stock = float(item.get('current_stock', 0))
                total_quantity = float(item.get('total_quantity', 0))
                total_revenue = float(item.get('total_revenue', 0))
                
                # Calculate total cost if not present
                if 'total_cost' in item:
                    total_cost = float(item.get('total_cost', 0))
                else:
                    unit_cost = float(item.get('unit_cost', 0))
                    total_cost = unit_cost * total_quantity
                
                total_profit = float(item.get('total_profit', 0))
                profit_margin = float(item.get('profit_margin', 0))
                avg_price = float(item.get('avg_price', 0))
                sale_count = item.get('sale_count', 0)
                days_active = item.get('days_active', 0)
                daily_avg_qty = float(item.get('daily_avg_qty', 0))
                stock_status = item.get('stock_status', 'Unknown')
                turnover_rate = float(item.get('turnover_rate', 0))
                
                writer.writerow([
                    product_name,
                    barcode,
                    category,
                    current_stock,
                    total_quantity,
                    total_revenue,
                    total_cost,
                    total_profit,
                    profit_margin,
                    avg_price,
                    sale_count,
                    days_active,
                    daily_avg_qty,
                    stock_status,
                    turnover_rate
                ])
            except Exception as e:
                # Skip problematic rows but continue export
                print(f"Error exporting row: {e}")
                continue
        
        return response
    
    elif export_format == 'excel':
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="product_performance_{start_date}_to_{end_date}.xlsx"'
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Product Performance"
        
        # Add headers
        headers = [
            'Product Name', 'Barcode', 'Category', 'Current Stock', 
            'Quantity Sold', 'Total Revenue (KES)', 'Total Cost (KES)', 
            'Total Profit (KES)', 'Profit Margin %', 'Avg Selling Price',
            'Number of Sales', 'Days Active', 'Daily Avg Quantity',
            'Stock Status', 'Turnover Rate %'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
        
        # Add data with error handling
        row_num = 2
        for item in report_data:
            try:
                product = item.get('product')
                if not product:
                    continue
                    
                # Safely get values
                product_name = getattr(product, 'name', 'Unknown')
                barcode = getattr(product, 'barcode', '')
                category = item.get('category', 'Uncategorized')
                current_stock = float(item.get('current_stock', 0))
                total_quantity = float(item.get('total_quantity', 0))
                total_revenue = float(item.get('total_revenue', 0))
                
                # Calculate total cost if not present
                if 'total_cost' in item:
                    total_cost = float(item.get('total_cost', 0))
                else:
                    unit_cost = float(item.get('unit_cost', 0))
                    total_cost = unit_cost * total_quantity
                
                total_profit = float(item.get('total_profit', 0))
                profit_margin = float(item.get('profit_margin', 0))
                avg_price = float(item.get('avg_price', 0))
                sale_count = item.get('sale_count', 0)
                days_active = item.get('days_active', 0)
                daily_avg_qty = float(item.get('daily_avg_qty', 0))
                stock_status = item.get('stock_status', 'Unknown')
                turnover_rate = float(item.get('turnover_rate', 0))
                
                ws.cell(row=row_num, column=1, value=product_name)
                ws.cell(row=row_num, column=2, value=barcode)
                ws.cell(row=row_num, column=3, value=category)
                ws.cell(row=row_num, column=4, value=current_stock)
                ws.cell(row=row_num, column=5, value=total_quantity)
                ws.cell(row=row_num, column=6, value=total_revenue)
                ws.cell(row=row_num, column=7, value=total_cost)
                ws.cell(row=row_num, column=8, value=total_profit)
                ws.cell(row=row_num, column=9, value=profit_margin)
                ws.cell(row=row_num, column=10, value=avg_price)
                ws.cell(row=row_num, column=11, value=sale_count)
                ws.cell(row=row_num, column=12, value=days_active)
                ws.cell(row=row_num, column=13, value=daily_avg_qty)
                ws.cell(row=row_num, column=14, value=stock_status)
                ws.cell(row=row_num, column=15, value=turnover_rate)
                
                row_num += 1
            except Exception as e:
                print(f"Error exporting row to Excel: {e}")
                continue
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add summary sheet if there's data
        if row_num > 2:
            ws2 = wb.create_sheet(title="Summary")
            
            # Calculate summary statistics
            total_quantity = sum(item.get('total_quantity', 0) for item in report_data)
            total_revenue = sum(item.get('total_revenue', 0) for item in report_data)
            total_profit = sum(item.get('total_profit', 0) for item in report_data)
            total_products = len(report_data)
            
            try:
                avg_margin = (total_profit / total_revenue * 100) if total_revenue > 0 else 0
            except (ZeroDivisionError, TypeError):
                avg_margin = 0
            
            # Write summary
            ws2.cell(row=1, column=1, value="SUMMARY STATISTICS").font = Font(bold=True, size=14)
            
            summary_data = [
                ("Report Period", f"{start_date} to {end_date}"),
                ("Total Products Sold", total_products),
                ("Total Quantity Sold", float(total_quantity)),
                ("Total Revenue (KES)", float(total_revenue)),
                ("Total Profit (KES)", float(total_profit)),
                ("Average Profit Margin", f"{avg_margin:.2f}%"),
                ("Generated On", timezone.now().strftime('%Y-%m-%d %H:%M:%S')),
            ]
            
            for i, (label, value) in enumerate(summary_data, 3):
                ws2.cell(row=i, column=1, value=label).font = Font(bold=True)
                ws2.cell(row=i, column=2, value=value)
        
        wb.save(response)
        return response
    
    return HttpResponse('Invalid export format', status=400)



# views.py - Add this new function
from datetime import datetime, timedelta
from decimal import Decimal
from django.db.models import Sum, Count, Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.utils import timezone
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import json

@login_required
def weekly_sales_profit_report(request):
    """
    Weekly sales and profit report with working exports
    """
    try:
        # 1. DATE FILTERING
        today = timezone.now().date()
        
        # Get date parameters or use defaults
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
        
        if start_date_str and end_date_str:
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            except ValueError:
                # Default to last 7 days if invalid dates
                end_date = today
                start_date = end_date - timedelta(days=6)
        else:
            # Default to last 7 days
            end_date = today
            start_date = end_date - timedelta(days=6)
        
        # 2. OTHER FILTERS
        category_id = request.GET.get('category')
        sale_type = request.GET.get('sale_type')
        sort_by = request.GET.get('sort', 'quantity_desc')
        
        # 3. CHECK FOR EXPORT FIRST
        export_format = request.GET.get('export')
        if export_format in ['csv', 'excel']:
            # We'll generate data and export
            report_data = generate_report_data(start_date, end_date, category_id, sale_type)
            if export_format == 'csv':
                return export_to_csv(report_data, start_date, end_date)
            elif export_format == 'excel':
                return export_to_excel(report_data, start_date, end_date)
        
        # 4. GENERATE REPORT DATA FOR DISPLAY
        report_data = generate_report_data(start_date, end_date, category_id, sale_type)
        
        # 5. SORT DATA
        if sort_by == 'quantity_desc':
            report_data.sort(key=lambda x: x['total_quantity'], reverse=True)
        elif sort_by == 'quantity_asc':
            report_data.sort(key=lambda x: x['total_quantity'])
        elif sort_by == 'revenue_desc':
            report_data.sort(key=lambda x: x['total_revenue'], reverse=True)
        elif sort_by == 'revenue_asc':
            report_data.sort(key=lambda x: x['total_revenue'])
        elif sort_by == 'profit_desc':
            report_data.sort(key=lambda x: x['total_profit'], reverse=True)
        elif sort_by == 'profit_asc':
            report_data.sort(key=lambda x: x['total_profit'])
        elif sort_by == 'margin_desc':
            report_data.sort(key=lambda x: x['profit_margin'], reverse=True)
        elif sort_by == 'margin_asc':
            report_data.sort(key=lambda x: x['profit_margin'])
        
        # 6. CALCULATE TOTALS
        total_quantity = sum(item['total_quantity'] for item in report_data)
        total_revenue = sum(item['total_revenue'] for item in report_data)
        total_cost = sum(item['total_cost'] for item in report_data)
        total_profit = sum(item['total_profit'] for item in report_data)
        total_products = len(report_data)
        
        overall_margin = (total_profit / total_revenue * 100) if total_revenue > 0 else 0
        
        # 7. TOP PERFORMERS
        top_10_by_quantity = sorted(report_data, key=lambda x: x['total_quantity'], reverse=True)[:10]
        top_10_by_profit = sorted(report_data, key=lambda x: x['total_profit'], reverse=True)[:10]
        
        # 8. PREPARE CONTEXT
        categories = Category.objects.all().order_by('name')
        
        context = {
            'report_data': report_data,
            'categories': categories,
            'start_date': start_date.strftime('%Y-%m-%d'),
            'end_date': end_date.strftime('%Y-%m-%d'),
            'selected_category': category_id,
            'selected_sale_type': sale_type,
            'selected_sort': sort_by,
            'total_quantity': total_quantity,
            'total_revenue': total_revenue,
            'total_cost': total_cost,
            'total_profit': total_profit,
            'total_products': total_products,
            'overall_margin': overall_margin,
            'top_10_by_quantity': top_10_by_quantity,
            'top_10_by_profit': top_10_by_profit,
            'sale_types': [('', 'All'), ('retail', 'Retail'), ('wholesale', 'Wholesale')],
            'sort_options': [
                ('quantity_desc', 'Quantity (High to Low)'),
                ('quantity_asc', 'Quantity (Low to High)'),
                ('revenue_desc', 'Revenue (High to Low)'),
                ('revenue_asc', 'Revenue (Low to High)'),
                ('profit_desc', 'Profit (High to Low)'),
                ('profit_asc', 'Profit (Low to High)'),
                ('margin_desc', 'Profit Margin (High to Low)'),
                ('margin_asc', 'Profit Margin (Low to High)'),
            ],
            'today': today,
        }
        
        return render(request, 'pos/reports/weekly_sales_profit.html', context)
        
    except Exception as e:
        print(f"Error in weekly_sales_profit_report: {str(e)}")
        import traceback
        traceback.print_exc()
        
        # Return empty context on error
        return render(request, 'pos/reports/weekly_sales_profit.html', {
            'report_data': [],
            'categories': Category.objects.all(),
            'start_date': timezone.now().date().strftime('%Y-%m-%d'),
            'end_date': timezone.now().date().strftime('%Y-%m-%d'),
            'error': str(e),
        })

def generate_report_data(start_date, end_date, category_id=None, sale_type=None):
    """Generate report data from database"""
    try:
        # Get sales in date range
        sales = Sale.objects.filter(
            date__date__range=[start_date, end_date],
            is_completed=True
        )
        
        # Apply filters
        if category_id:
            sales = sales.filter(items__product__category_id=category_id)
        
        if sale_type:
            sales = sales.filter(sale_type=sale_type)
        
        # Get all sale items
        sale_items = SaleItem.objects.filter(
            sale__in=sales
        ).select_related('product', 'product__category')
        
        # Group by product
        product_data = {}
        
        for item in sale_items:
            if not item.product:
                continue
                
            product = item.product
            product_id = product.id
            
            if product_id not in product_data:
                product_data[product_id] = {
                    'product_id': product_id,
                    'product_name': product.name,
                    'product_code': product.barcode or '',
                    'category': product.category.name if product.category else 'Uncategorized',
                    'total_quantity': Decimal('0'),
                    'total_revenue': Decimal('0'),
                    'total_cost': Decimal('0'),
                    'total_profit': Decimal('0'),
                    'sale_count': 0,
                    'unit_cost': product.purchase_price or Decimal('0'),
                    'selling_price': product.selling_price or Decimal('0'),
                    'current_stock': product.quantity or Decimal('0'),
                    'reorder_level': product.reorder_level or Decimal('0'),
                }
            
            data = product_data[product_id]
            
            # Add item data
            quantity = item.quantity or Decimal('0')
            revenue = item.total or Decimal('0')
            
            data['total_quantity'] += quantity
            data['total_revenue'] += revenue
            
            # Calculate cost and profit
            unit_cost = product.purchase_price or Decimal('0')
            cost = unit_cost * quantity
            profit = revenue - cost
            
            data['total_cost'] += cost
            data['total_profit'] += profit
            data['sale_count'] += 1
        
        # Convert to list and calculate additional metrics
        report_list = []
        for product_id, data in product_data.items():
            # Calculate profit margin
            if data['total_revenue'] > 0:
                profit_margin = (data['total_profit'] / data['total_revenue']) * 100
            else:
                profit_margin = Decimal('0')
            
            # Calculate average selling price
            if data['total_quantity'] > 0:
                avg_selling_price = data['total_revenue'] / data['total_quantity']
            else:
                avg_selling_price = Decimal('0')
            
            # Determine stock status
            stock_status = 'Low' if data['current_stock'] <= data['reorder_level'] else 'Good'
            
            report_list.append({
                **data,
                'profit_margin': profit_margin,
                'avg_selling_price': avg_selling_price,
                'stock_status': stock_status,
            })
        
        return report_list
        
    except Exception as e:
        print(f"Error generating report data: {str(e)}")
        return []

def export_to_csv(report_data, start_date, end_date):
    """Export report data to CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="sales_report_{start_date}_to_{end_date}.csv"'
    
    writer = csv.writer(response)
    
    # Write header
    writer.writerow([
        'Product Name', 'Product Code', 'Category',
        'Quantity Sold', 'Total Revenue (KES)', 'Total Cost (KES)',
        'Total Profit (KES)', 'Profit Margin %', 'Unit Cost (KES)',
        'Selling Price (KES)', 'Average Price (KES)', 'Sales Count',
        'Current Stock', 'Reorder Level', 'Stock Status'
    ])
    
    # Write data rows
    for item in report_data:
        writer.writerow([
            item['product_name'],
            item['product_code'],
            item['category'],
            float(item['total_quantity']),
            float(item['total_revenue']),
            float(item['total_cost']),
            float(item['total_profit']),
            float(item['profit_margin']),
            float(item['unit_cost']),
            float(item['selling_price']),
            float(item.get('avg_selling_price', 0)),
            item['sale_count'],
            float(item['current_stock']),
            float(item['reorder_level']),
            item['stock_status'],
        ])
    
    # Add summary
    writer.writerow([])
    writer.writerow(['SUMMARY'])
    writer.writerow(['Report Period', f'{start_date} to {end_date}'])
    writer.writerow(['Total Products', len(report_data)])
    writer.writerow(['Total Quantity Sold', float(sum(item['total_quantity'] for item in report_data))])
    writer.writerow(['Total Revenue (KES)', float(sum(item['total_revenue'] for item in report_data))])
    writer.writerow(['Total Profit (KES)', float(sum(item['total_profit'] for item in report_data))])
    writer.writerow(['Generated On', timezone.now().strftime('%Y-%m-%d %H:%M:%S')])
    
    return response

def export_to_excel(report_data, start_date, end_date):
    """Export report data to Excel"""
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="sales_report_{start_date}_to_{end_date}.xlsx"'
    
    # Create workbook
    wb = Workbook()
    
    # Sheet 1: Main Data
    ws1 = wb.active
    ws1.title = "Sales Report"
    
    # Title
    ws1.merge_cells('A1:O1')
    title_cell = ws1['A1']
    title_cell.value = f"WEEKLY SALES & PROFIT REPORT: {start_date} to {end_date}"
    title_cell.font = Font(bold=True, size=16, color='FFFFFF')
    title_cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    title_cell.alignment = Alignment(horizontal='center')
    
    # Header
    headers = [
        'Product Name', 'Product Code', 'Category',
        'Quantity Sold', 'Total Revenue (KES)', 'Total Cost (KES)',
        'Total Profit (KES)', 'Profit Margin %', 'Unit Cost (KES)',
        'Selling Price (KES)', 'Average Price (KES)', 'Sales Count',
        'Current Stock', 'Reorder Level', 'Stock Status'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws1.cell(row=3, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    row_num = 4
    for item in report_data:
        ws1.cell(row=row_num, column=1, value=item['product_name'])
        ws1.cell(row=row_num, column=2, value=item['product_code'])
        ws1.cell(row=row_num, column=3, value=item['category'])
        ws1.cell(row=row_num, column=4, value=float(item['total_quantity']))
        ws1.cell(row=row_num, column=5, value=float(item['total_revenue']))
        ws1.cell(row=row_num, column=6, value=float(item['total_cost']))
        
        # Profit with conditional formatting
        profit_cell = ws1.cell(row=row_num, column=7, value=float(item['total_profit']))
        if item['total_profit'] >= 0:
            profit_cell.font = Font(color='006100', bold=True)
        else:
            profit_cell.font = Font(color='FF0000', bold=True)
        
        # Margin with conditional formatting
        margin_cell = ws1.cell(row=row_num, column=8, value=float(item['profit_margin']))
        if item['profit_margin'] >= 20:
            margin_cell.font = Font(color='006100', bold=True)
        elif item['profit_margin'] >= 10:
            margin_cell.font = Font(color='FFA500', bold=True)
        else:
            margin_cell.font = Font(color='FF0000', bold=True)
        
        ws1.cell(row=row_num, column=9, value=float(item['unit_cost']))
        ws1.cell(row=row_num, column=10, value=float(item['selling_price']))
        ws1.cell(row=row_num, column=11, value=float(item.get('avg_selling_price', 0)))
        ws1.cell(row=row_num, column=12, value=item['sale_count'])
        ws1.cell(row=row_num, column=13, value=float(item['current_stock']))
        ws1.cell(row=row_num, column=14, value=float(item['reorder_level']))
        
        # Stock status
        status_cell = ws1.cell(row=row_num, column=15, value=item['stock_status'])
        if item['stock_status'] == 'Low':
            status_cell.font = Font(color='FF0000', bold=True)
        else:
            status_cell.font = Font(color='006100', bold=True)
        
        row_num += 1
    
    # Auto-adjust column widths
    for column in ws1.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws1.column_dimensions[column_letter].width = adjusted_width
    
    # Sheet 2: Summary
    ws2 = wb.create_sheet(title="Summary")
    
    # Calculate totals
    total_quantity = sum(item['total_quantity'] for item in report_data)
    total_revenue = sum(item['total_revenue'] for item in report_data)
    total_cost = sum(item['total_cost'] for item in report_data)
    total_profit = sum(item['total_profit'] for item in report_data)
    
    # Summary title
    ws2.merge_cells('A1:B1')
    ws2['A1'] = "REPORT SUMMARY"
    ws2['A1'].font = Font(bold=True, size=14)
    
    # Summary data
    summary_data = [
        ("Report Period", f"{start_date} to {end_date}"),
        ("Total Products", len(report_data)),
        ("Total Quantity Sold", float(total_quantity)),
        ("Total Revenue (KES)", float(total_revenue)),
        ("Total Cost (KES)", float(total_cost)),
        ("Total Profit (KES)", float(total_profit)),
        ("Profit Margin", f"{(total_profit/total_revenue*100) if total_revenue > 0 else 0:.2f}%"),
        ("Generated On", timezone.now().strftime('%Y-%m-%d %H:%M:%S')),
    ]
    
    for i, (label, value) in enumerate(summary_data, 3):
        ws2.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws2.cell(row=i, column=2, value=value)
    
    # Top performers
    top_start = len(summary_data) + 5
    ws2.cell(row=top_start, column=1, value="TOP 5 PRODUCTS BY QUANTITY").font = Font(bold=True, size=12)
    
    top_by_qty = sorted(report_data, key=lambda x: x['total_quantity'], reverse=True)[:5]
    for i, item in enumerate(top_by_qty, 1):
        ws2.cell(row=top_start + i, column=1, value=f"{i}. {item['product_name'][:30]}")
        ws2.cell(row=top_start + i, column=2, value=f"Quantity: {float(item['total_quantity']):.0f}")
        ws2.cell(row=top_start + i, column=3, value=f"Profit: KSh {float(item['total_profit']):.2f}")
    
    wb.save(response)
    return response



# Add these views to your views.py

from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.db import transaction
from django.core.paginator import Paginator
from django.views.decorators.http import require_POST, require_GET
from django.utils.crypto import get_random_string
import json
import csv
from openpyxl import Workbook
from datetime import datetime

def is_admin(user):
    """Check if user has admin privileges"""
    return user.is_authenticated and (user.is_superuser or 
           (hasattr(user, 'profile') and 
            user.profile.role and 
            user.profile.role.name == 'admin'))

@login_required
@user_passes_test(is_admin)
def user_management_dashboard(request):
    """Main dashboard for user management"""
    # Get statistics
    total_users = User.objects.count()
    active_users = User.objects.filter(is_active=True).count()
    total_roles = Role.objects.count()
    active_roles = Role.objects.filter(is_active=True).count()
    
    # Recent activities
    recent_activities = UserActivityLog.objects.select_related('user').order_by('-timestamp')[:10]
    
    # User distribution by role
    user_distribution = []
    for role in Role.objects.all():
        count = User.objects.filter(
            profile__role=role,
            is_active=True
        ).count()
        if count > 0:
            user_distribution.append({
                'role': role.get_name_display(),
                'count': count,
                'percentage': (count / active_users * 100) if active_users > 0 else 0
            })
    
    # Recent user logins (last 7 days)
    seven_days_ago = timezone.now() - timedelta(days=7)
    recent_logins = UserActivityLog.objects.filter(
        action_type='login',
        timestamp__gte=seven_days_ago
    ).values('user__username').annotate(
        login_count=Count('id'),
        last_login=Max('timestamp')
    ).order_by('-last_login')[:5]
    
    context = {
        'total_users': total_users,
        'active_users': active_users,
        'total_roles': total_roles,
        'active_roles': active_roles,
        'recent_activities': recent_activities,
        'user_distribution': user_distribution,
        'recent_logins': recent_logins,
    }
    return render(request, 'pos/admin/user_management_dashboard.html', context)


@login_required
@user_passes_test(is_admin)
def user_list(request):
    """List all users with filters"""
    users = User.objects.select_related('profile', 'profile__role').all()
    
    # Apply filters
    form = UserFilterForm(request.GET)
    if form.is_valid():
        role = form.cleaned_data.get('role')
        is_active = form.cleaned_data.get('is_active')
        search = form.cleaned_data.get('search')
        
        if role:
            users = users.filter(profile__role=role)
        
        if is_active == 'true':
            users = users.filter(is_active=True)
        elif is_active == 'false':
            users = users.filter(is_active=False)
        
        if search:
            users = users.filter(
                Q(username__icontains=search) |
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(email__icontains=search) |
                Q(profile__phone__icontains=search)
            )
    
    # Pagination
    paginator = Paginator(users.order_by('date_joined'), 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Export functionality
    export_format = request.GET.get('export')
    if export_format:
        return export_users(export_format, users)
    
    context = {
        'users': page_obj,
        'form': form,
        'total_users': users.count(),
        'roles': Role.objects.filter(is_active=True),
    }
    return render(request, 'pos/admin/user_list.html', context)


@login_required
# pos/views.py - Update create_user view
@login_required
@user_passes_test(is_admin)
def create_user(request):
    """Create a new user"""
    if request.method == 'POST':
        user_form = CustomUserCreationForm(request.POST)
        profile_form = UserProfileForm(request.POST, request.FILES)
        
        if user_form.is_valid() and profile_form.is_valid():
            try:
                with transaction.atomic():
                    # Create user
                    user = user_form.save(commit=False)
                    user.is_staff = False  # Regular user, not Django admin
                    user.save()
                    
                    # Create profile
                    profile = profile_form.save(commit=False)
                    profile.user = user
                    profile.save()
                    
                    # Log activity
                    UserActivityLog.objects.create(
                        user=request.user,
                        action_type='create',
                        model_name='User',
                        object_id=str(user.id),
                        description=f'Created user {user.username} with role {profile.role}',
                        ip_address=request.META.get('REMOTE_ADDR'),
                        user_agent=request.META.get('HTTP_USER_AGENT', '')
                    )
                    
                    messages.success(request, f'User {user.username} created successfully!')
                    
                    # Send welcome email if requested
                    if request.POST.get('send_welcome_email'):
                        send_welcome_email(user, request)
                    
                    return redirect('user_detail', pk=user.id)
                    
            except Exception as e:
                messages.error(request, f'Error creating user: {str(e)}')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        user_form = CustomUserCreationForm()
        profile_form = UserProfileForm()
    
    # Get roles with user counts
    roles_with_counts = []
    roles = Role.objects.filter(is_active=True)
    for role in roles:
        user_count = User.objects.filter(profile__role=role, is_active=True).count()
        roles_with_counts.append({
            'role': role,
            'user_count': user_count
        })
    
    context = {
        'user_form': user_form,
        'profile_form': profile_form,
        'title': 'Create New User',
        'roles': roles_with_counts,
        'permission_categories': get_permission_categories(),  # Add this
    }
    return render(request, 'pos/admin/create_user.html', context)

@login_required
@user_passes_test(is_admin)
def edit_user(request, pk):
    """Edit an existing user"""
    user = get_object_or_404(User, pk=pk)
    profile, created = UserProfile.objects.get_or_create(user=user)
    
    if request.method == 'POST':
        user_form = CustomUserEditForm(request.POST, instance=user)
        profile_form = UserProfileForm(request.POST, request.FILES, instance=profile)
        
        if user_form.is_valid() and profile_form.is_valid():
            try:
                with transaction.atomic():
                    # Update user
                    user = user_form.save()
                    
                    # Update profile
                    profile = profile_form.save()
                    
                    # Log activity
                    UserActivityLog.objects.create(
                        user=request.user,
                        action_type='update',
                        model_name='User',
                        object_id=str(user.id),
                        description=f'Updated user {user.username}',
                        ip_address=request.META.get('REMOTE_ADDR'),
                        user_agent=request.META.get('HTTP_USER_AGENT', '')
                    )
                    
                    messages.success(request, f'User {user.username} updated successfully!')
                    return redirect('user_detail', pk=user.id)
                    
            except Exception as e:
                messages.error(request, f'Error updating user: {str(e)}')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        user_form = CustomUserEditForm(instance=user)
        profile_form = UserProfileForm(instance=profile)
    
    context = {
        'user_form': user_form,
        'profile_form': profile_form,
        'user': user,
        'profile': profile,
        'title': f'Edit User: {user.username}',
    }
    return render(request, 'pos/admin/edit_user.html', context)


@login_required
@user_passes_test(is_admin)
def user_detail(request, pk):
    """View user details"""
    user = get_object_or_404(User.objects.select_related('profile', 'profile__role'), pk=pk)
    profile = user.profile if hasattr(user, 'profile') else None
    
    # Get user permissions
    permissions = profile.get_all_permissions() if profile else {}
    
    # Get user activities
    activities = UserActivityLog.objects.filter(user=user).order_by('-timestamp')[:20]
    
    # Get user statistics
    user_stats = {
        'total_sales': Sale.objects.filter(user=user).count(),
        'total_purchases': Purchase.objects.filter(user=user).count(),
        'last_login': UserActivityLog.objects.filter(
            user=user, 
            action_type='login'
        ).order_by('-timestamp').first(),
        'created_at': user.date_joined,
    }
    
    context = {
        'user': user,
        'profile': profile,
        'permissions': permissions,
        'activities': activities,
        'stats': user_stats,
    }
    return render(request, 'pos/admin/user_detail.html', context)


@login_required
@user_passes_test(is_admin)
@require_POST
def toggle_user_status(request, pk):
    """Activate/Deactivate a user"""
    user = get_object_or_404(User, pk=pk)
    
    try:
        user.is_active = not user.is_active
        user.save()
        
        # Update profile if exists
        if hasattr(user, 'profile'):
            user.profile.is_active = user.is_active
            user.profile.save()
        
        # Log activity
        UserActivityLog.objects.create(
            user=request.user,
            action_type='update',
            model_name='User',
            object_id=str(user.id),
            description=f'{"Activated" if user.is_active else "Deactivated"} user {user.username}',
            ip_address=request.META.get('REMOTE_ADDR'),
            user_agent=request.META.get('HTTP_USER_AGENT', '')
        )
        
        status = "activated" if user.is_active else "deactivated"
        messages.success(request, f'User {user.username} has been {status}.')
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'message': f'User {user.username} {status} successfully',
                'is_active': user.is_active
            })
            
    except Exception as e:
        error_message = f'Error updating user status: {str(e)}'
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'message': error_message
            }, status=500)
        messages.error(request, error_message)
    
    return redirect('user_list')


@login_required
@user_passes_test(is_admin)
def reset_user_password(request, pk):
    """Admin reset user password"""
    user = get_object_or_404(User, pk=pk)
    
    if request.method == 'POST':
        form = PasswordResetAdminForm(request.POST)
        form.fields['user'].queryset = User.objects.filter(pk=pk)
        
        if form.is_valid():
            try:
                new_password = form.cleaned_data['new_password']
                user.set_password(new_password)
                user.save()
                
                # Log activity
                UserActivityLog.objects.create(
                    user=request.user,
                    action_type='update',
                    model_name='User',
                    object_id=str(user.id),
                    description=f'Reset password for user {user.username}',
                    ip_address=request.META.get('REMOTE_ADDR'),
                    user_agent=request.META.get('HTTP_USER_AGENT', '')
                )
                
                messages.success(request, f'Password for {user.username} has been reset successfully!')
                return redirect('user_detail', pk=user.id)
                
            except Exception as e:
                messages.error(request, f'Error resetting password: {str(e)}')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = PasswordResetAdminForm(initial={'user': user})
    
    context = {
        'form': form,
        'user': user,
    }
    return render(request, 'pos/admin/reset_password.html', context)


@login_required
@user_passes_test(is_admin)
def role_list(request):
    """List all roles"""
    roles = Role.objects.all()
    
    # Apply filters
    form = RoleFilterForm(request.GET)
    if form.is_valid():
        is_active = form.cleaned_data.get('is_active')
        search = form.cleaned_data.get('search')
        
        if is_active == 'true':
            roles = roles.filter(is_active=True)
        elif is_active == 'false':
            roles = roles.filter(is_active=False)
        
        if search:
            roles = roles.filter(
                Q(name__icontains=search) |
                Q(description__icontains=search)
            )
    
    # Count users per role
    for role in roles:
        role.user_count = User.objects.filter(
            profile__role=role,
            is_active=True
        ).count()
    
    # Export functionality
    export_format = request.GET.get('export')
    if export_format:
        return export_roles(export_format, roles)
    
    context = {
        'roles': roles,
        'form': form,
        'total_roles': roles.count(),
        'permission_categories': get_permission_categories(),
    }
    return render(request, 'pos/admin/role_list.html', context)


@login_required
@user_passes_test(is_admin)
def create_role(request):
    """Create a new role"""
    if request.method == 'POST':
        form = RoleForm(request.POST)
        
        if form.is_valid():
            try:
                role = form.save()
                
                # Log activity
                UserActivityLog.objects.create(
                    user=request.user,
                    action_type='create',
                    model_name='Role',
                    object_id=str(role.id),
                    description=f'Created role {role.get_name_display()}',
                    ip_address=request.META.get('REMOTE_ADDR'),
                    user_agent=request.META.get('HTTP_USER_AGENT', '')
                )
                
                messages.success(request, f'Role {role.get_name_display()} created successfully!')
                return redirect('role_detail', pk=role.id)
                
            except Exception as e:
                messages.error(request, f'Error creating role: {str(e)}')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = RoleForm()
    
    context = {
        'form': form,
        'title': 'Create New Role',
        'permission_categories': get_permission_categories(),
    }
    return render(request, 'pos/admin/create_role.html', context)


@login_required
@user_passes_test(is_admin)
def edit_role(request, pk):
    """Edit an existing role"""
    role = get_object_or_404(Role, pk=pk)
    
    if request.method == 'POST':
        form = RoleForm(request.POST, instance=role)
        
        if form.is_valid():
            try:
                role = form.save()
                
                # Log activity
                UserActivityLog.objects.create(
                    user=request.user,
                    action_type='update',
                    model_name='Role',
                    object_id=str(role.id),
                    description=f'Updated role {role.get_name_display()}',
                    ip_address=request.META.get('REMOTE_ADDR'),
                    user_agent=request.META.get('HTTP_USER_AGENT', '')
                )
                
                messages.success(request, f'Role {role.get_name_display()} updated successfully!')
                return redirect('role_detail', pk=role.id)
                
            except Exception as e:
                messages.error(request, f'Error updating role: {str(e)}')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = RoleForm(instance=role)
    
    context = {
        'form': form,
        'role': role,
        'title': f'Edit Role: {role.get_name_display()}',
        'permission_categories': get_permission_categories(),
    }
    return render(request, 'pos/admin/edit_role.html', context)


@login_required
@user_passes_test(is_admin)
def role_detail(request, pk):
    """View role details"""
    role = get_object_or_404(Role, pk=pk)
    
    # Get users with this role
    users = User.objects.filter(profile__role=role, is_active=True).select_related('profile')
    
    # Get permissions by category
    permissions_by_category = {}
    for category_name, fields in get_permission_categories().items():
        permissions_by_category[category_name] = []
        for field_name, field_label in fields:
            if hasattr(role, field_name):
                permissions_by_category[category_name].append({
                    'name': field_name,
                    'label': field_label,
                    'value': getattr(role, field_name)
                })
    
    context = {
        'role': role,
        'users': users,
        'permissions_by_category': permissions_by_category,
        'user_count': users.count(),
    }
    return render(request, 'pos/admin/role_detail.html', context)


@login_required
@user_passes_test(is_admin)
@require_POST
def toggle_role_status(request, pk):
    """Activate/Deactivate a role"""
    role = get_object_or_404(Role, pk=pk)
    
    try:
        role.is_active = not role.is_active
        role.save()
        
        # Log activity
        UserActivityLog.objects.create(
            user=request.user,
            action_type='update',
            model_name='Role',
            object_id=str(role.id),
            description=f'{"Activated" if role.is_active else "Deactivated"} role {role.get_name_display()}',
            ip_address=request.META.get('REMOTE_ADDR'),
            user_agent=request.META.get('HTTP_USER_AGENT', '')
        )
        
        status = "activated" if role.is_active else "deactivated"
        messages.success(request, f'Role {role.get_name_display()} has been {status}.')
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'message': f'Role {status} successfully',
                'is_active': role.is_active
            })
            
    except Exception as e:
        error_message = f'Error updating role status: {str(e)}'
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'message': error_message
            }, status=500)
        messages.error(request, error_message)
    
    return redirect('role_list')


@login_required
@user_passes_test(is_admin)
def duplicate_role(request, pk):
    """Duplicate an existing role"""
    original_role = get_object_or_404(Role, pk=pk)
    
    if request.method == 'POST':
        try:
            # Create new role with same permissions
            new_role = Role.objects.create(
                name='copy_of_' + original_role.name,
                description=f"Copy of {original_role.get_name_display()} - {original_role.description}",
                is_active=original_role.is_active,
            )
            
            # Copy all permission fields
            for field in original_role._meta.get_fields():
                if field.name.startswith('can_'):
                    setattr(new_role, field.name, getattr(original_role, field.name))
            
            new_role.save()
            
            # Log activity
            UserActivityLog.objects.create(
                user=request.user,
                action_type='create',
                model_name='Role',
                object_id=str(new_role.id),
                description=f'Duplicated role {original_role.get_name_display()} to {new_role.get_name_display()}',
                ip_address=request.META.get('REMOTE_ADDR'),
                user_agent=request.META.get('HTTP_USER_AGENT', '')
            )
            
            messages.success(request, f'Role duplicated successfully!')
            return redirect('edit_role', pk=new_role.id)
            
        except Exception as e:
            messages.error(request, f'Error duplicating role: {str(e)}')
    
    context = {
        'role': original_role,
    }
    return render(request, 'pos/admin/duplicate_role.html', context)


@login_required
@user_passes_test(is_admin)
def activity_logs(request):
    """View system activity logs"""
    logs = UserActivityLog.objects.select_related('user').all()
    
    # Apply filters
    action_type = request.GET.get('action_type')
    user_id = request.GET.get('user')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if action_type:
        logs = logs.filter(action_type=action_type)
    
    if user_id:
        logs = logs.filter(user_id=user_id)
    
    if start_date:
        logs = logs.filter(timestamp__date__gte=start_date)
    
    if end_date:
        logs = logs.filter(timestamp__date__lte=end_date)
    
    # Pagination
    paginator = Paginator(logs.order_by('-timestamp'), 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Get filter options
    users = User.objects.filter(is_active=True)
    action_types = UserActivityLog.ACTION_TYPES
    
    # Export functionality
    export_format = request.GET.get('export')
    if export_format:
        return export_activity_logs(export_format, logs)
    
    context = {
        'logs': page_obj,
        'users': users,
        'action_types': action_types,
        'selected_action': action_type,
        'selected_user': user_id,
        'start_date': start_date,
        'end_date': end_date,
    }
    return render(request, 'pos/admin/activity_logs.html', context)


@login_required
@user_passes_test(is_admin)
def bulk_import_users(request):
    """Bulk import users from Excel file"""
    if request.method == 'POST':
        form = BulkUserImportForm(request.POST, request.FILES)
        
        if form.is_valid():
            try:
                import pandas as pd
                from io import BytesIO
                
                file = request.FILES['file']
                send_email = form.cleaned_data.get('send_email', False)
                generate_passwords = form.cleaned_data.get('generate_passwords', True)
                
                # Read Excel file
                df = pd.read_excel(BytesIO(file.read()))
                
                imported_count = 0
                errors = []
                
                with transaction.atomic():
                    for index, row in df.iterrows():
                        try:
                            # Extract user data
                            username = str(row.get('username', '')).strip()
                            email = str(row.get('email', '')).strip()
                            first_name = str(row.get('first_name', '')).strip()
                            last_name = str(row.get('last_name', '')).strip()
                            role_name = str(row.get('role', '')).strip()
                            phone = str(row.get('phone', '')).strip()
                            
                            # Validate required fields
                            if not username or not email:
                                errors.append(f"Row {index+1}: Missing username or email")
                                continue
                            
                            # Check if user already exists
                            if User.objects.filter(username=username).exists():
                                errors.append(f"Row {index+1}: Username {username} already exists")
                                continue
                            
                            if User.objects.filter(email=email).exists():
                                errors.append(f"Row {index+1}: Email {email} already exists")
                                continue
                            
                            # Generate password
                            if generate_passwords:
                                password = get_random_string(12)
                            else:
                                password = 'TempPass123!'  # Default temporary password
                            
                            # Create user
                            user = User.objects.create_user(
                                username=username,
                                email=email,
                                password=password,
                                first_name=first_name,
                                last_name=last_name,
                                is_active=True
                            )
                            
                            # Get or create role
                            role = None
                            if role_name:
                                try:
                                    role = Role.objects.get(name=role_name)
                                except Role.DoesNotExist:
                                    # Create custom role if not found
                                    role = Role.objects.create(
                                        name=role_name.lower(),
                                        description=f"Auto-created role for {role_name}",
                                        is_active=True
                                    )
                            
                            # Create profile
                            UserProfile.objects.create(
                                user=user,
                                role=role,
                                phone=phone,
                                is_active=True
                            )
                            
                            imported_count += 1
                            
                            # Send welcome email if requested
                            if send_email:
                                send_welcome_email(user, request, password)
                            
                        except Exception as e:
                            errors.append(f"Row {index+1}: {str(e)}")
                
                if errors:
                    messages.warning(request, f"Import completed with {len(errors)} errors. {imported_count} users imported successfully.")
                    request.session['import_errors'] = errors[:50]
                else:
                    messages.success(request, f'Successfully imported {imported_count} users!')
                
                return redirect('user_list')
                
            except Exception as e:
                messages.error(request, f'Error during import: {str(e)}')
        else:
            messages.error(request, 'Invalid file format')
    else:
        form = BulkUserImportForm()
    
    # Get import errors from session
    errors = request.session.pop('import_errors', [])
    
    context = {
        'form': form,
        'errors': errors,
        'roles': Role.objects.filter(is_active=True),
    }
    return render(request, 'pos/admin/bulk_import_users.html', context)


@login_required
@user_passes_test(is_admin)
# pos/views.py - Update user_permissions_summary view
# pos/views.py - Update user_permissions_summary view
@login_required
@user_passes_test(is_admin)
def user_permissions_summary(request):
    """Display summary of user permissions"""
    # Get all roles with their permissions
    roles = Role.objects.filter(is_active=True)
    
    # Get permission matrix
    permission_matrix = []
    permission_fields = [f.name for f in Role._meta.get_fields() if f.name.startswith('can_')]
    
    # Calculate permission counts per role
    for role in roles:
        enabled_count = 0
        for field in permission_fields:
            if getattr(role, field, False):
                enabled_count += 1
        
        permission_matrix.append({
            'role': role,
            'enabled_count': enabled_count,
            'user_count': User.objects.filter(profile__role=role, is_active=True).count()
        })
    
    # Calculate most and least permissions
    if permission_matrix:
        most_permissions = max(permission_matrix, key=lambda x: x['enabled_count'])
        least_permissions = min(permission_matrix, key=lambda x: x['enabled_count'])
    else:
        most_permissions = {'role': None, 'enabled_count': 0}
        least_permissions = {'role': None, 'enabled_count': 0}
    
    # Calculate most common permissions
    permission_counts = []
    for field in permission_fields:
        count = 0
        for role in roles:
            if getattr(role, field, False):
                count += 1
        permission_counts.append({
            'name': field,
            'count': count
        })
    
    most_common = sorted(permission_counts, key=lambda x: x['count'], reverse=True)[:5]
    least_common = sorted(permission_counts, key=lambda x: x['count'])[:5]
    
    total_permissions_count = sum(item['enabled_count'] for item in permission_matrix)
    
    context = {
        'permission_matrix': permission_matrix,
        'permission_fields': permission_fields,
        'total_roles': roles.count(),
        'most_permissions': most_permissions,
        'least_permissions': least_permissions,
        'most_common_permissions': most_common,
        'least_common_permissions': least_common,
        'total_permissions': total_permissions_count,
    }
    return render(request, 'pos/admin/permissions_summary.html', context)


@login_required
@user_passes_test(is_admin)
def get_user_permissions(request, pk):
    """Get user permissions for AJAX request"""
    user = get_object_or_404(User, pk=pk)
    
    if hasattr(user, 'profile') and user.profile:
        permissions = user.profile.get_all_permissions()
    else:
        permissions = {}
    
    return JsonResponse({
        'success': True,
        'permissions': permissions,
        'username': user.username
    })


# Helper Functions
def get_permission_categories():
    """Group permissions by category"""
    return {
        'Dashboard & POS': [
            ('can_access_dashboard', 'Access Dashboard'),
            ('can_access_pos', 'Access POS Interface'),
            ('can_process_sales', 'Process Sales'),
            ('can_edit_sales', 'Edit Sales'),
            ('can_delete_sales', 'Delete Sales'),
            ('can_view_all_sales', 'View All Sales'),
            ('can_view_own_sales', 'View Own Sales'),
        ],
        'Products Management': [
            ('can_view_products', 'View Products'),
            ('can_add_products', 'Add Products'),
            ('can_edit_products', 'Edit Products'),
            ('can_delete_products', 'Delete Products'),
            ('can_import_products', 'Import Products'),
        ],
        'Inventory Management': [
            ('can_view_inventory', 'View Inventory'),
            ('can_manage_stock', 'Manage Stock'),
            ('can_view_reports', 'View Reports'),
        ],
        'Purchases Management': [
            ('can_view_purchases', 'View Purchases'),
            ('can_add_purchases', 'Add Purchases'),
            ('can_edit_purchases', 'Edit Purchases'),
            ('can_delete_purchases', 'Delete Purchases'),
        ],
        'Customers Management': [
            ('can_view_customers', 'View Customers'),
            ('can_add_customers', 'Add Customers'),
            ('can_edit_customers', 'Edit Customers'),
            ('can_delete_customers', 'Delete Customers'),
        ],
        'Suppliers Management': [
            ('can_view_suppliers', 'View Suppliers'),
            ('can_add_suppliers', 'Add Suppliers'),
            ('can_edit_suppliers', 'Edit Suppliers'),
            ('can_delete_suppliers', 'Delete Suppliers'),
        ],
        'Reports & Analytics': [
            ('can_view_sales_reports', 'View Sales Reports'),
            ('can_view_inventory_reports', 'View Inventory Reports'),
            ('can_view_profit_reports', 'View Profit Reports'),
            ('can_view_customer_reports', 'View Customer Reports'),
            ('can_export_reports', 'Export Reports'),
        ],
        'Financial Management': [
            ('can_view_financials', 'View Financials'),
            ('can_process_refunds', 'Process Refunds'),
            ('can_view_credit_sales', 'View Credit Sales'),
            ('can_process_credit_payments', 'Process Credit Payments'),
        ],
        'System Settings': [
            ('can_manage_settings', 'Manage Settings'),
            ('can_manage_users', 'Manage Users'),
            ('can_manage_roles', 'Manage Roles'),
        ],
    }


def send_welcome_email(user, request, password=None):
    """Send welcome email to new user"""
    try:
        from django.core.mail import send_mail
        from django.template.loader import render_to_string
        from django.conf import settings
        
        subject = f"Welcome to {Company.objects.first().name} POS System"
        
        context = {
            'user': user,
            'company': Company.objects.first(),
            'password': password,
            'login_url': request.build_absolute_uri('/'),
        }
        
        message = render_to_string('pos/emails/welcome_email.html', context)
        
        send_mail(
            subject,
            message,
            settings.DEFAULT_FROM_EMAIL,
            [user.email],
            fail_silently=True,
            html_message=message
        )
        
        return True
    except Exception as e:
        print(f"Error sending welcome email: {e}")
        return False


def export_users(format, queryset):
    """Export users to CSV or Excel"""
    if format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="users_export.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            'Username', 'Email', 'First Name', 'Last Name',
            'Role', 'Phone', 'Is Active', 'Date Joined', 'Last Login'
        ])
        
        for user in queryset:
            role = user.profile.role.get_name_display() if hasattr(user, 'profile') and user.profile.role else 'No Role'
            phone = user.profile.phone if hasattr(user, 'profile') else ''
            
            writer.writerow([
                user.username,
                user.email,
                user.first_name,
                user.last_name,
                role,
                phone,
                'Yes' if user.is_active else 'No',
                user.date_joined.strftime('%Y-%m-%d %H:%M:%S'),
                user.last_login.strftime('%Y-%m-%d %H:%M:%S') if user.last_login else 'Never'
            ])
        
        return response
    
    elif format == 'excel':
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="users_export.xlsx"'
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Users"
        
        headers = [
            'Username', 'Email', 'First Name', 'Last Name',
            'Role', 'Phone', 'Is Active', 'Date Joined', 'Last Login'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
        
        row_num = 2
        for user in queryset:
            role = user.profile.role.get_name_display() if hasattr(user, 'profile') and user.profile.role else 'No Role'
            phone = user.profile.phone if hasattr(user, 'profile') else ''
            
            ws.cell(row=row_num, column=1, value=user.username)
            ws.cell(row=row_num, column=2, value=user.email)
            ws.cell(row=row_num, column=3, value=user.first_name)
            ws.cell(row=row_num, column=4, value=user.last_name)
            ws.cell(row=row_num, column=5, value=role)
            ws.cell(row=row_num, column=6, value=phone)
            ws.cell(row=row_num, column=7, value='Yes' if user.is_active else 'No')
            ws.cell(row=row_num, column=8, value=user.date_joined.strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row_num, column=9, value=user.last_login.strftime('%Y-%m-%d %H:%M:%S') if user.last_login else 'Never')
            
            row_num += 1
        
        wb.save(response)
        return response
    
    return HttpResponse('Invalid export format', status=400)


def export_roles(format, queryset):
    """Export roles to CSV or Excel"""
    if format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="roles_export.csv"'
        
        writer = csv.writer(response)
        
        # Get permission fields
        permission_fields = [f.name for f in Role._meta.get_fields() if f.name.startswith('can_')]
        
        # Write headers
        headers = ['Role Name', 'Description', 'Is Active', 'Users Count']
        headers.extend([f.replace('can_', '').replace('_', ' ').title() for f in permission_fields])
        
        writer.writerow(headers)
        
        # Write data
        for role in queryset:
            row = [
                role.get_name_display(),
                role.description,
                'Yes' if role.is_active else 'No',
                User.objects.filter(profile__role=role, is_active=True).count()
            ]
            
            # Add permission values
            for field in permission_fields:
                row.append('Yes' if getattr(role, field, False) else 'No')
            
            writer.writerow(row)
        
        return response
    
    return HttpResponse('Invalid export format', status=400)


def export_activity_logs(format, queryset):
    """Export activity logs to CSV"""
    if format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="activity_logs_export.csv"'
        
        writer = csv.writer(response)
        writer.writerow([
            'Timestamp', 'User', 'Action Type', 'Model', 
            'Object ID', 'Description', 'IP Address'
        ])
        
        for log in queryset:
            writer.writerow([
                log.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                log.user.username if log.user else 'System',
                log.get_action_type_display(),
                log.model_name,
                log.object_id or '',
                log.description,
                log.ip_address or ''
            ])
        
        return response
    
    return HttpResponse('Invalid export format', status=400)


# views.py - Add these views
# views.py - Update product_level_credit_payment view
@login_required
def product_level_credit_payment(request, pk):
    """Process credit payment at product level"""
    try:
        sale = Sale.objects.get(pk=pk)
        if not sale.is_credit:
            messages.warning(request, 'This sale is not a credit sale.')
            return redirect('credit_payments')
            
        if sale.balance <= 0:
            messages.info(request, 'This credit sale has been fully paid.')
            return redirect('credit_payments')
            
    except Sale.DoesNotExist:
        messages.error(request, f'Sale #{pk} does not exist.')
        return redirect('credit_payments')
    
    # Initialize credit items (for new credit sales)
    try:
        credit_items = []
        for sale_item in sale.items.all():
            # Get or create credit sale item
            credit_item, created = CreditSaleItem.objects.get_or_create(
                sale_item=sale_item,
                defaults={
                    'quantity_credited': sale_item.quantity,
                    'original_quantity': sale_item.quantity,
                    'price_per_unit': sale_item.price,
                    'total_amount': sale_item.total,
                    'remaining_quantity': sale_item.quantity,
                    'balance_amount': sale_item.total
                }
            )
            
            # If it exists but hasn't been initialized properly
            if not created and credit_item.quantity_credited == 0:
                credit_item.quantity_credited = sale_item.quantity
                credit_item.original_quantity = sale_item.quantity
                credit_item.price_per_unit = sale_item.price
                credit_item.total_amount = sale_item.total
                credit_item.remaining_quantity = sale_item.quantity
                credit_item.balance_amount = sale_item.total
                credit_item.save()
            
            credit_items.append(credit_item)
        
        # Get unpaid items
        unpaid_items = [item for item in credit_items if not item.is_fully_paid]
        
        # Calculate totals
        total_credited = sum(item.total_amount for item in credit_items)
        total_paid = sum(item.amount_paid for item in credit_items)
        total_balance = sum(item.balance_amount for item in credit_items)
        
        # Get payment history
        payment_history = CreditPayment.objects.filter(sale=sale).order_by('-payment_date')[:5]
        
        if request.method == 'POST':
            try:
                with transaction.atomic():
                    payment_amount = Decimal(request.POST.get('total_amount', '0.00'))
                    payment_method = request.POST.get('payment_method', 'cash')
                    reference = request.POST.get('reference', '')
                    notes = request.POST.get('notes', '')
                    
                    # Validate payment amount
                    if payment_amount <= 0:
                        messages.error(request, 'Payment amount must be greater than 0')
                        return redirect('product_level_credit_payment', pk=sale.id)
                    
                    # Validate against remaining balance
                    max_payment = min(total_balance, sale.balance)
                    if payment_amount > max_payment:
                        messages.error(request, f'Payment amount cannot exceed KSh {max_payment:.2f}')
                        return redirect('product_level_credit_payment', pk=sale.id)
                    
                    # Create credit payment record
                    credit_payment = CreditPayment.objects.create(
                        sale=sale,
                        amount=payment_amount,
                        payment_date=timezone.now().date(),
                        payment_method=payment_method,
                        reference=reference,
                        notes=notes,
                        processed_by=request.user
                    )
                    
                    # Process each product payment
                    remaining_amount = payment_amount
                    payment_details = []
                    
                    for item in unpaid_items:
                        if remaining_amount <= 0:
                            break
                        
                        # Check if payment was made for this item
                        item_key = f"amount_{item.id}"
                        item_paid_amount = Decimal(request.POST.get(item_key, '0.00'))
                        
                        if item_paid_amount > 0:
                            # Calculate how many units this amount covers
                            units_paid = min(
                                item_paid_amount / item.price_per_unit,
                                item.remaining_quantity
                            )
                            
                            # Ensure we don't exceed remaining payment amount
                            if units_paid * item.price_per_unit > remaining_amount:
                                units_paid = remaining_amount / item.price_per_unit
                            
                            actual_amount = units_paid * item.price_per_unit
                            
                            # Update credit sale item
                            item.quantity_paid += units_paid
                            item.amount_paid += actual_amount
                            item.save()
                            
                            # Create payment detail
                            payment_detail = CreditPaymentDetail.objects.create(
                                credit_payment=credit_payment,
                                credit_sale_item=item,
                                quantity_paid=units_paid,
                                amount_paid=actual_amount
                            )
                            payment_details.append(payment_detail)
                            
                            remaining_amount -= actual_amount
                    
                    # Update sale payment details
                    sale.amount_paid += payment_amount
                    sale.balance -= payment_amount
                    
                    if sale.balance <= 0:
                        sale.balance = Decimal('0.00')
                        sale.is_credit = False
                    
                    sale.save()
                    
                    # Update customer balance
                    if sale.customer:
                        customer = sale.customer
                        customer.balance -= payment_amount
                        if customer.balance < 0:
                            customer.balance = Decimal('0.00')
                        customer.save()
                    
                    # Log activity
                    UserActivityLog.objects.create(
                        user=request.user,
                        action_type='credit_payment',
                        model_name='Sale',
                        object_id=str(sale.id),
                        description=f'Product-level credit payment of KSh {payment_amount:.2f} for sale #{sale.sale_number}',
                        ip_address=request.META.get('REMOTE_ADDR'),
                        user_agent=request.META.get('HTTP_USER_AGENT', '')
                    )
                    
                    messages.success(request, f'Payment of KSh {payment_amount:.2f} recorded successfully!')
                    
                    # If all items are paid, show success message
                    all_paid = all(item.is_fully_paid for item in unpaid_items)
                    if all_paid:
                        messages.success(request, 'All products have been fully paid!')
                    
                    return redirect('product_level_credit_payment', pk=sale.id)
                    
            except Exception as e:
                messages.error(request, f'Error processing payment: {str(e)}')
                return redirect('product_level_credit_payment', pk=sale.id)
        
        context = {
            'sale': sale,
            'credit_items': unpaid_items,
            'total_credited': total_credited,
            'total_paid': total_paid,
            'total_balance': total_balance,
            'max_payment': min(total_balance, sale.balance),
            'payment_history': payment_history,
            'today': timezone.now().date(),
        }
        
        return render(request, 'pos/product_level_credit_payment.html', context)
        
    except Exception as e:
        messages.error(request, f'Error loading credit payment page: {str(e)}')
        return redirect('credit_payments')

@login_required
def credit_payment_history(request, pk):
    """View credit payment history for a sale"""
    sale = get_object_or_404(Sale, pk=pk, is_credit=True)
    credit_payments = CreditPayment.objects.filter(sale=sale).order_by('-payment_date')
    credit_items = CreditSaleItem.objects.filter(sale_item__sale=sale)
    
    context = {
        'sale': sale,
        'credit_payments': credit_payments,
        'credit_items': credit_items,
    }
    return render(request, 'pos/credit_payment_history.html', context)

@login_required
def credit_payment_detail(request, payment_id):
    """View details of a specific credit payment"""
    payment = get_object_or_404(CreditPayment, pk=payment_id)
    payment_details = payment.details.all().select_related('credit_sale_item__sale_item__product')
    
    context = {
        'payment': payment,
        'payment_details': payment_details,
    }
    return render(request, 'pos/credit_payment_detail.html', context)



# views.py - Add batch profit report view

@login_required
def batch_profit_report(request):
    """Report showing profit by batch"""
    # Date filtering
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if not start_date:
        # Default to last 30 days
        end_date = timezone.now().date()
        start_date = end_date - timedelta(days=30)
    
    # Get all batches with sales in date range
    batches = Batch.objects.filter(
        saleitem__sale__date__date__range=[start_date, end_date],
        saleitem__sale__is_completed=True
    ).distinct().select_related('product', 'product__category', 'purchase', 'purchase__supplier')
    
    # Prepare batch profit data
    batch_data = []
    total_profit = Decimal('0')
    total_revenue = Decimal('0')
    total_cost = Decimal('0')
    
    for batch in batches:
        # Get sales items for this batch in date range
        sale_items = batch.saleitem_set.filter(
            sale__date__date__range=[start_date, end_date],
            sale__is_completed=True
        )
        
        # Calculate totals
        revenue = sale_items.aggregate(total=Sum('total'))['total'] or Decimal('0')
        quantity_sold = sale_items.aggregate(total=Sum('quantity'))['total'] or Decimal('0')
        cost = quantity_sold * batch.purchase_price
        profit = revenue - cost
        
        # Calculate profit margin
        profit_margin = (profit / revenue * 100) if revenue > 0 else 0
        
        # Get batch purchase details
        purchase_info = None
        if batch.purchase_item:
            purchase_info = {
                'invoice_number': batch.purchase_item.purchase.invoice_number,
                'supplier': batch.purchase_item.purchase.supplier.name,
                'purchase_date': batch.purchase_item.purchase.date,
            }
        
        batch_data.append({
            'batch': batch,
            'product': batch.product,
            'batch_number': batch.batch_number,
            'category': batch.product.category.name if batch.product.category else 'Uncategorized',
            'supplier': batch.purchase.supplier.name if batch.purchase else 'Unknown',
            'purchase_price': batch.purchase_price,
            'selling_price': batch.selling_price,
            'quantity_purchased': batch.quantity + quantity_sold,
            'quantity_sold': quantity_sold,
            'quantity_remaining': batch.quantity,
            'revenue': revenue,
            'cost': cost,
            'profit': profit,
            'profit_margin': profit_margin,
            'purchase_date': batch.date_received,
            'purchase_info': purchase_info,
            'expiry_date': batch.expiry_date,
            'days_in_stock': (timezone.now().date() - batch.date_received).days,
        })
        
        total_profit += profit
        total_revenue += revenue
        total_cost += cost
    
    # Sort by profit (highest first)
    batch_data.sort(key=lambda x: x['profit'], reverse=True)
    
    # Calculate overall margin
    overall_margin = (total_profit / total_revenue * 100) if total_revenue > 0 else 0
    
    # Get filter options
    products = Product.objects.filter(is_active=True)
    suppliers = Supplier.objects.all()
    
    context = {
        'batch_data': batch_data,
        'total_profit': total_profit,
        'total_revenue': total_revenue,
        'total_cost': total_cost,
        'overall_margin': overall_margin,
        'total_batches': len(batch_data),
        'start_date': start_date,
        'end_date': end_date,
        'products': products,
        'suppliers': suppliers,
        'today': timezone.now().date(),
    }
    
    return render(request, 'pos/reports/batch_profit_report.html', context)



# views.py - Add expiry tracking view

@login_required
def expiry_tracking(request):
    """Track product expiry dates from batches"""
    try:
        # Get filter parameters
        days_to_expiry = request.GET.get('days', '30')  # Default to 30 days
        product_id = request.GET.get('product')
        status = request.GET.get('status', 'all')
        
        # Base queryset - get all batches with expiry dates
        today = timezone.now().date()
        thirty_days_later = today + timedelta(days=30)
        ninety_days_later = today + timedelta(days=90)
        
        batches = Batch.objects.filter(
            quantity__gt=0,  # Only batches with stock
            is_active=True
        ).select_related('product', 'product__category', 'purchase', 'purchase__supplier').order_by('expiry_date')
        
        # Apply filters
        if product_id:
            batches = batches.filter(product_id=product_id)
        
        # Status filter
        if status == 'expired':
            batches = batches.filter(expiry_date__lt=today)
        elif status == 'expiring_soon':
            batches = batches.filter(
                expiry_date__gte=today,
                expiry_date__lte=thirty_days_later
            )
        elif status == 'expiring_medium':
            batches = batches.filter(
                expiry_date__gt=thirty_days_later,
                expiry_date__lte=ninety_days_later
            )
        elif status == 'expiring_later':
            batches = batches.filter(expiry_date__gt=ninety_days_later)
        elif status == 'no_expiry':
            batches = batches.filter(expiry_date__isnull=True)
        
        # Days to expiry filter
        if days_to_expiry and days_to_expiry.isdigit() and status == 'all':
            cutoff_date = today + timedelta(days=int(days_to_expiry))
            batches = batches.filter(
                expiry_date__lte=cutoff_date,
                expiry_date__gte=today
            )
        
        # Calculate statistics
        total_batches = batches.count()
        total_quantity = batches.aggregate(total=Sum('quantity'))['total'] or Decimal('0')
        total_value = batches.aggregate(
            total=Sum(F('quantity') * F('purchase_price'))
        )['total'] or Decimal('0')
        
        # Expiry statistics
        expired_count = Batch.objects.filter(
            quantity__gt=0,
            is_active=True,
            expiry_date__lt=today
        ).count()
        
        expiring_soon_count = Batch.objects.filter(
            quantity__gt=0,
            is_active=True,
            expiry_date__gte=today,
            expiry_date__lte=thirty_days_later
        ).count()
        
        expiring_medium_count = Batch.objects.filter(
            quantity__gt=0,
            is_active=True,
            expiry_date__gt=thirty_days_later,
            expiry_date__lte=ninety_days_later
        ).count()
        
        no_expiry_count = Batch.objects.filter(
            quantity__gt=0,
            is_active=True,
            expiry_date__isnull=True
        ).count()
        
        # Prepare batch data with expiry information
        batch_data = []
        total_at_risk_value = Decimal('0')
        
        for batch in batches:
            days_until_expiry = None
            status_class = ''
            status_text = ''
            
            if batch.expiry_date:
                days_until_expiry = (batch.expiry_date - today).days
                
                if days_until_expiry < 0:
                    status_class = 'bg-red-100 text-red-800'
                    status_text = 'Expired'
                    total_at_risk_value += batch.quantity * batch.purchase_price
                elif days_until_expiry <= 30:
                    status_class = 'bg-orange-100 text-orange-800'
                    status_text = f'Expiring in {days_until_expiry} days'
                    total_at_risk_value += batch.quantity * batch.purchase_price
                elif days_until_expiry <= 90:
                    status_class = 'bg-yellow-100 text-yellow-800'
                    status_text = f'Expiring in {days_until_expiry} days'
                else:
                    status_class = 'bg-green-100 text-green-800'
                    status_text = f'Valid ({days_until_expiry} days)'
            else:
                status_class = 'bg-gray-100 text-gray-800'
                status_text = 'No expiry date'
            
            # Calculate potential loss if expired
            potential_loss = Decimal('0')
            if batch.expiry_date and batch.expiry_date < today:
                potential_loss = batch.quantity * batch.purchase_price
            
            batch_data.append({
                'batch': batch,
                'product': batch.product,
                'category': batch.product.category.name if batch.product.category else 'Uncategorized',
                'supplier': batch.purchase.supplier.name if batch.purchase else 'Unknown',
                'batch_number': batch.batch_number,
                'quantity': batch.quantity,
                'purchase_price': batch.purchase_price,
                'selling_price': batch.selling_price,
                'total_cost': batch.quantity * batch.purchase_price,
                'potential_revenue': batch.quantity * batch.selling_price,
                'potential_profit': (batch.quantity * batch.selling_price) - (batch.quantity * batch.purchase_price),
                'expiry_date': batch.expiry_date,
                'days_until_expiry': days_until_expiry,
                'status_class': status_class,
                'status_text': status_text,
                'potential_loss': potential_loss,
                'date_received': batch.date_received,
                'purchase_invoice': batch.purchase.invoice_number if batch.purchase else 'N/A',
            })
        
        # Get products for filter dropdown
        products = Product.objects.filter(is_active=True).order_by('name')
        
        # Export functionality
        export_format = request.GET.get('export')
        if export_format:
            return export_expiry_report(export_format, batch_data, today)
        
        context = {
            'batch_data': batch_data,
            'products': products,
            'total_batches': total_batches,
            'total_quantity': total_quantity,
            'total_value': total_value,
            'total_at_risk_value': total_at_risk_value,
            'expired_count': expired_count,
            'expiring_soon_count': expiring_soon_count,
            'expiring_medium_count': expiring_medium_count,
            'no_expiry_count': no_expiry_count,
            'today': today,
            'selected_days': days_to_expiry,
            'selected_product': product_id,
            'selected_status': status,
        }
        
        return render(request, 'pos/reports/expiry_tracking.html', context)
        
    except Exception as e:
        messages.error(request, f'Error generating expiry report: {str(e)}')
        return redirect('reports_dashboard')


# views.py - Fixed export_expiry_report function

def export_expiry_report(export_format, batch_data, report_date):
    """Export expiry report to CSV or Excel"""
    
    if export_format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="expiry_tracking_{report_date}.csv"'
        
        writer = csv.writer(response)
        
        # Write headers
        writer.writerow([
            'Product Name', 'Category', 'Supplier', 'Batch Number',
            'Quantity', 'Purchase Price', 'Selling Price', 'Total Cost',
            'Potential Revenue', 'Potential Profit', 'Expiry Date',
            'Days Until Expiry', 'Status', 'Date Received', 'Purchase Invoice'
        ])
        
        # Write data
        for item in batch_data:
            writer.writerow([
                item['product'].name,
                item['category'],
                item['supplier'],
                item['batch_number'],
                float(item['quantity']),
                float(item['purchase_price']),
                float(item['selling_price']),
                float(item['total_cost']),
                float(item['potential_revenue']),
                float(item['potential_profit']),
                item['expiry_date'].strftime('%Y-%m-%d') if item['expiry_date'] else 'N/A',
                item['days_until_expiry'] if item['days_until_expiry'] else 'N/A',
                item['status_text'],
                item['date_received'].strftime('%Y-%m-%d'),
                item['purchase_invoice'],
            ])
        
        return response
    
    elif export_format == 'excel':
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            # If openpyxl is not installed
            response = HttpResponse(content_type='text/plain')
            response['Content-Disposition'] = f'attachment; filename="error.txt"'
            response.write("Excel export requires openpyxl. Please install it with: pip install openpyxl")
            return response
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="expiry_tracking_{report_date}.xlsx"'
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Expiry Tracking"
        
        # Title
        ws.merge_cells('A1:O1')
        title_cell = ws['A1']
        title_cell.value = f"EXPIRY DATE TRACKING REPORT - Generated on {report_date.strftime('%Y-%m-%d')}"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center')
        
        # Generation info
        ws['A2'] = f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(italic=True)
        
        # Headers
        headers = [
            'Product Name', 'Category', 'Supplier', 'Batch Number',
            'Quantity', 'Purchase Price', 'Selling Price', 'Total Cost',
            'Potential Revenue', 'Potential Profit', 'Expiry Date',
            'Days Until Expiry', 'Status', 'Date Received', 'Purchase Invoice'
        ]
        
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        row_num = 5
        for item in batch_data:
            # Product Name
            ws.cell(row=row_num, column=1, value=item['product'].name)
            
            # Category
            ws.cell(row=row_num, column=2, value=item['category'])
            
            # Supplier
            ws.cell(row=row_num, column=3, value=item['supplier'])
            
            # Batch Number
            ws.cell(row=row_num, column=4, value=item['batch_number'])
            
            # Quantity
            ws.cell(row=row_num, column=5, value=float(item['quantity']))
            
            # Purchase Price
            ws.cell(row=row_num, column=6, value=float(item['purchase_price']))
            
            # Selling Price
            ws.cell(row=row_num, column=7, value=float(item['selling_price']))
            
            # Total Cost
            cost_cell = ws.cell(row=row_num, column=8, value=float(item['total_cost']))
            cost_cell.number_format = '#,##0.00'
            
            # Potential Revenue
            revenue_cell = ws.cell(row=row_num, column=9, value=float(item['potential_revenue']))
            revenue_cell.number_format = '#,##0.00'
            
            # Potential Profit with conditional formatting
            profit_cell = ws.cell(row=row_num, column=10, value=float(item['potential_profit']))
            profit_cell.number_format = '#,##0.00'
            if item['potential_profit'] >= 0:
                profit_cell.font = Font(color='006100', bold=True)
            else:
                profit_cell.font = Font(color='FF0000', bold=True)
            
            # Expiry Date
            if item['expiry_date']:
                ws.cell(row=row_num, column=11, value=item['expiry_date'].strftime('%Y-%m-%d'))
            else:
                ws.cell(row=row_num, column=11, value='N/A')
            
            # Days Until Expiry
            if item['days_until_expiry'] is not None:
                days_cell = ws.cell(row=row_num, column=12, value=item['days_until_expiry'])
                if item['days_until_expiry'] < 0:
                    days_cell.font = Font(color='FF0000', bold=True)
                elif item['days_until_expiry'] <= 30:
                    days_cell.font = Font(color='FFA500', bold=True)
            else:
                ws.cell(row=row_num, column=12, value='N/A')
            
            # Status with color
            status_cell = ws.cell(row=row_num, column=13, value=item['status_text'])
            if 'Expired' in item['status_text']:
                status_cell.font = Font(color='FF0000', bold=True)
            elif 'Expiring soon' in item['status_text'] or 'Expiring in' in item['status_text']:
                if '30' in item['status_text']:
                    status_cell.font = Font(color='FFA500', bold=True)
            
            # Date Received
            ws.cell(row=row_num, column=14, value=item['date_received'].strftime('%Y-%m-%d'))
            
            # Purchase Invoice
            ws.cell(row=row_num, column=15, value=item['purchase_invoice'])
            
            row_num += 1
        
        # Add summary at the bottom
        summary_row = row_num + 2
        
        # Calculate totals
        total_batches = len(batch_data)
        total_quantity = sum(item['quantity'] for item in batch_data)
        total_value = sum(item['total_cost'] for item in batch_data)
        total_at_risk = sum(item['total_cost'] for item in batch_data if item['days_until_expiry'] is not None and item['days_until_expiry'] <= 30)
        
        ws.merge_cells(f'A{summary_row}:C{summary_row}')
        summary_title = ws.cell(row=summary_row, column=1, value="SUMMARY STATISTICS")
        summary_title.font = Font(bold=True, size=12)
        
        summary_data = [
            ("Total Batches:", total_batches, f"C{summary_row+1}"),
            ("Total Quantity:", f"{float(total_quantity):,.2f}", f"C{summary_row+2}"),
            ("Total Inventory Value:", f"KSh {float(total_value):,.2f}", f"C{summary_row+3}"),
            ("At-Risk Value (≤30 days):", f"KSh {float(total_at_risk):,.2f}", f"C{summary_row+4}"),
        ]
        
        for i, (label, value, cell_ref) in enumerate(summary_data, 1):
            ws.cell(row=summary_row + i, column=1, value=label).font = Font(bold=True)
            ws.cell(row=summary_row + i, column=2, value=value)
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            
            # Check header length
            header_length = len(headers[col-1])
            max_length = max(max_length, header_length)
            
            # Check data rows
            for row in range(5, row_num):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    cell_length = len(str(cell_value))
                    max_length = max(max_length, cell_length)
            
            # Set width with a cap
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add borders to the data table
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in range(4, row_num):
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
        
        wb.save(response)
        return response
    
    return HttpResponse('Invalid export format', status=400)



# views.py - Add expected profits view

@login_required
def expected_profits_report(request):
    """Calculate expected profits from current inventory based on batch data"""
    try:
        # Get filter parameters
        category_id = request.GET.get('category')
        supplier_id = request.GET.get('supplier')
        product_id = request.GET.get('product')
        min_profit = request.GET.get('min_profit')
        sort_by = request.GET.get('sort', 'profit_desc')
        
        # Base queryset - get all active batches with stock
        batches = Batch.objects.filter(
            quantity__gt=0,
            is_active=True
        ).select_related('product', 'product__category', 'purchase', 'purchase__supplier')
        
        # Apply filters
        if category_id:
            batches = batches.filter(product__category_id=category_id)
        
        if supplier_id:
            batches = batches.filter(purchase__supplier_id=supplier_id)
        
        if product_id:
            batches = batches.filter(product_id=product_id)
        
        # Calculate expected profits by batch
        batch_profits = []
        total_inventory_cost = Decimal('0')
        total_expected_revenue = Decimal('0')
        total_expected_profit = Decimal('0')
        total_quantity = Decimal('0')
        
        for batch in batches:
            # Calculate expected revenue and profit
            quantity = batch.quantity
            purchase_price = batch.purchase_price
            selling_price = batch.selling_price
            
            cost_value = quantity * purchase_price
            revenue_value = quantity * selling_price
            expected_profit = revenue_value - cost_value
            profit_margin = (expected_profit / revenue_value * 100) if revenue_value > 0 else 0
            
            # Get purchase info
            purchase_info = None
            if batch.purchase:
                purchase_info = {
                    'invoice': batch.purchase.invoice_number,
                    'date': batch.purchase.date,
                    'supplier': batch.purchase.supplier.name if batch.purchase.supplier else 'Unknown'
                }
            
            # Check if batch is near expiry (for risk assessment)
            days_to_expiry = None
            if batch.expiry_date:
                days_to_expiry = (batch.expiry_date - timezone.now().date()).days
            
            batch_profits.append({
                'batch': batch,
                'product': batch.product,
                'category': batch.product.category.name if batch.product.category else 'Uncategorized',
                'supplier': batch.purchase.supplier.name if batch.purchase and batch.purchase.supplier else 'Unknown',
                'batch_number': batch.batch_number,
                'quantity': quantity,
                'purchase_price': purchase_price,
                'selling_price': selling_price,
                'cost_value': cost_value,
                'revenue_value': revenue_value,
                'expected_profit': expected_profit,
                'profit_margin': profit_margin,
                'expiry_date': batch.expiry_date,
                'days_to_expiry': days_to_expiry,
                'purchase_info': purchase_info,
                'date_received': batch.date_received,
            })
            
            # Update totals
            total_inventory_cost += cost_value
            total_expected_revenue += revenue_value
            total_expected_profit += expected_profit
            total_quantity += quantity
        
        # Apply minimum profit filter
        if min_profit and min_profit.replace('.', '').isdigit():
            min_profit_val = Decimal(min_profit)
            batch_profits = [b for b in batch_profits if b['expected_profit'] >= min_profit_val]
        
        # Sort by selected criteria
        if sort_by == 'profit_desc':
            batch_profits.sort(key=lambda x: x['expected_profit'], reverse=True)
        elif sort_by == 'profit_asc':
            batch_profits.sort(key=lambda x: x['expected_profit'])
        elif sort_by == 'margin_desc':
            batch_profits.sort(key=lambda x: x['profit_margin'], reverse=True)
        elif sort_by == 'margin_asc':
            batch_profits.sort(key=lambda x: x['profit_margin'])
        elif sort_by == 'quantity_desc':
            batch_profits.sort(key=lambda x: x['quantity'], reverse=True)
        elif sort_by == 'value_desc':
            batch_profits.sort(key=lambda x: x['cost_value'], reverse=True)
        
        # Calculate summary statistics
        avg_profit_margin = (total_expected_profit / total_expected_revenue * 100) if total_expected_revenue > 0 else 0
        
        # Group by category for breakdown
        category_breakdown = {}
        for item in batch_profits:
            cat = item['category']
            if cat not in category_breakdown:
                category_breakdown[cat] = {
                    'quantity': Decimal('0'),
                    'cost': Decimal('0'),
                    'revenue': Decimal('0'),
                    'profit': Decimal('0'),
                }
            category_breakdown[cat]['quantity'] += item['quantity']
            category_breakdown[cat]['cost'] += item['cost_value']
            category_breakdown[cat]['revenue'] += item['revenue_value']
            category_breakdown[cat]['profit'] += item['expected_profit']
        
        # Group by supplier for breakdown
        supplier_breakdown = {}
        for item in batch_profits:
            sup = item['supplier']
            if sup not in supplier_breakdown:
                supplier_breakdown[sup] = {
                    'quantity': Decimal('0'),
                    'cost': Decimal('0'),
                    'revenue': Decimal('0'),
                    'profit': Decimal('0'),
                }
            supplier_breakdown[sup]['quantity'] += item['quantity']
            supplier_breakdown[sup]['cost'] += item['cost_value']
            supplier_breakdown[sup]['revenue'] += item['revenue_value']
            supplier_breakdown[sup]['profit'] += item['expected_profit']
        
        # Get filter dropdown data
        categories = Category.objects.all().order_by('name')
        suppliers = Supplier.objects.all().order_by('name')
        products = Product.objects.filter(is_active=True).order_by('name')
        
        # Export functionality
        export_format = request.GET.get('export')
        if export_format:
            return export_expected_profits(export_format, batch_profits, category_breakdown, supplier_breakdown)
        
        context = {
            'batch_profits': batch_profits,
            'category_breakdown': category_breakdown,
            'supplier_breakdown': supplier_breakdown,
            'total_inventory_cost': total_inventory_cost,
            'total_expected_revenue': total_expected_revenue,
            'total_expected_profit': total_expected_profit,
            'avg_profit_margin': avg_profit_margin,
            'total_quantity': total_quantity,
            'total_batches': len(batch_profits),
            'categories': categories,
            'suppliers': suppliers,
            'products': products,
            'selected_category': category_id,
            'selected_supplier': supplier_id,
            'selected_product': product_id,
            'selected_min_profit': min_profit,
            'selected_sort': sort_by,
            'today': timezone.now().date(),
            'sort_options': [
                ('profit_desc', 'Profit (High to Low)'),
                ('profit_asc', 'Profit (Low to High)'),
                ('margin_desc', 'Margin (High to Low)'),
                ('margin_asc', 'Margin (Low to High)'),
                ('quantity_desc', 'Quantity (High to Low)'),
                ('value_desc', 'Inventory Value (High to Low)'),
            ],
        }
        
        return render(request, 'pos/reports/expected_profits.html', context)
        
    except Exception as e:
        messages.error(request, f'Error generating expected profits report: {str(e)}')
        return redirect('reports_dashboard')


# views.py - Fixed export_expected_profits function

def export_expected_profits(export_format, batch_profits, category_breakdown, supplier_breakdown):
    """Export expected profits report to CSV or Excel"""
    
    if export_format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="expected_profits_{timezone.now().date()}.csv"'
        
        writer = csv.writer(response)
        
        # Write main data headers
        writer.writerow(['EXPECTED PROFITS FROM INVENTORY - DETAILED REPORT'])
        writer.writerow(['Generated:', timezone.now().strftime('%Y-%m-%d %H:%M:%S')])
        writer.writerow([])
        
        # Detailed batch data
        writer.writerow(['BATCH LEVEL DETAILS'])
        writer.writerow([
            'Product', 'Category', 'Supplier', 'Batch Number',
            'Quantity', 'Purchase Price', 'Selling Price',
            'Cost Value', 'Expected Revenue', 'Expected Profit',
            'Profit Margin %', 'Expiry Date', 'Days to Expiry'
        ])
        
        for item in batch_profits:
            writer.writerow([
                item['product'].name,
                item['category'],
                item['supplier'],
                item['batch_number'],
                float(item['quantity']),
                float(item['purchase_price']),
                float(item['selling_price']),
                float(item['cost_value']),
                float(item['revenue_value']),
                float(item['expected_profit']),
                f"{float(item['profit_margin']):.2f}",
                item['expiry_date'].strftime('%Y-%m-%d') if item['expiry_date'] else 'N/A',
                item['days_to_expiry'] if item['days_to_expiry'] else 'N/A',
            ])
        
        # Category breakdown
        writer.writerow([])
        writer.writerow(['CATEGORY BREAKDOWN'])
        writer.writerow(['Category', 'Quantity', 'Cost Value', 'Expected Revenue', 'Expected Profit'])
        
        for cat, data in category_breakdown.items():
            writer.writerow([
                cat,
                float(data['quantity']),
                float(data['cost']),
                float(data['revenue']),
                float(data['profit']),
            ])
        
        # Supplier breakdown
        writer.writerow([])
        writer.writerow(['SUPPLIER BREAKDOWN'])
        writer.writerow(['Supplier', 'Quantity', 'Cost Value', 'Expected Revenue', 'Expected Profit'])
        
        for sup, data in supplier_breakdown.items():
            writer.writerow([
                sup,
                float(data['quantity']),
                float(data['cost']),
                float(data['revenue']),
                float(data['profit']),
            ])
        
        # Summary
        writer.writerow([])
        writer.writerow(['SUMMARY'])
        writer.writerow(['Total Batches', len(batch_profits)])
        writer.writerow(['Total Quantity', float(sum(b['quantity'] for b in batch_profits))])
        writer.writerow(['Total Inventory Cost', float(sum(b['cost_value'] for b in batch_profits))])
        writer.writerow(['Total Expected Revenue', float(sum(b['revenue_value'] for b in batch_profits))])
        writer.writerow(['Total Expected Profit', float(sum(b['expected_profit'] for b in batch_profits))])
        
        return response
    
    elif export_format == 'excel':
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            # If openpyxl is not installed
            response = HttpResponse(content_type='text/plain')
            response['Content-Disposition'] = f'attachment; filename="error.txt"'
            response.write("Excel export requires openpyxl. Please install it with: pip install openpyxl")
            return response
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="expected_profits_{timezone.now().date()}.xlsx"'
        
        wb = Workbook()
        
        # Sheet 1: Detailed Batch Data
        ws1 = wb.active
        ws1.title = "Batch Details"
        
        # Title
        ws1.merge_cells('A1:M1')
        title_cell = ws1['A1']
        title_cell.value = f"EXPECTED PROFITS FROM INVENTORY - Detailed Batch Report"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center')
        
        # Generation info
        ws1['A2'] = f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws1['A2'].font = Font(italic=True)
        
        # Headers
        headers = [
            'Product', 'Category', 'Supplier', 'Batch Number',
            'Quantity', 'Purchase Price', 'Selling Price',
            'Cost Value', 'Expected Revenue', 'Expected Profit',
            'Profit Margin %', 'Expiry Date', 'Days to Expiry'
        ]
        
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws1.cell(row=4, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        row_num = 5
        for item in batch_profits:
            ws1.cell(row=row_num, column=1, value=item['product'].name)
            ws1.cell(row=row_num, column=2, value=item['category'])
            ws1.cell(row=row_num, column=3, value=item['supplier'])
            ws1.cell(row=row_num, column=4, value=item['batch_number'])
            ws1.cell(row=row_num, column=5, value=float(item['quantity']))
            ws1.cell(row=row_num, column=6, value=float(item['purchase_price']))
            ws1.cell(row=row_num, column=7, value=float(item['selling_price']))
            ws1.cell(row=row_num, column=8, value=float(item['cost_value']))
            ws1.cell(row=row_num, column=9, value=float(item['revenue_value']))
            
            # Profit with color
            profit_cell = ws1.cell(row=row_num, column=10, value=float(item['expected_profit']))
            if item['expected_profit'] >= 0:
                profit_cell.font = Font(color='006100', bold=True)
            else:
                profit_cell.font = Font(color='FF0000', bold=True)
            
            # Margin with color
            margin_cell = ws1.cell(row=row_num, column=11, value=float(item['profit_margin']))
            if item['profit_margin'] >= 20:
                margin_cell.font = Font(color='006100', bold=True)
            elif item['profit_margin'] >= 10:
                margin_cell.font = Font(color='FFA500', bold=True)
            else:
                margin_cell.font = Font(color='FF0000', bold=True)
            
            ws1.cell(row=row_num, column=12, value=item['expiry_date'].strftime('%Y-%m-%d') if item['expiry_date'] else 'N/A')
            ws1.cell(row=row_num, column=13, value=item['days_to_expiry'] if item['days_to_expiry'] else 'N/A')
            
            row_num += 1
        
        # Auto-adjust column widths for Sheet 1
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            for row in range(4, row_num):
                cell_value = ws1.cell(row=row, column=col).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            adjusted_width = min(max_length + 2, 30)
            ws1.column_dimensions[column_letter].width = adjusted_width
        
        # Sheet 2: Category Breakdown
        ws2 = wb.create_sheet(title="Category Breakdown")
        
        ws2.merge_cells('A1:E1')
        ws2['A1'] = "CATEGORY BREAKDOWN"
        ws2['A1'].font = Font(bold=True, size=14)
        ws2['A1'].alignment = Alignment(horizontal='center')
        
        cat_headers = ['Category', 'Quantity', 'Cost Value', 'Expected Revenue', 'Expected Profit']
        for col_num, header in enumerate(cat_headers, 1):
            cell = ws2.cell(row=3, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.font = Font(color='FFFFFF', bold=True)
        
        row_num = 4
        for cat, data in category_breakdown.items():
            ws2.cell(row=row_num, column=1, value=cat)
            ws2.cell(row=row_num, column=2, value=float(data['quantity']))
            ws2.cell(row=row_num, column=3, value=float(data['cost']))
            ws2.cell(row=row_num, column=4, value=float(data['revenue']))
            
            profit_cell = ws2.cell(row=row_num, column=5, value=float(data['profit']))
            if data['profit'] >= 0:
                profit_cell.font = Font(color='006100')
            else:
                profit_cell.font = Font(color='FF0000')
            
            row_num += 1
        
        # Auto-adjust column widths for Sheet 2
        for col in range(1, 6):
            column_letter = get_column_letter(col)
            max_length = 0
            for row in range(3, row_num):
                cell_value = ws2.cell(row=row, column=col).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            adjusted_width = min(max_length + 2, 25)
            ws2.column_dimensions[column_letter].width = adjusted_width
        
        # Sheet 3: Supplier Breakdown
        ws3 = wb.create_sheet(title="Supplier Breakdown")
        
        ws3.merge_cells('A1:E1')
        ws3['A1'] = "SUPPLIER BREAKDOWN"
        ws3['A1'].font = Font(bold=True, size=14)
        ws3['A1'].alignment = Alignment(horizontal='center')
        
        sup_headers = ['Supplier', 'Quantity', 'Cost Value', 'Expected Revenue', 'Expected Profit']
        for col_num, header in enumerate(sup_headers, 1):
            cell = ws3.cell(row=3, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.font = Font(color='FFFFFF', bold=True)
        
        row_num = 4
        for sup, data in supplier_breakdown.items():
            ws3.cell(row=row_num, column=1, value=sup)
            ws3.cell(row=row_num, column=2, value=float(data['quantity']))
            ws3.cell(row=row_num, column=3, value=float(data['cost']))
            ws3.cell(row=row_num, column=4, value=float(data['revenue']))
            
            profit_cell = ws3.cell(row=row_num, column=5, value=float(data['profit']))
            if data['profit'] >= 0:
                profit_cell.font = Font(color='006100')
            else:
                profit_cell.font = Font(color='FF0000')
            
            row_num += 1
        
        # Auto-adjust column widths for Sheet 3
        for col in range(1, 6):
            column_letter = get_column_letter(col)
            max_length = 0
            for row in range(3, row_num):
                cell_value = ws3.cell(row=row, column=col).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            adjusted_width = min(max_length + 2, 25)
            ws3.column_dimensions[column_letter].width = adjusted_width
        
        # Sheet 4: Summary
        ws4 = wb.create_sheet(title="Summary")
        
        ws4.merge_cells('A1:B1')
        ws4['A1'] = "SUMMARY STATISTICS"
        ws4['A1'].font = Font(bold=True, size=14)
        ws4['A1'].alignment = Alignment(horizontal='center')
        
        # Calculate totals
        total_quantity = sum(b['quantity'] for b in batch_profits)
        total_cost = sum(b['cost_value'] for b in batch_profits)
        total_revenue = sum(b['revenue_value'] for b in batch_profits)
        total_profit = sum(b['expected_profit'] for b in batch_profits)
        avg_margin = (total_profit / total_revenue * 100) if total_revenue > 0 else 0
        
        summary_data = [
            ("Report Generated:", timezone.now().strftime('%Y-%m-%d %H:%M:%S')),
            ("Total Batches:", len(batch_profits)),
            ("Total Quantity:", f"{float(total_quantity):,.2f}"),
            ("Total Inventory Cost:", f"KSh {float(total_cost):,.2f}"),
            ("Total Expected Revenue:", f"KSh {float(total_revenue):,.2f}"),
            ("Total Expected Profit:", f"KSh {float(total_profit):,.2f}"),
            ("Average Profit Margin:", f"{float(avg_margin):.2f}%"),
        ]
        
        for i, (label, value) in enumerate(summary_data, 3):
            ws4.cell(row=i, column=1, value=label).font = Font(bold=True)
            ws4.cell(row=i, column=2, value=value)
        
        # Add some formatting to Summary sheet
        for col in range(1, 3):
            column_letter = get_column_letter(col)
            ws4.column_dimensions[column_letter].width = 25
        
        wb.save(response)
        return response
    
    return HttpResponse('Invalid export format', status=400)


# views.py - Add/Update batch detail view

@login_required
def batch_detail(request, pk):
    """View detailed information about a specific batch"""
    batch = get_object_or_404(
        Batch.objects.select_related(
            'product', 
            'product__category',
            'purchase', 
            'purchase__supplier',
            'purchase_item'
        ), 
        pk=pk
    )
    
    # Get all sales of this batch
    sales = SaleItem.objects.filter(
        batch=batch,
        sale__is_completed=True
    ).select_related('sale', 'sale__customer').order_by('-sale__date')
    
    # Calculate sales statistics
    total_sold = sales.aggregate(total=Sum('quantity'))['total'] or Decimal('0')
    total_revenue = sales.aggregate(total=Sum('total'))['total'] or Decimal('0')
    total_cost = total_sold * batch.purchase_price
    total_profit = total_revenue - total_cost
    
    # Get stock movements (if you have StockJournal)
    stock_movements = StockJournalItem.objects.filter(
        batch=batch
    ).select_related('journal').order_by('-journal__date')
    
    # Calculate days in stock
    days_in_stock = (timezone.now().date() - batch.date_received).days
    
    # Check if batch can be deleted
    can_delete = not sales.exists() and batch.quantity == 0
    
    # Get profit margin
    profit_margin = ((batch.selling_price - batch.purchase_price) / batch.purchase_price * 100) if batch.purchase_price > 0 else 0
    
    # Get purchase details
    purchase_details = None
    if batch.purchase:
        purchase_details = {
            'invoice': batch.purchase.invoice_number,
            'date': batch.purchase.date,
            'supplier': batch.purchase.supplier,
            'total_amount': batch.purchase.total,
            'payment_status': batch.purchase.payment_status,
        }
    
    context = {
        'batch': batch,
        'sales': sales,
        'stock_movements': stock_movements,
        'total_sold': total_sold,
        'total_revenue': total_revenue,
        'total_cost': total_cost,
        'total_profit': total_profit,
        'profit_margin': profit_margin,
        'days_in_stock': days_in_stock,
        'can_delete': can_delete,
        'purchase_details': purchase_details,
        'remaining_value': batch.quantity * batch.purchase_price,
        'potential_revenue': batch.quantity * batch.selling_price,
        'potential_profit': (batch.quantity * batch.selling_price) - (batch.quantity * batch.purchase_price),
        'today': timezone.now().date(),
    }
    
    return render(request, 'pos/batch_detail.html', context)


@login_required
@require_POST
def delete_batch(request, pk):
    """Delete a batch if it has no sales and zero quantity"""
    batch = get_object_or_404(Batch, pk=pk)
    
    # Check if batch can be deleted
    has_sales = SaleItem.objects.filter(batch=batch).exists()
    
    if has_sales:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'message': 'Cannot delete batch because it has associated sales records.'
            }, status=400)
        messages.error(request, 'Cannot delete batch because it has associated sales records.')
        return redirect('batch_detail', pk=batch.id)
    
    if batch.quantity > 0:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'message': f'Cannot delete batch because it still has {batch.quantity} units in stock.'
            }, status=400)
        messages.error(request, f'Cannot delete batch because it still has {batch.quantity} units in stock.')
        return redirect('batch_detail', pk=batch.id)
    
    try:
        with transaction.atomic():
            batch_number = batch.batch_number
            product_name = batch.product.name if batch.product else 'Unknown'
            
            # Log the deletion
            UserActivityLog.objects.create(
                user=request.user,
                action_type='delete',
                model_name='Batch',
                object_id=str(batch.id),
                description=f'Deleted batch {batch_number} for product {product_name}',
                ip_address=request.META.get('REMOTE_ADDR'),
                user_agent=request.META.get('HTTP_USER_AGENT', '')
            )
            
            batch.delete()
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': f'Batch {batch_number} deleted successfully!',
                    'redirect_url': reverse('batch_list')
                })
            
            messages.success(request, f'Batch {batch_number} deleted successfully!')
            return redirect('batch_list')
            
    except Exception as e:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'message': f'Error deleting batch: {str(e)}'
            }, status=500)
        messages.error(request, f'Error deleting batch: {str(e)}')
        return redirect('batch_detail', pk=batch.id)



import requests
import json
import hmac
import hashlib
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.conf import settings

# Add these settings to your settings.py
# PAYSTACK_SECRET_KEY = 'sk_live_49057bf7256d0e2f681337fbedf0258b1262291a'
# PAYSTACK_PUBLIC_KEY = 'pk_live_0208a917a7e468760e03bd06c22d4805f31ae91e'
# MPESA_TILL_NUMBER = 'your_till_number_here'  # Add your MPesa till number

def initialize_paystack_payment(request):
    """
    Initialize Paystack payment (for card, QR, etc.)
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
        amount = int(float(data.get('amount')) * 100)  # Paystack uses kobo/cent
        email = data.get('email', request.user.email if request.user.is_authenticated else 'customer@example.com')
        sale_id = data.get('sale_id')
        
        # Prepare metadata
        metadata = {
            'sale_id': sale_id,
            'customer_name': data.get('customer_name', 'Walk-in Customer'),
            'payment_type': 'pos_sale',
            'payment_methods': data.get('payment_methods', ['card', 'qr', 'bank_transfer'])
        }
        
        # Initialize transaction
        headers = {
            'Authorization': f'Bearer {settings.PAYSTACK_SECRET_KEY}',
            'Content-Type': 'application/json'
        }
        
        payload = {
            'email': email,
            'amount': amount,
            'metadata': metadata,
            'channels': ['card', 'qr', 'bank_transfer'],  # Enable multiple channels
            'callback_url': request.build_absolute_uri('/pos/payment-callback/')
        }
        
        response = requests.post(
            'https://api.paystack.co/transaction/initialize',
            json=payload,
            headers=headers
        )
        
        if response.status_code == 200:
            result = response.json()
            if result['status']:
                return JsonResponse({
                    'success': True,
                    'authorization_url': result['data']['authorization_url'],
                    'reference': result['data']['reference'],
                    'access_code': result['data']['access_code']
                })
        
        return JsonResponse({
            'success': False,
            'message': 'Failed to initialize payment'
        }, status=400)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


def initialize_mpesa_payment(request):
    """
    Initialize MPesa payment using Paystack's MPesa integration
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        data = json.loads(request.body)
        phone_number = data.get('phone_number', '').strip()
        amount = int(float(data.get('amount')))
        sale_id = data.get('sale_id')
        
        # Validate phone number
        if not phone_number:
            return JsonResponse({
                'success': False,
                'message': 'Phone number is required for MPesa payment'
            }, status=400)
        
        # Format phone number (ensure it starts with 254)
        if phone_number.startswith('0'):
            phone_number = '254' + phone_number[1:]
        elif phone_number.startswith('+'):
            phone_number = phone_number[1:]
        
        # Paystack MPesa API endpoint
        headers = {
            'Authorization': f'Bearer {settings.PAYSTACK_SECRET_KEY}',
            'Content-Type': 'application/json'
        }
        
        # MPesa specific payload
        payload = {
            'email': data.get('email', 'customer@example.com'),
            'amount': amount * 100,  # Convert to cents
            'currency': 'KES',
            'mobile_money': {
                'phone': phone_number,
                'provider': 'mpesa'
            },
            'metadata': {
                'sale_id': sale_id,
                'till_number': settings.MPESA_TILL_NUMBER,
                'payment_type': 'mpesa_pos'
            },
            'callback_url': request.build_absolute_uri('/pos/mpesa-callback/')
        }
        
        # Initialize MPesa charge
        response = requests.post(
            'https://api.paystack.co/charge',
            json=payload,
            headers=headers
        )
        
        if response.status_code == 200:
            result = response.json()
            if result['status']:
                return JsonResponse({
                    'success': True,
                    'reference': result['data']['reference'],
                    'message': 'Please check your phone for STK push prompt'
                })
        
        return JsonResponse({
            'success': False,
            'message': 'Failed to initialize MPesa payment'
        }, status=400)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


def generate_qr_payment(request, sale_id):
    """
    Generate QR code for payment
    """
    try:
        sale = Sale.objects.get(id=sale_id)
        
        # Prepare payment data for QR code
        payment_data = {
            'sale_id': sale.id,
            'amount': float(sale.total),
            'merchant': 'Succeed Cereal Hub',
            'till_number': settings.MPESA_TILL_NUMBER,
            'paystack_public_key': settings.PAYSTACK_PUBLIC_KEY
        }
        
        # Generate QR code URL using Paystack's QR code endpoint
        headers = {
            'Authorization': f'Bearer {settings.PAYSTACK_SECRET_KEY}',
            'Content-Type': 'application/json'
        }
        
        response = requests.post(
            'https://api.paystack.co/qr/create',
            json={
                'amount': int(sale.total * 100),
                'type': 'dynamic',
                'metadata': {
                    'sale_id': sale.id
                }
            },
            headers=headers
        )
        
        if response.status_code == 200:
            result = response.json()
            return JsonResponse({
                'success': True,
                'qr_code_url': result['data']['qr_code_url'],
                'qr_code_data': result['data']['qr_code_data']
            })
        
        # Fallback: Generate a simple payment QR code
        import qrcode
        from io import BytesIO
        import base64
        
        # Create a QR code with payment information
        payment_string = f"""
        PAYMENT DETAILS:
        Till: {settings.MPESA_TILL_NUMBER}
        Amount: KES {sale.total}
        Sale ID: {sale.id}
        Merchant: Succeed Cereal Hub
        """
        
        qr = qrcode.QRCode(version=1, box_size=10, border=5)
        qr.add_data(payment_string)
        qr.make(fit=True)
        
        img = qr.make_image(fill_color="black", back_color="white")
        
        # Convert to base64
        buffer = BytesIO()
        img.save(buffer, format='PNG')
        img_str = base64.b64encode(buffer.getvalue()).decode()
        
        return JsonResponse({
            'success': True,
            'qr_code_base64': img_str,
            'payment_info': payment_data
        })
        
    except Sale.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Sale not found'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


@csrf_exempt
def paystack_webhook(request):
    """
    Handle Paystack webhook callbacks
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    # Verify webhook signature
    signature = request.headers.get('x-paystack-signature')
    if not signature:
        return JsonResponse({'error': 'No signature'}, status=400)
    
    # Compute expected signature
    payload = request.body
    expected_signature = hmac.new(
        settings.PAYSTACK_SECRET_KEY.encode('utf-8'),
        payload,
        hashlib.sha512
    ).hexdigest()
    
    if signature != expected_signature:
        return JsonResponse({'error': 'Invalid signature'}, status=400)
    
    # Process webhook
    event = json.loads(payload)
    
    if event['event'] == 'charge.success':
        # Payment successful
        reference = event['data']['reference']
        amount = event['data']['amount'] / 100  # Convert from cents
        metadata = event['data']['metadata']
        sale_id = metadata.get('sale_id')
        
        if sale_id:
            try:
                sale = Sale.objects.get(id=sale_id)
                
                # Update sale payment status
                if not sale.is_completed:
                    # Mark as paid
                    sale.amount_paid = amount
                    sale.balance = max(sale.total - amount, 0)
                    sale.is_credit = sale.balance > 0
                    sale.save()
                
                return JsonResponse({'status': 'success'})
                
            except Sale.DoesNotExist:
                pass
    
    return JsonResponse({'status': 'received'})


def verify_payment(request):
    """
    Verify payment status
    """
    reference = request.GET.get('reference')
    if not reference:
        return JsonResponse({'error': 'No reference'}, status=400)
    
    headers = {
        'Authorization': f'Bearer {settings.PAYSTACK_SECRET_KEY}',
    }
    
    response = requests.get(
        f'https://api.paystack.co/transaction/verify/{reference}',
        headers=headers
    )
    
    if response.status_code == 200:
        result = response.json()
        return JsonResponse({
            'success': result['status'],
            'data': result['data']
        })
    
    return JsonResponse({
        'success': False,
        'message': 'Verification failed'
    }, status=400)


from django.shortcuts import render, redirect
from django.contrib import messages

def payment_callback(request):
    """
    Handle Paystack payment callback after customer completes payment
    """
    reference = request.GET.get('reference')
    trxref = request.GET.get('trxref')
    
    # Use reference or trxref (Paystack uses both sometimes)
    payment_reference = reference or trxref
    
    if not payment_reference:
        messages.error(request, 'No payment reference provided')
        return redirect('pos')
    
    # Verify the payment
    headers = {
        'Authorization': f'Bearer {settings.PAYSTACK_SECRET_KEY}',
    }
    
    try:
        response = requests.get(
            f'https://api.paystack.co/transaction/verify/{payment_reference}',
            headers=headers
        )
        
        if response.status_code == 200:
            result = response.json()
            
            if result['status'] and result['data']['status'] == 'success':
                # Payment was successful
                metadata = result['data'].get('metadata', {})
                sale_id = metadata.get('sale_id')
                
                if sale_id:
                    try:
                        sale = Sale.objects.get(id=sale_id)
                        # Mark payment as completed if needed
                        # (Your existing process_sale view already handles this)
                        
                        messages.success(request, 'Payment completed successfully!')
                        
                        # Render a success page or redirect to receipt
                        return render(request, 'pos/payment_success.html', {
                            'sale': sale,
                            'reference': payment_reference
                        })
                    except Sale.DoesNotExist:
                        messages.warning(request, 'Payment successful but sale record not found')
                else:
                    messages.success(request, 'Payment completed successfully!')
            else:
                messages.error(request, 'Payment verification failed')
        else:
            messages.error(request, 'Could not verify payment')
            
    except Exception as e:
        messages.error(request, f'Error verifying payment: {str(e)}')
    
    return redirect('pos')


    